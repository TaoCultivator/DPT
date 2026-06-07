"""验证 Excel 导出是否写入当前 result（含手动微调后的数值）。"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from PyQt6.QtWidgets import QApplication

from dpt_extractor.config.loader import load_config
from dpt_extractor.export.excel_export import export_to_excel
from dpt_extractor.export.mcu2506_layout import COL_ON, DATA_ROW
from dpt_extractor.gui.main_window import MainWindow
from dpt_extractor.io.tek_parser import TekParser
from dpt_extractor.models.bridge_profile import LOWER_BRIDGE
from dpt_extractor.pipeline.extract import extract_all


def main() -> int:
    csv = ROOT / "WL_480V_800A_000_ALL.csv"
    if not csv.is_file():
        print("WL CSV not found")
        return 1
    app = QApplication.instance() or QApplication(sys.argv)
    bundle = TekParser().parse(csv)
    win = MainWindow()
    win.bundle = bundle
    win.profile = LOWER_BRIDGE
    win.result = extract_all(bundle, LOWER_BRIDGE, load_config())
    win.cfg = load_config()
    win.result_table.set_result(win.result)
    win.wave_plot.plot_waveforms(bundle, LOWER_BRIDGE, win.result)

    before = float(win.result.turn_on.turn_on_current)
    marker = before + 50.0
    win.result.turn_on.turn_on_current = marker
    win.result_table.set_metric_value("开通", "开通电流", marker)

    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "verify.xlsx"
        export_to_excel(win.result, path)
        from openpyxl import load_workbook

        xls = float(
            load_workbook(path, data_only=True).active.cell(
                DATA_ROW, COL_ON["turn_on_current"]
            ).value
        )

    ok = abs(xls - marker) < 0.5
    print(f"before={before} manual={marker} excel={xls} OK={ok}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
