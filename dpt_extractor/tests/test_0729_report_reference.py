"""0729 report-backed cursor regressions for the re-imported KSU2577 batch."""

from __future__ import annotations

import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from dpt_extractor.config.loader import load_config
from dpt_extractor.io.waveform_loader import load_waveform
from dpt_extractor.metrics.iec_windows import (
    err_energy_markers,
    rr_completed_measurement_window_indices,
)
from dpt_extractor.models.bridge_profile import (
    guess_profile_from_path,
    make_profile,
)
from dpt_extractor.models.channel_mapping import (
    apply_mapping,
    infer_best_mapping_from_bundle,
)
from dpt_extractor.models.waveform import bundle_reverse_recovery_current
from dpt_extractor.pipeline.extract import extract_all


ROOT = Path(__file__).resolve().parents[2]
REFERENCE_ROOT = (
    ROOT
    / "示例文件"
    / "songzhenxi"
    / "KSU2577"
    / "07CF2C1000 20260717"
    / "SMC"
)
REPORT_0729 = next(
    iter(sorted(REFERENCE_ROOT.parent.glob("*20260729*.xlsx"))),
    None,
)

LT_REPORT_ERR_MJ = {
    "UH": 9.420,
    "UL": 6.534,
    "VH": 7.592,
    "VL": 6.874,
    "WH": 8.354,
    "WL": 13.105,
}

LT_REPORT_RR_DVDT_V_PER_NS = {
    "UH": 5.2769,
    "UL": 5.3617,
    "VH": 5.3307,
    "VL": 5.4807,
    "WH": 8.6487,
    "WL": 5.3973,
}

STABLE_PLATFORM_RR_DVDT_V_PER_NS = {
    ("HT", "UH"): 18.05629729392376,
    ("HT", "UL"): 16.43715401342251,
    ("HT", "VH"): 17.943293659263166,
    ("HT", "VL"): 16.721052212254328,
    ("HT", "WH"): 18.376296838315675,
    ("HT", "WL"): 16.85790246194822,
    ("RT", "UH"): 8.722280364658834,
    ("RT", "UL"): 9.20825974222445,
    ("RT", "VH"): 8.85922515237225,
    ("RT", "VL"): 9.361448143800159,
    ("RT", "WH"): 8.835369622310722,
    ("RT", "WL"): 9.355705848950942,
}

UNCHANGED_ERR_MJ = {
    ("HT", "UH"): 13.731502732671764,
    ("HT", "UL"): 19.993488195295914,
    ("HT", "VH"): 11.50478793001307,
    ("HT", "VL"): 19.19821183908417,
    ("HT", "WH"): 12.09336839009178,
    ("HT", "WL"): 19.902304197834884,
    ("RT", "UH"): 6.74651066446076,
    ("RT", "UL"): 9.293883724971456,
    ("RT", "VH"): 6.45866352055415,
    ("RT", "VL"): 10.96827992884032,
    ("RT", "WH"): 6.3178850247336396,
    ("RT", "WL"): 13.038302468178658,
}


@unittest.skipUnless(REFERENCE_ROOT.is_dir(), "0729 reference TSS batch missing")
class Test0729ReportReference(unittest.TestCase):
    @staticmethod
    def _load_case(temperature: str, code: str):
        path = REFERENCE_ROOT / temperature / f"{code}_750V_1048A_000.tss"
        bundle = load_waveform(path)
        guessed = guess_profile_from_path(path)
        base = make_profile(guessed.phase, guessed.bridge)
        mapping, _source = infer_best_mapping_from_bundle(
            bundle,
            guessed.bridge,
        )
        profile = apply_mapping(base, mapping) if mapping is not None else base
        result = extract_all(bundle, profile, load_config())
        return bundle, profile, result

    @staticmethod
    def _err_markers(bundle, profile, result):
        segments = result.segments
        assert segments is not None
        on0, on1 = segments.turn_on
        rr0, rr1 = segments.reverse_recovery
        v_diode = bundle.get(profile.v_diode)
        _rr_s0, rr_s1, completed = rr_completed_measurement_window_indices(
            on0,
            rr1,
            on1,
            v_diode,
            len(bundle.t),
            bundle.dt,
        )
        rr_context_i1 = rr_s1 if completed else rr1
        return err_energy_markers(
            bundle.t,
            bundle_reverse_recovery_current(bundle, profile),
            v_diode,
            rr0,
            rr_context_i1,
            bundle.dt,
            i_search_end=on1,
            vge=bundle.get(profile.vge),
            pulse1_off=segments.pulse1_off,
            pulse2_on=segments.pulse2_on,
            pulse2_off=segments.pulse2_off,
            dc_current=result.idc,
            lower_bridge_irr_from_ic_minus_il=profile.irr_from_ic_minus_il,
        )

    def test_lt_err_waits_for_the_visible_recovery_tail(self) -> None:
        """LT report images put A after the compact segment, on the settled tail."""
        for code, report_err in LT_REPORT_ERR_MJ.items():
            with self.subTest(code=code):
                bundle, profile, result = self._load_case("LT", code)
                segments = result.segments
                assert segments is not None
                markers = self._err_markers(bundle, profile, result)
                on1_s = float(bundle.t[segments.turn_on[1]])

                self.assertGreaterEqual(markers.t_start, on1_s + 50e-9)
                self.assertLessEqual(markers.t_start, on1_s + 350e-9)

                irr = bundle_reverse_recovery_current(bundle, profile)
                irr_at_a = float(np.interp(markers.t_start, bundle.t, irr))
                self.assertAlmostEqual(irr_at_a, markers.ha_v, places=6)

                tolerance_mj = max(0.75, 0.20 * report_err)
                self.assertLessEqual(
                    abs(result.reverse_recovery.err - report_err),
                    tolerance_mj,
                    (
                        f"{code} Err={result.reverse_recovery.err:.6f}mJ "
                        f"deviates from 0729 report {report_err:.3f}mJ"
                    ),
                )

    def test_ht_and_rt_err_values_do_not_drift(self) -> None:
        """The LT tail repair must not change already-aligned HT/RT extraction."""
        for (temperature, code), expected in UNCHANGED_ERR_MJ.items():
            with self.subTest(temperature=temperature, code=code):
                _bundle, _profile, result = self._load_case(temperature, code)
                self.assertAlmostEqual(
                    result.reverse_recovery.err,
                    expected,
                    places=9,
                )

    def test_lt_rr_dvdt_rejects_ringing_peak_as_the_voltage_top(self) -> None:
        """Low-Irr LT records use their settled Vd platform, not ringing overshoot."""
        for code, report_dvdt in LT_REPORT_RR_DVDT_V_PER_NS.items():
            with self.subTest(code=code):
                _bundle, _profile, result = self._load_case("LT", code)
                self.assertAlmostEqual(
                    result.reverse_recovery.dvdt_max,
                    report_dvdt,
                    delta=0.02 * report_dvdt,
                )

    def test_ht_and_rt_rr_dvdt_use_stable_blocking_platform(self) -> None:
        """HT/RT slopes use stable high/low platforms, not Vd overshoot/zero."""
        for (temperature, code), expected in STABLE_PLATFORM_RR_DVDT_V_PER_NS.items():
            with self.subTest(temperature=temperature, code=code):
                _bundle, _profile, result = self._load_case(temperature, code)
                self.assertAlmostEqual(
                    result.reverse_recovery.dvdt_max,
                    expected,
                    delta=1e-4,
                )

    @unittest.skipUnless(REPORT_0729 is not None, "0729 reference report missing")
    def test_all_558_report_numeric_slots_have_finite_production_results(self) -> None:
        """Every numeric report slot has a corresponding extracted value."""
        assert REPORT_0729 is not None
        workbook = load_workbook(REPORT_0729, read_only=True, data_only=True)
        try:
            report_numeric_count = sum(
                1
                for sheet_index in (6, 8, 10)
                for row in (5, 9, 13, 17, 21, 25)
                for col in range(6, 40)
                if isinstance(
                    workbook.worksheets[sheet_index].cell(row, col).value,
                    (int, float),
                )
            )
        finally:
            workbook.close()
        self.assertEqual(report_numeric_count, 558)

        production_values: list[float] = []
        for temperature in ("RT", "HT", "LT"):
            for code in ("UH", "UL", "VH", "VL", "WH", "WL"):
                with self.subTest(temperature=temperature, code=code):
                    _bundle, _profile, result = self._load_case(temperature, code)
                    off = result.turn_off
                    on = result.turn_on
                    rr = result.reverse_recovery
                    values = (
                        off.delta_vce,
                        off.ic_off_max,
                        off.vce_off_max,
                        off.dvdt,
                        off.didt,
                        off.ls_off,
                        off.toff,
                        off.td_off,
                        off.tf,
                        off.pmax,
                        off.eoff,
                        on.delta_vce,
                        on.ic_on_max,
                        on.vce_on_max,
                        on.turn_on_current,
                        on.dvdt,
                        on.didt,
                        on.ls_on,
                        on.ton,
                        on.td_on,
                        on.tr,
                        on.pmax,
                        on.eon,
                        rr.irr,
                        rr.trr,
                        rr.vrr,
                        rr.dvdt_max,
                        rr.didt_irr,
                        rr.pdmax,
                        rr.err,
                        off.eoff + on.eon + rr.err,
                    )
                    self.assertTrue(all(np.isfinite(value) for value in values))
                    production_values.extend(float(value) for value in values)
        self.assertEqual(len(production_values), 558)


if __name__ == "__main__":
    unittest.main()
