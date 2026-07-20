from __future__ import annotations

import base64
from pathlib import Path
import shutil
import tempfile
import unittest
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook


_PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMB"
    "/6X7W1sAAAAASUVORK5CYII="
)


def _write_tiny_png(path: Path) -> None:
    path.write_bytes(_PNG_1X1)


def _write_rgb_png(path: Path, size: tuple[int, int]) -> None:
    from PIL import Image

    Image.new("RGB", size, (16, 24, 32)).save(path)


def _drawing_anchor_counts(path: Path) -> dict[str, int]:
    tags = {
        "oneCellAnchor": "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor",
        "twoCellAnchor": "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor",
    }
    counts = {key: 0 for key in tags}
    with ZipFile(path) as zf:
        for name in zf.namelist():
            if not name.startswith("xl/drawings/drawing") or not name.endswith(".xml"):
                continue
            root = ET.fromstring(zf.read(name))
            for key, tag in tags.items():
                counts[key] += len(root.findall(f".//{tag}"))
    return counts


class TestReportTemplateWriter(unittest.TestCase):
    def test_report_condition_overrides_are_validated_before_path_fallback(self):
        from dpt_extractor.export.report_template import (
            _resolve_phase_code,
            _resolve_temp_code,
        )
        from dpt_extractor.models.results import ExtractResult

        ambiguous = str(Path("colleague") / "U_L" / "UH_750V_1048A_000.tss")
        explicit_lt = str(Path("colleague") / "LT" / "UH_750V_1048A_000.tss")
        self.assertEqual(_resolve_temp_code(ambiguous), "RT")
        self.assertEqual(_resolve_temp_code(explicit_lt), "LT")
        self.assertEqual(_resolve_temp_code(ambiguous, "lt"), "LT")
        with self.assertRaisesRegex(ValueError, "不支持的温度工况代码"):
            _resolve_temp_code(ambiguous, "LOW")
        result = ExtractResult(profile_code="UL")
        self.assertEqual(_resolve_phase_code(ambiguous, result), "UH")
        self.assertEqual(_resolve_phase_code(ambiguous, result, "vl"), "VL")
        with self.assertRaisesRegex(ValueError, "不支持的相位/桥臂代码"):
            _resolve_phase_code(ambiguous, result, "U")

    def test_dpt_missing_explicit_condition_never_falls_back_to_row_five(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "RT"
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "缺少工况组"):
                write_report_template(
                    ExtractResult(source_path="ambiguous.tss", profile_code="UL"),
                    report,
                    phase_code="UL",
                    temperature_code="LT",
                )

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(saved["A5"].value, "UH")
            self.assertEqual(saved["B5"].value, "RT")

    def test_short_missing_explicit_condition_never_relabels_row_five(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws["A5"] = "25℃"
            ws["B5"] = "UH"
            wb.save(report)

            result = ExtractResult(
                source_path="ambiguous.tss",
                profile_code="UL",
                short_circuit_mode=True,
            )
            with self.assertRaisesRegex(ValueError, "缺少短路工况行"):
                write_report_template(
                    result,
                    report,
                    phase_code="UL",
                    temperature_code="LT",
                )

            saved = load_workbook(report)["短路测试"]
            self.assertEqual(saved["A5"].value, "25℃")
            self.assertEqual(saved["B5"].value, "UH")

    def test_dpt_nonempty_images_require_waveform_sheet(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "缺少工作表：U相_双脉冲波形"):
                write_report_template(
                    ExtractResult(
                        source_path="UH_750V_1048A_000.tss",
                        profile_code="UH",
                    ),
                    report,
                    images={("关断过程", "ΔVce"): image},
                    phase_code="UH",
                    temperature_code="RT",
                )

    def test_short_nonempty_images_require_picture_sheet(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws["A5"] = "25℃"
            ws["B5"] = "UH"
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "缺少工作表：短路测试图片"):
                write_report_template(
                    ExtractResult(
                        source_path="UH_750V_1048A_000.tss",
                        profile_code="UH",
                        short_circuit_mode=True,
                    ),
                    report,
                    images={("短路过程", "短路电流Imax"): image},
                    phase_code="UH",
                    temperature_code="RT",
                )

    def test_unknown_image_key_does_not_require_an_image_sheet(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path="UH_750V_1048A_000.tss",
                    profile_code="UH",
                ),
                report,
                images={("未知过程", "未知参数"): Path(td) / "unused.png"},
                phase_code="UH",
                temperature_code="RT",
            )

            self.assertEqual(summary.images_written, 0)
            self.assertIsNone(summary.waveform_sheet)

    def test_short_nonempty_images_require_matching_picture_condition(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws["A5"] = "25℃"
            ws["B5"] = "UH"
            pic = wb.create_sheet("短路测试图片")
            pic.merge_cells("A1:E1")
            pic["A1"] = "UL_25℃"
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "缺少图片工况块：UH_25℃"):
                write_report_template(
                    ExtractResult(
                        source_path="UH_750V_1048A_000.tss",
                        profile_code="UH",
                        short_circuit_mode=True,
                    ),
                    report,
                    images={("短路过程", "短路电流Imax"): image},
                    phase_code="UH",
                    temperature_code="RT",
                )

    def test_dpt_nonempty_images_reject_missing_parameter_slot(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wb.create_sheet("U相_双脉冲波形")
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "双脉冲报告图片写入不完整"):
                write_report_template(
                    ExtractResult(
                        source_path="UH_750V_1048A_000.tss",
                        profile_code="UH",
                    ),
                    report,
                    images={("关断过程", "ΔVce"): image},
                    phase_code="UH",
                    temperature_code="RT",
                )

    def test_short_nonempty_images_reject_missing_parameter_slot(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws["A5"] = "25℃"
            ws["B5"] = "UH"
            pic = wb.create_sheet("短路测试图片")
            pic.merge_cells("A1:E1")
            pic["A1"] = "UH_25℃"
            wb.save(report)

            with self.assertRaisesRegex(ValueError, "短路报告图片写入不完整"):
                write_report_template(
                    ExtractResult(
                        source_path="UH_750V_1048A_000.tss",
                        profile_code="UH",
                        short_circuit_mode=True,
                    ),
                    report,
                    images={("短路过程", "短路电流Imax"): image},
                    phase_code="UH",
                    temperature_code="RT",
                )

    def test_inserted_waveform_image_keeps_original_aspect_ratio(self):
        from openpyxl.worksheet.cell_range import CellRange

        from dpt_extractor.export.report_template import _insert_image

        with tempfile.TemporaryDirectory() as td:
            image = Path(td) / "scope_16x9.png"
            _write_rgb_png(image, (160, 90))

            wb = Workbook()
            ws = wb.active
            ws.merge_cells("J2:Q17")
            _insert_image(ws, image, CellRange("J2:Q17"))

            inserted = ws._images[0]
            self.assertAlmostEqual(inserted.width / inserted.height, 160 / 90, delta=0.02)

    def test_insert_image_rejects_missing_file_instead_of_counting_blank_slot(self):
        from openpyxl.worksheet.cell_range import CellRange

        from dpt_extractor.export.report_template import _insert_image

        with tempfile.TemporaryDirectory() as td:
            missing = Path(td) / "missing_scope.png"
            wb = Workbook()
            ws = wb.active
            ws.merge_cells("J2:Q17")

            with self.assertRaisesRegex(FileNotFoundError, "报告图片不存在"):
                _insert_image(ws, missing, CellRange("J2:Q17"))

            self.assertEqual(ws._images, [])

    def test_dpt_template_writes_next_empty_row_and_fills_waveform_labels(self):
        from dpt_extractor.export.mcu2506_layout import COL_OFF
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in ((1, "750V_1050A"), (54, "750V_805A")):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(start_row=base, start_column=10, end_row=base, end_column=17)
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(start_row=base + 1, start_column=10, end_row=base + 16, end_column=17)
                wave.cell(base + 1, 10, 1)
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "RT" / "UH_750V_805A_000.tss"),
                profile_code="UH",
                turn_off=TurnOffResult(delta_vce=123.456),
            )
            progress_events: list[tuple[int, int, str]] = []
            summary = write_report_template(
                result,
                report,
                images={
                    ("关断过程", "ΔVce"): image,
                    ("不存在的分区", "不存在的参数"): image,
                },
                progress_callback=lambda done, total, label: progress_events.append(
                    (done, total, label)
                ),
            )

            saved = load_workbook(report)
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(summary.data_sheet, "U相_双脉冲数据")
            self.assertEqual(summary.data_row, 5)
            self.assertEqual(summary.waveform_anchor_row, 1)
            self.assertEqual(saved["U相_双脉冲数据"].cell(5, COL_OFF["delta_vce"]).value, 123.456)
            self.assertEqual(saved_wave.cell(1, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(18, 1).value, "750V_805A")
            self.assertEqual(saved_wave.cell(35, 1).value, "总概览图")
            self.assertLess(saved_wave.column_dimensions["J"].width, 13)
            self.assertEqual(len(saved_wave._images), 1)
            self.assertEqual(
                [event for event in progress_events if event[2] == "插入报告图片"],
                [(1, 1, "插入报告图片")],
            )

    def test_dpt_template_uses_current_temperature_label(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "RT"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_1050A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                ),
                report,
                temperature_labels={"RT": "26.5℃", "HT": "175℃", "LT": "-55℃"},
            )

            saved = load_workbook(report)
            self.assertEqual(saved["U相_双脉冲数据"].cell(5, 2).value, "26.5℃")
            self.assertEqual(saved["U相_双脉冲波形"].cell(1, 1).value, "UH_26.5℃")

    def test_dpt_same_code_keeps_distinct_actual_temperatures(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.mcu2506_layout import COL_OFF
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "preserved.png"
            _write_tiny_png(image)
            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            for start_row, temperature in ((5, "25℃"), (9, "150℃")):
                data.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row + 3,
                    end_column=1,
                )
                data.merge_cells(
                    start_row=start_row,
                    start_column=2,
                    end_row=start_row + 3,
                    end_column=2,
                )
                data.cell(start_row, 1, "UH")
                data.cell(start_row, 2, temperature)
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells("A1:H17")
            wave["A1"] = "UH_150℃"
            wave.merge_cells("A18:H34")
            wave["A18"] = "750V_1050A"
            wave.merge_cells("A35:H51")
            wave["A35"] = "总概览图"
            preserved = wb.create_sheet("保留内容")
            preserved["A1"] = "=1+1"
            preserved.add_image(XLImage(str(image)), "B2")
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "UH_750V_1050A_000.tss"),
                profile_code="UH",
                turn_off=TurnOffResult(eoff=1.25),
            )
            first = write_report_template(
                result,
                report,
                temperature_labels={"RT": "25℃", "HT": "150℃", "LT": "-40℃"},
                temperature_code="HT",
                phase_code="UH",
            )
            result.turn_off.eoff = 2.5
            second = write_report_template(
                result,
                report,
                temperature_labels={"RT": "25℃", "HT": "170℃", "LT": "-40℃"},
                temperature_code="HT",
                phase_code="UH",
            )

            saved = load_workbook(report, data_only=False)
            self.assertEqual(first.data_row, 9)
            self.assertEqual(second.data_row, 13)
            self.assertEqual(
                saved.sheetnames,
                ["U相_双脉冲数据", "U相_双脉冲波形", "保留内容"],
            )
            self.assertEqual(saved["U相_双脉冲数据"]["B5"].value, "25℃")
            self.assertEqual(saved["U相_双脉冲数据"]["B9"].value, "150℃")
            self.assertEqual(saved["U相_双脉冲数据"]["B13"].value, "170℃")
            self.assertEqual(saved["U相_双脉冲波形"]["A1"].value, "UH_150℃")
            self.assertEqual(saved["U相_双脉冲波形"]["A54"].value, "UH_170℃")
            self.assertIsNone(saved["U相_双脉冲数据"].cell(5, COL_OFF["eoff"]).value)
            self.assertEqual(
                saved["U相_双脉冲数据"].cell(9, COL_OFF["eoff"]).value,
                1.25,
            )
            self.assertEqual(
                saved["U相_双脉冲数据"].cell(13, COL_OFF["eoff"]).value,
                2.5,
            )
            self.assertEqual(saved["保留内容"]["A1"].value, "=1+1")
            self.assertEqual(len(saved["保留内容"]._images), 1)
            anchor = saved["保留内容"]._images[0].anchor._from
            self.assertEqual((anchor.row, anchor.col), (1, 1))
            for code in ("RT", "HT", "LT"):
                identity = saved.defined_names[f"_DPT_TEMP_IDENTITY_{code}"]
                self.assertTrue(identity.hidden)

    def test_default_template_retains_150_and_170_data_and_images(self):
        from dpt_extractor.export.report_template import (
            DPT_OVERVIEW_IMAGE_PARAM,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult, TurnOffResult
        from dpt_extractor.utils.app_paths import default_report_template_path

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "high_temperature_retention.xlsx"
            image = td_path / "overview.png"
            _write_rgb_png(image, (1280, 960))
            shutil.copyfile(default_report_template_path(), report)

            rows = []
            for temperature, eoff in ((150, 1.5), (170, 1.7)):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(
                            Path("samples") / "UH_900V_500A_000.tss"
                        ),
                        profile_code="UH",
                        turn_off=TurnOffResult(eoff=eoff),
                    ),
                    report,
                    images={DPT_OVERVIEW_IMAGE_PARAM: image},
                    temperature_labels={
                        "RT": "25℃",
                        "HT": f"{temperature}℃",
                        "LT": "-40℃",
                    },
                    temperature_code="HT",
                    phase_code="UH",
                )
                rows.append(summary.data_row)

            saved = load_workbook(report)
            data = saved["U相_双脉冲数据"]
            wave = saved["U相_双脉冲波形"]
            self.assertEqual(rows, [13, 17])
            self.assertEqual(data["B13"].value, "150℃")
            self.assertEqual(data["B17"].value, "170℃")
            self.assertEqual(wave["A1"].value, "UH_150℃")
            self.assertEqual(wave["A54"].value, "UH_170℃")
            self.assertEqual(len(wave._images), 2)
            self.assertEqual(
                sorted(item.anchor._from.row + 1 for item in wave._images),
                [35, 88],
            )

    def test_dpt_manual_page_conditions_override_ambiguous_path(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_CURRENT,
            COL_OFF,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            groups = (
                (5, "UH", "RT"),
                (9, "UL", "RT"),
                (13, "UH", "LT"),
                (17, "UL", "LT"),
            )
            for start_row, phase, temp in groups:
                data.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row + 3,
                    end_column=1,
                )
                data.merge_cells(
                    start_row=start_row,
                    start_column=2,
                    end_row=start_row + 3,
                    end_column=2,
                )
                data.cell(start_row, 1, phase)
                data.cell(start_row, 2, temp)
                data.cell(start_row, COL_VOLTAGE, 750)
                data.cell(start_row, COL_CURRENT, 1048)

            wave = wb.create_sheet("U相_双脉冲波形")
            for index, (_start_row, phase, temp) in enumerate(groups):
                base = 1 + index * 53
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, f"{phase}_{temp}")
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, "750V_1048A")
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(
                        Path("colleague")
                        / "U_L"
                        / "UH_750V_1048A_000.tss"
                    ),
                    profile_code="UH",
                    vdc_set=750.0,
                    idc_set=1048.0,
                    turn_off=TurnOffResult(
                        delta_vce=321.0,
                        ic_off_max=1048.0,
                    ),
                ),
                report,
                temperature_code="LT",
                phase_code="UL",
            )

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(summary.data_row, 17)
            self.assertEqual(summary.waveform_anchor_row, 160)
            self.assertIsNone(saved_data.cell(5, COL_OFF["delta_vce"]).value)
            self.assertIsNone(saved_data.cell(13, COL_OFF["delta_vce"]).value)
            self.assertEqual(saved_data.cell(17, COL_OFF["delta_vce"]).value, 321)
            self.assertEqual(saved_data.cell(17, 2).value, "-40℃")
            self.assertEqual(saved_wave.cell(160, 1).value, "UL_-40℃")

    def test_dpt_multi_pulse_images_follow_current_page_result(self):
        from openpyxl.drawing.image import Image as OpenpyxlImage

        from dpt_extractor.export.report_template import (
            DPT_OVERVIEW_IMAGE_PARAM,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "overview.png"
            _write_tiny_png(image)

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            data.merge_cells("A5:A8")
            data.merge_cells("B5:B8")
            data["A5"] = "UH"
            data["B5"] = "RT"
            data["D5"] = 750
            data["E5"] = 100
            data["D6"] = 750
            data["E6"] = 200
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in ((1, "750V_100A"), (54, "750V_200A")):
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
            old_first = OpenpyxlImage(image)
            old_first.anchor = "A35"
            wave.add_image(old_first)
            old_second = OpenpyxlImage(image)
            old_second.anchor = "A88"
            wave.add_image(old_second)
            wb.save(report)

            source = str(Path("samples") / "RT" / "UH_capture.tss")
            rows = [
                ExtractResult(
                    source_path=source,
                    profile_code="UH",
                    vdc_set=750.0,
                    idc_set=100.0,
                    off_pulse_index=1,
                    on_pulse_index=2,
                    turn_off=TurnOffResult(delta_vce=101.0),
                ),
                ExtractResult(
                    source_path=source,
                    profile_code="UH",
                    vdc_set=750.0,
                    idc_set=300.0,
                    off_pulse_index=2,
                    on_pulse_index=3,
                    turn_off=TurnOffResult(delta_vce=202.0),
                ),
                ExtractResult(
                    source_path=source,
                    profile_code="UH",
                    vdc_set=750.0,
                    idc_set=400.0,
                    off_pulse_index=3,
                    on_pulse_index=4,
                    turn_off=TurnOffResult(delta_vce=303.0),
                ),
            ]
            summary = write_report_template(
                rows,
                report,
                images={DPT_OVERVIEW_IMAGE_PARAM: image},
                temperature_code="RT",
                phase_code="UH",
                image_result_index=2,
            )

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(summary.waveform_anchor_row, 107)
            self.assertEqual(
                [saved_data.cell(row, 5).value for row in range(5, 9)],
                [100, 300, 400, 200],
            )
            self.assertEqual(saved_wave.cell(18, 1).value, "750V_100A")
            self.assertEqual(saved_wave.cell(71, 1).value, "750V_300A")
            self.assertEqual(saved_wave.cell(124, 1).value, "750V_400A")
            self.assertEqual(saved_wave.cell(177, 1).value, "750V_200A")
            self.assertEqual(len(saved_wave._images), 3)
            self.assertEqual(
                sorted(item.anchor._from.row + 1 for item in saved_wave._images),
                [35, 141, 194],
            )

    def test_dpt_write_normalizes_short_picture_temperature_labels(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_1050A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            short_pic = wb.create_sheet("短路测试图片")
            short_pic.merge_cells("A2:E5")
            short_pic["A2"] = "UH_RT"
            short_pic.merge_cells("A23:E26")
            short_pic["A23"] = "UH_HT"
            short_pic.merge_cells("A44:E47")
            short_pic["A44"] = "UH_LT"
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                ),
                report,
            )

            saved_short = load_workbook(report)["短路测试图片"]
            self.assertEqual(saved_short["A2"].value, "UH_25℃")
            self.assertEqual(saved_short["A23"].value, "UH_150℃")
            self.assertEqual(saved_short["A44"].value, "UH_-40℃")

    def test_dpt_template_leaves_unavailable_metric_cells_blank(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_CURRENT,
            COL_OFF,
            COL_ON,
            COL_RR,
            COL_TAIL,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import (
            ExtractResult,
            ReverseRecoveryResult,
            TurnOffResult,
            TurnOnResult,
        )

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1050)
            ws.cell(5, COL_OFF["crosstalk"], "stale")
            ws.cell(5, COL_RR["err"], 999)
            ws.cell(5, COL_TAIL["etotal"], 999)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(
                        crosstalk_vmax=12.3,
                        crosstalk_vmin=-4.5,
                        eoff=10.0,
                    ),
                    turn_on=TurnOnResult(eon=20.0),
                    reverse_recovery=ReverseRecoveryResult(err=30.0),
                    unavailable_metrics={
                        ("关断过程", "串扰电压"),
                        ("反向恢复", "Err"),
                    },
                ),
                report,
            )

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertIsNone(saved.cell(5, COL_OFF["crosstalk"]).value)
            self.assertIsNone(saved.cell(5, COL_RR["err"]).value)
            self.assertIsNone(saved.cell(5, COL_TAIL["etotal"]).value)
            self.assertEqual(saved.cell(5, COL_OFF["eoff"]).value, 10)
            self.assertEqual(saved.cell(5, COL_ON["eon"]).value, 20)

    def test_excel_layout_leaves_unavailable_metric_cells_blank(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_OFF,
            COL_ON,
            COL_RR,
            COL_TAIL,
            DATA_ROW,
            build_mcu2506_workbook,
            fill_data_row,
        )
        from dpt_extractor.models.results import (
            ExtractResult,
            ReverseRecoveryResult,
            TurnOffResult,
            TurnOnResult,
        )

        result = ExtractResult(
            source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
            profile_code="UH",
            turn_off=TurnOffResult(eoff=10.0),
            turn_on=TurnOnResult(eon=20.0),
            reverse_recovery=ReverseRecoveryResult(vrr=900.0, err=30.0),
            unavailable_metrics={
                ("开通", "Eon"),
                ("反向恢复", "Vrr"),
            },
        )
        wb = build_mcu2506_workbook(result)
        ws = wb.active
        ws.cell(DATA_ROW, COL_ON["eon"], 999)
        ws.cell(DATA_ROW, COL_RR["vrr"], 999)
        ws.cell(DATA_ROW, COL_TAIL["etotal"], 999)

        fill_data_row(ws, DATA_ROW, result)

        self.assertEqual(ws.cell(DATA_ROW, COL_OFF["eoff"]).value, 10)
        self.assertIsNone(ws.cell(DATA_ROW, COL_ON["eon"]).value)
        self.assertIsNone(ws.cell(DATA_ROW, COL_RR["vrr"]).value)
        self.assertEqual(ws.cell(DATA_ROW, COL_RR["err"]).value, 30)
        self.assertIsNone(ws.cell(DATA_ROW, COL_TAIL["etotal"]).value)

    def test_dpt_old_template_is_upgraded_with_power_columns(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_OFF,
            COL_ON,
            COL_RR,
            DATA_ROW,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import (
            ExtractResult,
            ReverseRecoveryResult,
            TurnOffResult,
            TurnOnResult,
        )

        old_headers = [
            "测试相",
            "Temp",
            "测试条件",
            "Recorded Voltage",
            "Recorded Current",
            "△Vce",
            "Ic_off_max",
            "Vce_off_max",
            "dv/dt",
            "di/dt",
            "Ls_off",
            "Toff",
            "Td_off",
            "Tf",
            "串扰电压",
            "Eoff",
            "△Vce",
            "Ic_on_max",
            "Vce_on_max",
            "开通电流",
            "dv/dt",
            "di/dt",
            "Ls_on",
            "Ton",
            "Td_on",
            "Tr",
            "串扰电压",
            "Eon",
            "Irr",
            "Trr",
            "Vrr",
            "Dvdt_max",
            "Didt_Irr",
            "Err",
            "Deadtime",
            "Etotal（all）",
            "Uaveform",
        ]
        old_units = [
            "",
            "°C",
            "",
            "V",
            "A",
            "V",
            "A",
            "V",
            "V/ns",
            "A/ns",
            "nH",
            "ns",
            "ns",
            "ns",
            "V",
            "mJ",
            "V",
            "A",
            "V",
            "A",
            "V/ns",
            "A/ns",
            "nH",
            "ns",
            "ns",
            "ns",
            "V",
            "mJ",
            "A",
            "ns",
            "V",
            "V/ns",
            "A/ns",
            "mJ",
            "ns",
            "mJ",
            "Picture number",
        ]

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "old_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            for col, (header, unit) in enumerate(zip(old_headers, old_units), start=1):
                ws.cell(3, col, header)
                ws.cell(4, col, unit)
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(DATA_ROW, 4, 750)
            ws.cell(DATA_ROW, 5, 1050)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(pmax=801.25, eoff=10.0),
                    turn_on=TurnOnResult(pmax=650.5, eon=20.0),
                    reverse_recovery=ReverseRecoveryResult(pdmax=120.75, err=30.0),
                ),
                report,
            )

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(saved.cell(3, COL_OFF["pmax"]).value, "Pmax")
            self.assertEqual(saved.cell(4, COL_OFF["pmax"]).value, "KW")
            self.assertEqual(saved.cell(3, COL_ON["pmax"]).value, "Pmax")
            self.assertEqual(saved.cell(3, COL_RR["pdmax"]).value, "Pdmax")
            self.assertEqual(saved.cell(DATA_ROW, COL_OFF["pmax"]).value, 801.25)
            self.assertEqual(saved.cell(DATA_ROW, COL_OFF["eoff"]).value, 10)
            self.assertEqual(saved.cell(DATA_ROW, COL_ON["pmax"]).value, 650.5)
            self.assertEqual(saved.cell(DATA_ROW, COL_ON["eon"]).value, 20)
            self.assertEqual(saved.cell(DATA_ROW, COL_RR["pdmax"]).value, 120.75)
            self.assertEqual(saved.cell(DATA_ROW, COL_RR["err"]).value, 30)

    def test_short_template_leaves_unavailable_metric_cells_blank(self):
        from dpt_extractor.export.short_circuit_layout import COL_ESC_OTHER, COL_VPEAK_OTHER
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "short_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws.merge_cells("A5:A6")
            ws["A5"] = 25
            ws["B5"] = "UH"
            ws["B6"] = "UL"
            ws.cell(5, COL_ESC_OTHER, 999)
            ws.cell(5, COL_VPEAK_OTHER, 999)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
                    profile_code="UH",
                    short_circuit_mode=True,
                    short_circuit=ShortCircuitResult(
                        esc_other=12.3456,
                        vpeak_other=789.123,
                    ),
                    unavailable_metrics={("短路过程", "短路能量Esc_对管")},
                ),
                report,
            )

            saved = load_workbook(report)["短路测试"]
            self.assertIsNone(saved.cell(5, COL_ESC_OTHER).value)
            self.assertEqual(saved.cell(5, COL_VPEAK_OTHER).value, 789.123)

    def test_dpt_missing_images_clear_only_active_condition_slots(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "dpt_stale_images.xlsx"
            image = td_path / "old_scope.png"
            _write_tiny_png(image)

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            data.merge_cells("A5:A8")
            data.merge_cells("B5:B8")
            data["A5"] = "UH"
            data["B5"] = "25℃"
            data.cell(5, COL_VOLTAGE, 750)
            data.cell(5, COL_CURRENT, 1050)
            data.cell(5, COL_OFF["delta_vce"], 999)
            data.cell(6, COL_VOLTAGE, 750)
            data.cell(6, COL_CURRENT, 805)

            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in ((1, "750V_1050A"), (54, "750V_805A")):
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(
                    start_row=base,
                    start_column=10,
                    end_row=base,
                    end_column=17,
                )
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(
                    start_row=base + 1,
                    start_column=10,
                    end_row=base + 16,
                    end_column=17,
                )
            for anchor in ("A35", "J2", "A88", "J55"):
                old_image = XLImage(str(image))
                old_image.anchor = anchor
                wave.add_image(old_image)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=321.0),
                    unavailable_metrics={("关断过程", "ΔVce")},
                ),
                report,
            )

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(summary.waveform_anchor_row, 1)
            self.assertIsNone(saved_data.cell(5, COL_OFF["delta_vce"]).value)
            self.assertEqual(len(saved_wave._images), 2)
            self.assertEqual(
                sorted(
                    (item.anchor._from.row + 1, item.anchor._from.col + 1)
                    for item in saved_wave._images
                ),
                [(55, 10), (88, 1)],
            )

    def test_short_missing_images_clear_only_active_condition_slots(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.export.short_circuit_layout import COL_ICMAX
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "short_stale_images.xlsx"
            image = td_path / "old_scope.png"
            _write_tiny_png(image)

            wb = Workbook()
            data = wb.active
            data.title = "短路测试"
            data.merge_cells("A5:A6")
            data["A5"] = "25℃"
            data["B5"] = "UH"
            data["B6"] = "UL"
            data.cell(5, COL_ICMAX, 999)
            data.cell(6, COL_ICMAX, 2222)

            picture = wb.create_sheet("短路测试图片")
            for base, phase_temp in ((1, "UH_RT"), (12, "UL_RT")):
                picture.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 3,
                    end_column=5,
                )
                picture.cell(base, 1, phase_temp)
                picture.merge_cells(
                    start_row=base,
                    start_column=6,
                    end_row=base,
                    end_column=8,
                )
                picture.cell(base, 6, "短路电流Imax")
                picture.merge_cells(
                    start_row=base + 1,
                    start_column=6,
                    end_row=base + 9,
                    end_column=8,
                )
            for anchor in ("F2", "F13"):
                old_image = XLImage(str(image))
                old_image.anchor = anchor
                picture.add_image(old_image)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
                    profile_code="UH",
                    short_circuit_mode=True,
                    short_circuit=ShortCircuitResult(ic_max=1234.0),
                    unavailable_metrics={("短路过程", "短路电流Imax")},
                ),
                report,
            )

            saved = load_workbook(report)
            saved_data = saved["短路测试"]
            saved_picture = saved["短路测试图片"]
            self.assertEqual(summary.waveform_anchor_row, 1)
            self.assertIsNone(saved_data.cell(5, COL_ICMAX).value)
            self.assertEqual(saved_data.cell(6, COL_ICMAX).value, 2222)
            self.assertEqual(len(saved_picture._images), 1)
            anchor = saved_picture._images[0].anchor._from
            self.assertEqual((anchor.row + 1, anchor.col + 1), (13, 6))

    def test_short_layout_leaves_unavailable_metric_cells_blank(self):
        from dpt_extractor.export.short_circuit_layout import (
            COL_ESC_OTHER,
            COL_VPEAK_OTHER,
            DATA_START_ROW,
            build_short_circuit_workbook,
            fill_short_circuit_row,
        )
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        result = ExtractResult(
            source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
            profile_code="UH",
            short_circuit_mode=True,
            short_circuit=ShortCircuitResult(
                esc_other=12.3456,
                vpeak_other=789.123,
            ),
            unavailable_metrics={("短路过程", "短路能量Esc_对管")},
        )
        wb = build_short_circuit_workbook(result)
        ws = wb.active
        ws.cell(DATA_START_ROW, COL_ESC_OTHER, 999)
        ws.cell(DATA_START_ROW, COL_VPEAK_OTHER, 999)

        fill_short_circuit_row(ws, DATA_START_ROW, result)

        self.assertIsNone(ws.cell(DATA_START_ROW, COL_ESC_OTHER).value)
        self.assertEqual(ws.cell(DATA_START_ROW, COL_VPEAK_OTHER).value, 789.123)

    def test_dpt_template_rewrites_same_condition_in_first_waveform_block(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1048)
            ws.cell(6, COL_VOLTAGE, 750)
            ws.cell(6, COL_CURRENT, 806)
            wave = wb.create_sheet("U相_双脉冲波形")
            for idx in range(2):
                base = 1 + idx * 53
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.merge_cells(start_row=base, start_column=10, end_row=base, end_column=17)
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(start_row=base + 1, start_column=10, end_row=base + 16, end_column=17)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1048A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=1048.0),
                ),
                report,
            )

            saved = load_workbook(report)
            self.assertEqual(summary.data_row, 5)
            self.assertEqual(summary.waveform_anchor_row, 1)
            self.assertEqual(saved["U相_双脉冲数据"].cell(5, COL_OFF["delta_vce"]).value, 1048)
            self.assertEqual(saved["U相_双脉冲波形"].cell(1, 1).value, "UH_25℃")
            self.assertEqual(saved["U相_双脉冲波形"].cell(18, 1).value, "750V_1048A")

    def test_dpt_template_clears_stale_duplicate_waveform_block(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A7")
            ws.merge_cells("B5:B7")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1048)
            ws.cell(6, COL_VOLTAGE, 750)
            ws.cell(6, COL_CURRENT, 806)
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in (
                (1, "750V_1048A"),
                (54, "750V_806A"),
                (107, "750V_806A"),
            ):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(start_row=base, start_column=9, end_row=base + 16, end_column=9)
                wave.cell(base, 9, "关断")
                wave.merge_cells(start_row=base, start_column=10, end_row=base, end_column=17)
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(start_row=base + 1, start_column=10, end_row=base + 16, end_column=17)
            stale_image = XLImage(str(image))
            stale_image.anchor = "J108"
            wave.add_image(stale_image)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_806A_000.tss"),
                    profile_code="UH",
                ),
                report,
                images={("关断过程", "ΔVce"): image},
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(summary.data_row, 6)
            self.assertEqual(summary.waveform_anchor_row, 54)
            self.assertEqual(saved_wave.cell(54, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(71, 1).value, "750V_806A")
            self.assertIsNone(saved_wave.cell(107, 1).value)
            self.assertIsNone(saved_wave.cell(124, 1).value)
            self.assertIsNone(saved_wave.cell(141, 1).value)
            self.assertEqual(saved_wave.cell(107, 9).value, "关断")
            self.assertEqual(saved_wave.cell(107, 10).value, "△Vce（V）")
            self.assertEqual(len(saved_wave._images), 1)
            image_rows = [
                img.anchor._from.row + 1
                for img in saved_wave._images
                if hasattr(img.anchor, "_from")
            ]
            self.assertEqual(image_rows, [55])

    def test_dpt_template_clears_same_setpoint_stale_signature_block(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.mcu2506_layout import (
            COL_CONDITION,
            COL_CURRENT,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "same_setpoint_stale_signature.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            for col in (1, 2, COL_CONDITION):
                data.merge_cells(
                    start_row=5,
                    start_column=col,
                    end_row=8,
                    end_column=col,
                )
            data["A5"] = "UH"
            data["B5"] = "25℃"
            data.cell(
                5,
                COL_CONDITION,
                "Rg_on = 5.6 ohm\nRg_off = 7.5 ohm\nCg = 22 nf",
            )
            data.cell(5, COL_VOLTAGE, 750)
            data.cell(5, COL_CURRENT, 805)

            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in (
                (1, "750V_805A_Rg_on5.6_Rg_off7.5_Cg22"),
                (54, "750V_805A_Rg_on3.3_Rg_off3.6_Cg10"),
            ):
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(
                    start_row=base,
                    start_column=10,
                    end_row=base,
                    end_column=17,
                )
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(
                    start_row=base + 1,
                    start_column=10,
                    end_row=base + 16,
                    end_column=17,
                )
            stale_image = XLImage(str(image))
            stale_image.anchor = "J55"
            wave.add_image(stale_image)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(
                        Path("samples")
                        / "RT"
                        / "Rg_on5.6_Rg_off7.5_Cg22_UH_750V_805A_000.tss"
                    ),
                    profile_code="UH",
                ),
                report,
                images={("关断过程", "ΔVce"): image},
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(summary.waveform_anchor_row, 1)
            self.assertEqual(
                saved_wave.cell(18, 1).value,
                "750V_805A_Rg_on5.6_Rg_off7.5_Cg22",
            )
            self.assertIsNone(saved_wave.cell(54, 1).value)
            self.assertIsNone(saved_wave.cell(71, 1).value)
            self.assertIsNone(saved_wave.cell(88, 1).value)
            self.assertEqual(len(saved_wave._images), 1)
            self.assertEqual(saved_wave._images[0].anchor._from.row + 1, 2)

    def test_dpt_different_gate_conditions_keep_images_in_both_current_orders(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_CONDITION,
            COL_CURRENT,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        cases = (
            (
                (805, "Rg_on3.3_Rg_off3.6_Cg10"),
                (1050, "Rg_on5.6_Rg_off7.5_Cg22"),
            ),
            (
                (1050, "Rg_on5.6_Rg_off7.5_Cg22"),
                (805, "Rg_on3.3_Rg_off3.6_Cg10"),
            ),
        )
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            for case_index, order in enumerate(cases):
                with self.subTest(order=order):
                    report = td_path / f"different_gate_conditions_{case_index}.xlsx"
                    wb = Workbook()
                    data = wb.active
                    data.title = "U相_双脉冲数据"
                    for col in (1, 2, COL_CONDITION):
                        data.merge_cells(
                            start_row=5,
                            start_column=col,
                            end_row=8,
                            end_column=col,
                        )
                    data["A5"] = "UH"
                    data["B5"] = "25℃"
                    data.cell(
                        5,
                        COL_CONDITION,
                        "Rg_on =  ohm\nRg_off =  ohm\nCg =  nf",
                    )

                    wave = wb.create_sheet("U相_双脉冲波形")
                    for base in (1, 54):
                        wave.merge_cells(
                            start_row=base,
                            start_column=1,
                            end_row=base + 16,
                            end_column=8,
                        )
                        wave.merge_cells(
                            start_row=base + 17,
                            start_column=1,
                            end_row=base + 33,
                            end_column=8,
                        )
                        wave.merge_cells(
                            start_row=base + 34,
                            start_column=1,
                            end_row=base + 50,
                            end_column=8,
                        )
                        wave.merge_cells(
                            start_row=base,
                            start_column=10,
                            end_row=base,
                            end_column=17,
                        )
                        wave.cell(base, 10, "△Vce（V）")
                        wave.merge_cells(
                            start_row=base + 1,
                            start_column=10,
                            end_row=base + 16,
                            end_column=17,
                        )
                    wb.save(report)

                    for current, condition in order:
                        write_report_template(
                            ExtractResult(
                                source_path=str(
                                    Path("samples")
                                    / "RT"
                                    / f"{condition}_UH_750V_{current}A_000.tss"
                                ),
                                profile_code="UH",
                            ),
                            report,
                            images={("关断过程", "ΔVce"): image},
                        )

                    saved = load_workbook(report)
                    saved_data = saved["U相_双脉冲数据"]
                    saved_wave = saved["U相_双脉冲波形"]
                    self.assertEqual(
                        [saved_data.cell(row, COL_VOLTAGE).value for row in (5, 6)],
                        [750, 750],
                    )
                    self.assertEqual(
                        [saved_data.cell(row, COL_CURRENT).value for row in (5, 6)],
                        [item[0] for item in order],
                    )
                    self.assertEqual(len(saved_wave._images), 2)
                    self.assertEqual(
                        sorted(image.anchor._from.row + 1 for image in saved_wave._images),
                        [2, 55],
                    )
                    for base, (current, condition) in zip((1, 54), order):
                        label = str(saved_wave.cell(base + 17, 1).value)
                        self.assertEqual(label, f"750V_{current}A_{condition}")

    def test_dpt_decimal_setpoint_blocks_keep_all_existing_images(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_CONDITION,
            COL_CURRENT,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        setpoints = (
            (900, 494.9),
            (900, 693.37),
            (850, 777.7),
            (850, 1061.01),
        )
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "decimal_setpoint_retention.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            for col in (1, 2, COL_CONDITION):
                data.merge_cells(
                    start_row=5,
                    start_column=col,
                    end_row=8,
                    end_column=col,
                )
            data["A5"] = "UH"
            data["B5"] = "25℃"
            data.cell(
                5,
                COL_CONDITION,
                "Rg_on =  ohm\nRg_off =  ohm\nCg =  nf",
            )

            wave = wb.create_sheet("U相_双脉冲波形")
            for block_index in range(len(setpoints)):
                base = 1 + block_index * 53
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.merge_cells(
                    start_row=base,
                    start_column=10,
                    end_row=base,
                    end_column=17,
                )
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(
                    start_row=base + 1,
                    start_column=10,
                    end_row=base + 16,
                    end_column=17,
                )
            wb.save(report)

            anchors = []
            for voltage, current in setpoints:
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(
                            Path("samples")
                            / "RT"
                            / f"uh-{voltage}v-{current}a_000.tss"
                        ),
                        profile_code="UH",
                    ),
                    report,
                    images={("关断过程", "ΔVce"): image},
                )
                anchors.append(summary.waveform_anchor_row)

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(anchors, [1, 54, 107, 160])
            self.assertEqual(
                [saved_data.cell(row, COL_VOLTAGE).value for row in range(5, 9)],
                [item[0] for item in setpoints],
            )
            self.assertEqual(
                [saved_data.cell(row, COL_CURRENT).value for row in range(5, 9)],
                [round(item[1], 1) for item in setpoints],
            )
            self.assertEqual(len(saved_wave._images), 4)
            self.assertEqual(
                sorted(item.anchor._from.row + 1 for item in saved_wave._images),
                [2, 55, 108, 161],
            )
            expected_labels = (
                "900V_494.9A",
                "900V_693.37A",
                "850V_777.7A",
                "850V_1061A",
            )
            for block_index, expected_label in enumerate(expected_labels):
                base = 1 + block_index * 53
                self.assertEqual(saved_wave.cell(base, 1).value, "UH_25℃")
                self.assertEqual(saved_wave.cell(base + 17, 1).value, expected_label)

    def test_dpt_cleanup_preserves_unparseable_legacy_condition_block(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.mcu2506_layout import (
            COL_CURRENT,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "legacy_condition_retention.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            data.merge_cells("A5:A6")
            data.merge_cells("B5:B6")
            data["A5"] = "UH"
            data["B5"] = "25℃"
            data.cell(5, COL_VOLTAGE, 900)
            data.cell(5, COL_CURRENT, 693.4)
            data.cell(6, COL_VOLTAGE, 900)
            data.cell(6, COL_CURRENT, 494.9)

            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in (
                (1, "900V_693.37A"),
                (54, "900V_494.9安"),
            ):
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(
                    start_row=base,
                    start_column=10,
                    end_row=base,
                    end_column=17,
                )
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(
                    start_row=base + 1,
                    start_column=10,
                    end_row=base + 16,
                    end_column=17,
                )
            legacy_image = XLImage(str(image))
            legacy_image.anchor = "J55"
            wave.add_image(legacy_image)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(
                        Path("samples") / "RT" / "UH_900V_693.37A_000.tss"
                    ),
                    profile_code="UH",
                ),
                report,
                images={("关断过程", "ΔVce"): image},
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(saved_wave.cell(54, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(71, 1).value, "900V_494.9安")
            self.assertEqual(len(saved_wave._images), 2)
            self.assertEqual(
                sorted(item.anchor._from.row + 1 for item in saved_wave._images),
                [2, 55],
            )

    def test_dpt_waveform_insert_shifts_existing_image_anchors(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "insert_with_existing_images.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            data = wb.active
            data.title = "U相_双脉冲数据"
            for start, phase in ((5, "UH"), (9, "UL")):
                for col in (1, 2, 3):
                    data.merge_cells(
                        start_row=start,
                        start_column=col,
                        end_row=start + 3,
                        end_column=col,
                    )
                data.cell(start, 1, phase)
                data.cell(start, 2, "25℃")

            wave = wb.create_sheet("U相_双脉冲波形")
            for base, phase in ((1, "UH_RT"), (54, "UL_RT")):
                wave.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 16,
                    end_column=8,
                )
                wave.cell(base, 1, phase)
                wave.merge_cells(
                    start_row=base + 17,
                    start_column=1,
                    end_row=base + 33,
                    end_column=8,
                )
                wave.cell(base + 17, 1, "750V_805A")
                wave.merge_cells(
                    start_row=base + 34,
                    start_column=1,
                    end_row=base + 50,
                    end_column=8,
                )
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(
                    start_row=base,
                    start_column=10,
                    end_row=base,
                    end_column=17,
                )
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(
                    start_row=base + 1,
                    start_column=10,
                    end_row=base + 16,
                    end_column=17,
                )
            wb.save(report)

            for phase, current in (("UH", 1050), ("UL", 805), ("UH", 805)):
                write_report_template(
                    ExtractResult(
                        source_path=str(
                            Path("samples")
                            / "RT"
                            / f"{phase}_750V_{current}A_000.tss"
                        ),
                        profile_code=phase,
                    ),
                    report,
                    images={("关断过程", "ΔVce"): image},
                )

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(saved_data.cell(5, COL_VOLTAGE).value, 750)
            self.assertEqual(saved_data.cell(5, COL_CURRENT).value, 1050)
            self.assertEqual(saved_data.cell(6, COL_CURRENT).value, 805)
            self.assertEqual(saved_data.cell(9, COL_CURRENT).value, 805)
            self.assertEqual(
                [saved_wave.cell(base, 1).value for base in (1, 54, 107)],
                ["UH_25℃", "UH_25℃", "UL_25℃"],
            )
            self.assertEqual(len(saved_wave._images), 3)
            self.assertEqual(
                sorted(image.anchor._from.row + 1 for image in saved_wave._images),
                [2, 55, 108],
            )

    def test_dpt_condition_rows_survive_narrow_left_blocks_and_overview_images(self):
        from dpt_extractor.export.mcu2506_layout import COL_OFF
        from dpt_extractor.export.report_template import (
            DPT_OVERVIEW_IMAGE_PARAM,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope.png"
            _write_tiny_png(image)

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in ((1, "750V_1050A"), (54, "750V_805A")):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=7)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=7)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=7)
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(start_row=base, start_column=9, end_row=base, end_column=16)
                wave.cell(base, 9, "△Vce（V）")
                wave.merge_cells(start_row=base + 1, start_column=9, end_row=base + 16, end_column=16)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=1050.0),
                ),
                report,
                images={
                    DPT_OVERVIEW_IMAGE_PARAM: image,
                    ("关断过程", "ΔVce"): image,
                },
            )
            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_805A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=805.0),
                ),
                report,
                images={
                    DPT_OVERVIEW_IMAGE_PARAM: image,
                    ("关断过程", "ΔVce"): image,
                },
            )

            saved = load_workbook(report)
            saved_data = saved["U相_双脉冲数据"]
            saved_wave = saved["U相_双脉冲波形"]
            self.assertEqual(saved_data.cell(5, COL_OFF["delta_vce"]).value, 1050)
            self.assertEqual(saved_data.cell(6, COL_OFF["delta_vce"]).value, 805)
            self.assertEqual(saved_data.cell(5, 5).value, 1050)
            self.assertEqual(saved_data.cell(6, 5).value, 805)
            self.assertIsNone(saved_wave.cell(35, 1).value)
            self.assertIsNone(saved_wave.cell(88, 1).value)

    def test_dpt_unmatched_condition_uses_empty_data_row_without_overwriting_reserved_row(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A7")
            ws.merge_cells("B5:B7")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition in ((1, "750V_1050A"), (54, "750V_805A")):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
            wb.save(report)

            rows = []
            for filename, marker in (
                ("UH_750V_1048A_000.tss", 1048.0),
                ("UH_750V_806A_000.tss", 806.0),
                ("UH_600V_403A_000.tss", 403.0),
            ):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / filename),
                        profile_code="UH",
                        turn_off=TurnOffResult(delta_vce=marker),
                    ),
                    report,
                )
                rows.append(summary.data_row)

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(rows, [5, 6, 7])
            self.assertEqual(saved.cell(5, COL_VOLTAGE).value, 750)
            self.assertEqual(saved.cell(5, COL_CURRENT).value, 1048)
            self.assertEqual(saved.cell(5, COL_OFF["delta_vce"]).value, 1048)
            self.assertEqual(saved.cell(6, COL_CURRENT).value, 806)
            self.assertEqual(saved.cell(6, COL_OFF["delta_vce"]).value, 806)
            self.assertEqual(saved.cell(7, COL_VOLTAGE).value, 600)
            self.assertEqual(saved.cell(7, COL_CURRENT).value, 403)
            self.assertEqual(saved.cell(7, COL_OFF["delta_vce"]).value, 403)

    def test_dpt_filename_setpoints_are_strict_report_keys(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A7")
            ws.merge_cells("B5:B7")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1050)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1048A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=1048.0),
                ),
                report,
            )

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(summary.data_row, 6)
            self.assertIsNone(saved.cell(5, COL_OFF["delta_vce"]).value)
            self.assertEqual(saved.cell(6, COL_VOLTAGE).value, 750)
            self.assertEqual(saved.cell(6, COL_CURRENT).value, 1048)
            self.assertEqual(saved.cell(6, COL_OFF["delta_vce"]).value, 1048)

    def test_dpt_missing_filename_setpoint_fallback_is_limited_to_ten_units(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        def build_report(path: Path) -> None:
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A7")
            ws.merge_cells("B5:B7")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1050)
            wb.save(path)

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "within.xlsx"
            build_report(report)
            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
                    profile_code="UH",
                    idc_set=1041.0,
                    turn_off=TurnOffResult(delta_vce=1041.0, ic_off_max=1041.0),
                ),
                report,
            )
            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(summary.data_row, 5)
            self.assertEqual(saved.cell(5, COL_CURRENT).value, 1041)
            self.assertEqual(saved.cell(5, COL_OFF["delta_vce"]).value, 1041)

            report = Path(td) / "outside.xlsx"
            build_report(report)
            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
                    profile_code="UH",
                    idc_set=1039.0,
                    turn_off=TurnOffResult(delta_vce=1039.0, ic_off_max=1039.0),
                ),
                report,
            )
            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(summary.data_row, 6)
            self.assertIsNone(saved.cell(5, COL_OFF["delta_vce"]).value)
            self.assertEqual(saved.cell(6, COL_CURRENT).value, 1039)

    def test_dpt_waveform_blocks_do_not_keep_blank_reserved_rows_between_groups(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.merge_cells("A9:A12")
            ws.merge_cells("B9:B12")
            ws["A9"] = "UL"
            ws["B9"] = "25℃"

            wave = wb.create_sheet("U相_双脉冲波形")
            blocks = (
                ("UH_RT", "750V_1050A"),
                ("UH_RT", "750V_805A"),
                ("UH_RT", "750V_50A"),
                ("UH_RT", "600V_285A"),
                ("UL_RT", "750V_1050A"),
                ("UL_RT", "750V_805A"),
            )
            for idx, (phase_temp, condition) in enumerate(blocks):
                base = 1 + idx * 53
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, phase_temp)
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
            wb.save(report)

            anchors = []
            for phase, filename, marker in (
                ("UH", "UH_750V_1048A_000.tss", 1048.0),
                ("UH", "UH_750V_806A_000.tss", 806.0),
                ("UH", "UH_600V_403A_000.tss", 403.0),
                ("UL", "UL_750V_1048A_000.tss", 1048.0),
                ("UL", "UL_750V_806A_000.tss", 806.0),
            ):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / filename),
                        profile_code=phase,
                        turn_off=TurnOffResult(delta_vce=marker),
                    ),
                    report,
                )
                anchors.append(summary.waveform_anchor_row)

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(anchors, [1, 54, 107, 160, 213])
            self.assertEqual(saved_wave.cell(107, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(124, 1).value, "600V_403A")
            self.assertEqual(saved_wave.cell(160, 1).value, "UL_25℃")
            self.assertEqual(saved_wave.cell(177, 1).value, "750V_1048A")
            self.assertEqual(saved_wave.cell(194, 1).value, "总概览图")
            self.assertEqual(saved_wave.cell(213, 1).value, "UL_25℃")
            self.assertEqual(saved_wave.cell(230, 1).value, "750V_806A")
            self.assertIsNone(saved_wave.cell(266, 1).value)
            self.assertIsNone(saved_wave.cell(283, 1).value)
            self.assertIsNone(saved_wave.cell(300, 1).value)

    def test_dpt_template_inserts_extra_condition_row_with_matching_style(self):
        from openpyxl.styles import Alignment, Border, PatternFill, Side

        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws.merge_cells("C5:C8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws["C5"] = "Rg_on =  ohm\nRg_off =  ohm\nCg =  nf"
            ws.merge_cells("A9:A12")
            ws.merge_cells("B9:B12")
            ws.merge_cells("C9:C12")
            ws["A9"] = "UL"
            ws["B9"] = "25℃"
            ws["C9"] = "Rg_on =  ohm\nRg_off =  ohm\nCg =  nf"

            fill = PatternFill("solid", fgColor="D9EAD3")
            border = Border(
                left=Side(style="thin", color="808080"),
                right=Side(style="thin", color="808080"),
                top=Side(style="thin", color="808080"),
                bottom=Side(style="thin", color="808080"),
            )
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in range(5, 13):
                ws.row_dimensions[row].height = 22
                for col in range(1, 38):
                    cell = ws.cell(row, col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = alignment
                    cell.number_format = "0.000"
            for row in range(5, 9):
                ws.cell(row, COL_VOLTAGE, 750)
                ws.cell(row, COL_CURRENT, 1000 - (row - 5) * 200)

            wave = wb.create_sheet("U相_双脉冲波形")
            conditions = ("750V_1000A", "750V_800A", "750V_600A", "750V_400A")
            for idx, condition in enumerate(conditions):
                base = 1 + idx * 53
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
            base = 1 + len(conditions) * 53
            wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
            wave.cell(base, 1, "UL_RT")
            wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
            wave.cell(base + 17, 1, "750V_1000A")
            wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
            wave.cell(base + 34, 1, "总概览图")
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_200A_000.tss"),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=200.0),
                ),
                report,
            )

            saved = load_workbook(report)["U相_双脉冲数据"]
            merges = {str(rng) for rng in saved.merged_cells.ranges}
            self.assertEqual(summary.data_row, 9)
            self.assertEqual(summary.waveform_anchor_row, 213)
            self.assertIn("A5:A9", merges)
            self.assertIn("B5:B9", merges)
            self.assertIn("C5:C9", merges)
            self.assertIn("A10:A13", merges)
            self.assertEqual(saved.cell(10, 1).value, "UL")
            self.assertEqual(saved.cell(9, COL_VOLTAGE).value, 750)
            self.assertEqual(saved.cell(9, COL_CURRENT).value, 200)
            self.assertEqual(saved.cell(9, COL_OFF["delta_vce"]).value, 200)
            self.assertEqual(saved.row_dimensions[9].height, saved.row_dimensions[8].height)
            self.assertEqual(
                saved.cell(9, COL_OFF["delta_vce"]).fill.fgColor.rgb,
                saved.cell(8, COL_OFF["delta_vce"]).fill.fgColor.rgb,
            )
            self.assertEqual(
                saved.cell(9, COL_OFF["delta_vce"]).border.left.style,
                saved.cell(8, COL_OFF["delta_vce"]).border.left.style,
            )
            self.assertEqual(
                saved.cell(9, COL_OFF["delta_vce"]).alignment.horizontal,
                saved.cell(8, COL_OFF["delta_vce"]).alignment.horizontal,
            )
            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(saved_wave.cell(213, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(230, 1).value, "750V_200A")
            self.assertEqual(saved_wave.cell(266, 1).value, "UL_25℃")

    def test_dpt_report_written_values_normalize_font_and_alignment(self):
        from openpyxl.styles import Alignment, Font

        from dpt_extractor.export.mcu2506_layout import COL_ON
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOnResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "W相_双脉冲数据"
            ws.merge_cells("A25:A28")
            ws.merge_cells("B25:B28")
            ws.merge_cells("C25:C28")
            ws["A25"] = "WL"
            ws["B25"] = "-40℃"
            ws["C25"] = "Rg_on =  ohm\nRg_off =  ohm\nCg =  nf"

            stale_style_cols = (
                COL_ON["ls_on"],
                COL_ON["tr"],
                COL_ON["crosstalk"],
            )
            for col in stale_style_cols:
                cell = ws.cell(25, col)
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal=None, vertical="center")
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "LT" / "WL_750V_1048A_000.tss"),
                    profile_code="WL",
                    turn_on=TurnOnResult(
                        ls_on=32.518,
                        tr=111.27,
                        crosstalk_vmax=2.42,
                        crosstalk_vmin=-5.45,
                    ),
                ),
                report,
            )

            saved = load_workbook(report)["W相_双脉冲数据"]
            self.assertEqual(summary.data_row, 25)
            for col in stale_style_cols:
                cell = saved.cell(25, col)
                self.assertIsNotNone(cell.value)
                self.assertEqual(cell.font.sz, 12)
                self.assertEqual(cell.alignment.horizontal, "center")
                self.assertEqual(cell.alignment.vertical, "center")

    def test_dpt_template_appends_waveform_block_with_two_separator_rows(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            for row, current in zip(range(5, 9), (1000, 800, 600, 400)):
                ws.cell(row, COL_VOLTAGE, 750)
                ws.cell(row, COL_CURRENT, current)

            wave = wb.create_sheet("U相_双脉冲波形")
            for idx, current in enumerate((1000, 800, 600, 400)):
                base = 1 + idx * 53
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, f"750V_{current}A")
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
            wave.row_dimensions[158].height = 6
            wave.row_dimensions[159].height = 6
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_200A_000.tss"),
                    profile_code="UH",
                ),
                report,
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(summary.waveform_anchor_row, 213)
            self.assertEqual(saved_wave.cell(213, 1).value, "UH_25℃")
            self.assertEqual(saved_wave.cell(230, 1).value, "750V_200A")
            self.assertEqual(saved_wave.cell(247, 1).value, "总概览图")
            self.assertIsNone(saved_wave.cell(211, 1).value)
            self.assertIsNone(saved_wave.cell(212, 1).value)

    def test_dpt_parameter_images_share_one_display_size_per_condition(self):
        from dpt_extractor.export.report_template import (
            REPORT_IMAGE_DISPLAY_SIZE_PX,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_1050A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            wave.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
            wave.cell(1, 10, "△Vce（V）")
            wave.merge_cells(start_row=2, start_column=10, end_row=17, end_column=13)
            wave.merge_cells(start_row=1, start_column=18, end_row=1, end_column=28)
            wave.cell(1, 18, "dv/dt(V/ns)")
            wave.merge_cells(start_row=2, start_column=18, end_row=17, end_column=28)
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                profile_code="UH",
            )
            write_report_template(
                result,
                report,
                images={
                    ("关断过程", "ΔVce"): image,
                    ("关断过程", "dv/dt"): image,
                },
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(len(saved_wave._images), 2)
            anchors = _drawing_anchor_counts(report)
            self.assertEqual(anchors["twoCellAnchor"], 2)
            self.assertEqual(anchors["oneCellAnchor"], 0)

    def test_dpt_parameter_images_share_one_display_size_across_conditions(self):
        from dpt_extractor.export.report_template import (
            REPORT_IMAGE_DISPLAY_SIZE_PX,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            for base, condition, row_height in (
                (1, "750V_1050A", 20.0),
                (54, "750V_805A", 10.0),
            ):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
                wave.merge_cells(start_row=base, start_column=10, end_row=base, end_column=17)
                wave.cell(base, 10, "△Vce（V）")
                wave.merge_cells(start_row=base + 1, start_column=10, end_row=base + 16, end_column=17)
                for row in range(base + 1, base + 17):
                    wave.row_dimensions[row].height = row_height
            wb.save(report)

            for current in ("UH_750V_1050A_000.tss", "UH_750V_805A_000.tss"):
                write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / current),
                        profile_code="UH",
                    ),
                    report,
                    images={("关断过程", "ΔVce"): image},
                )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(len(saved_wave._images), 2)
            anchors = _drawing_anchor_counts(report)
            self.assertEqual(anchors["twoCellAnchor"], 2)
            self.assertEqual(anchors["oneCellAnchor"], 0)

    def test_dpt_waveform_text_layout_and_open_zoom_are_normalized(self):
        from openpyxl.styles import Font

        from dpt_extractor.export.report_template import (
            _REPORT_VIEW_HORIZONTAL_MARGIN_PX,
            _REPORT_VIEW_MAX_ZOOM,
            _REPORT_VIEW_MIN_ZOOM,
            _WAVEFORM_HEADER_FONT_SIZE,
            _WAVEFORM_HEADER_ROW_HEIGHT_PX,
            _WAVEFORM_LEFT_LABEL_FONT_SIZE,
            _WAVEFORM_STATE_FONT_SIZE,
            _row_height_px,
            _used_columns_width_px,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, 4, 750)
            ws.cell(5, 5, 1050)
            ws.cell(6, 4, 750)
            ws.cell(6, 5, 805)
            wave = wb.create_sheet("U相_双脉冲波形")
            headers = ("△Vce（V）", "dv/dt(V/ns)", "di/dt(A/ns)", "Tf(ns)", "Td_off(ns)", "Eoff(mJ)")
            for base, condition, header_height in (
                (1, "750V_1050A", 6.0),
                (54, "750V_805A", 42.0),
            ):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
                for offset, label in ((0, "关断"), (17, "开通"), (34, "反向恢复")):
                    wave.merge_cells(
                        start_row=base + offset,
                        start_column=9,
                        end_row=base + offset + 16,
                        end_column=9,
                    )
                    wave.cell(base + offset, 9, label)
                for idx, header_text in enumerate(headers):
                    col = 10 + idx * 8
                    wave.merge_cells(
                        start_row=base,
                        start_column=col,
                        end_row=base,
                        end_column=col + 7,
                    )
                    cell = wave.cell(base, col, header_text)
                    cell.font = Font(size=22)
                    wave.row_dimensions[base].height = header_height
                    wave.merge_cells(
                        start_row=base + 1,
                        start_column=col,
                        end_row=base + 16,
                        end_column=col + 7,
                    )
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_805A_000.tss"),
                    profile_code="UH",
                ),
                report,
                target_screen_width_px=1600,
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            for row in (1, 18, 35, 54, 71, 88):
                self.assertEqual(_row_height_px(saved_wave, row), _WAVEFORM_HEADER_ROW_HEIGHT_PX)
            self.assertEqual(saved_wave.cell(1, 10).font.sz, _WAVEFORM_HEADER_FONT_SIZE)
            self.assertTrue(saved_wave.cell(1, 10).alignment.shrink_to_fit)
            self.assertEqual(saved_wave.cell(71, 1).font.sz, _WAVEFORM_LEFT_LABEL_FONT_SIZE)
            self.assertTrue(saved_wave.cell(71, 1).alignment.shrink_to_fit)
            self.assertEqual(saved_wave.cell(88, 9).font.sz, _WAVEFORM_STATE_FONT_SIZE)

            usable_width = max(960, 1600 - _REPORT_VIEW_HORIZONTAL_MARGIN_PX)
            expected_zoom = int(usable_width * 100 / max(1, _used_columns_width_px(saved_wave)))
            expected_zoom = max(
                _REPORT_VIEW_MIN_ZOOM,
                min(_REPORT_VIEW_MAX_ZOOM, expected_zoom),
            )
            self.assertEqual(saved_wave.sheet_view.zoomScale, expected_zoom)

    def test_dpt_irr_header_does_not_match_didt_irr_slot(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_805A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            wave.merge_cells(start_row=35, start_column=10, end_row=35, end_column=17)
            wave.cell(35, 10, "Didt_Irr90-10")
            wave.merge_cells(start_row=36, start_column=10, end_row=51, end_column=17)
            wave.merge_cells(start_row=35, start_column=18, end_row=35, end_column=25)
            wave.cell(35, 18, "Irr")
            wave.merge_cells(start_row=36, start_column=18, end_row=51, end_column=25)
            wave.cell(36, 18, 12)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_805A_000.tss"),
                    profile_code="UH",
                ),
                report,
                images={
                    ("反向恢复", "Irr"): image,
                    ("反向恢复", "di/dt"): image,
                },
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(len(saved_wave._images), 2)
            self.assertIsNone(saved_wave.cell(36, 18).value)

    def test_dpt_full_image_set_writes_every_slot_for_multiple_conditions(self):
        from dpt_extractor.export.report_template import (
            DPT_REPORT_IMAGE_PARAMS,
            _DPT_IMAGE_HEADERS,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            block_offsets = {"off": 0, "on": 17, "rr": 34}
            for base, condition in ((1, "750V_1050A"), (54, "750V_805A")):
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
                for block, row_offset in block_offsets.items():
                    headers = [
                        header_text
                        for _key, (item_block, header_text) in _DPT_IMAGE_HEADERS.items()
                        if item_block == block
                    ]
                    if block == "rr":
                        headers = ["Didt_Irr"] + [
                            header for header in headers if header != "Didt_Irr"
                        ]
                    for idx, header_text in enumerate(headers):
                        col = 10 + idx * 8
                        header_row = base + row_offset
                        wave.merge_cells(
                            start_row=header_row,
                            start_column=col,
                            end_row=header_row,
                            end_column=col + 7,
                        )
                        wave.cell(header_row, col, header_text)
                        wave.merge_cells(
                            start_row=header_row + 1,
                            start_column=col,
                            end_row=header_row + 16,
                            end_column=col + 7,
                        )
                        wave.cell(header_row + 1, col, "slot")
            wb.save(report)

            images = {param: image for param in DPT_REPORT_IMAGE_PARAMS}
            for current in ("UH_750V_1050A_000.tss", "UH_750V_805A_000.tss"):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / current),
                        profile_code="UH",
                    ),
                    report,
                    images=images,
                )
                self.assertEqual(summary.images_written, len(DPT_REPORT_IMAGE_PARAMS))

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(len(saved_wave._images), len(DPT_REPORT_IMAGE_PARAMS) * 2)
            leftovers = [
                (cell.row, cell.column)
                for row in saved_wave.iter_rows(min_row=1, max_row=104)
                for cell in row
                if cell.value == "slot"
            ]
            self.assertEqual(leftovers, [])

    def test_single_pulse_template_skips_turn_on_and_reverse_recovery_images(self):
        from dpt_extractor.export.report_template import (
            DPT_REPORT_IMAGE_PARAMS,
            _DPT_IMAGE_HEADERS,
            dpt_report_image_params_for_result,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_1050A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            block_offsets = {"off": 0, "on": 17, "rr": 34}
            block_counts = {"off": 0, "on": 0, "rr": 0}
            for _key, (block, header_text) in _DPT_IMAGE_HEADERS.items():
                col = 10 + block_counts[block] * 8
                block_counts[block] += 1
                header_row = 1 + block_offsets[block]
                wave.merge_cells(
                    start_row=header_row,
                    start_column=col,
                    end_row=header_row,
                    end_column=col + 7,
                )
                wave.cell(header_row, col, header_text)
                wave.merge_cells(
                    start_row=header_row + 1,
                    start_column=col,
                    end_row=header_row + 16,
                    end_column=col + 7,
                )
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                profile_code="UH",
                single_pulse_mode=True,
            )
            allowed_params = dpt_report_image_params_for_result(result)
            self.assertFalse(
                any(
                    section in {"开通", "反向恢复"}
                    for section, _name in allowed_params
                )
            )

            summary = write_report_template(
                result,
                report,
                images={param: image for param in DPT_REPORT_IMAGE_PARAMS},
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            self.assertEqual(summary.images_written, len(allowed_params))
            self.assertEqual(len(saved_wave._images), len(allowed_params))

    def test_dpt_image_slots_fit_fixed_4x3_display_size(self):
        from openpyxl.worksheet.cell_range import CellRange

        from dpt_extractor.export.report_template import (
            REPORT_IMAGE_DISPLAY_SIZE_PX,
            _column_width_px,
            _range_size_px,
            _row_height_px,
            write_report_template,
        )
        from dpt_extractor.models.results import ExtractResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "report.xlsx"
            image = td_path / "scope_4x3.png"
            _write_rgb_png(image, (1280, 960))

            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            wave.merge_cells(start_row=1, start_column=1, end_row=17, end_column=8)
            wave.cell(1, 1, "UH_RT")
            wave.merge_cells(start_row=18, start_column=1, end_row=34, end_column=8)
            wave.cell(18, 1, "750V_1050A")
            wave.merge_cells(start_row=35, start_column=1, end_row=51, end_column=8)
            wave.cell(35, 1, "总概览图")
            wave.merge_cells(start_row=1, start_column=10, end_row=1, end_column=17)
            wave.cell(1, 10, "△Vce（V）")
            wave.merge_cells(start_row=2, start_column=10, end_row=17, end_column=17)
            wb.save(report)

            write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1050A_000.tss"),
                    profile_code="UH",
                ),
                report,
                images={("关断过程", "ΔVce"): image},
            )

            saved_wave = load_workbook(report)["U相_双脉冲波形"]
            anchors = _drawing_anchor_counts(report)
            self.assertEqual(anchors["twoCellAnchor"], 1)
            self.assertEqual(anchors["oneCellAnchor"], 0)
            slot = CellRange("J2:Q17")
            slot_w, slot_h = _range_size_px(saved_wave, slot)
            self.assertGreaterEqual(slot_w, REPORT_IMAGE_DISPLAY_SIZE_PX[0] - 2)
            self.assertGreaterEqual(slot_h, REPORT_IMAGE_DISPLAY_SIZE_PX[1] - 2)
            self.assertGreaterEqual(
                sum(_column_width_px(saved_wave, col) for col in range(10, 18)),
                REPORT_IMAGE_DISPLAY_SIZE_PX[0],
            )
            self.assertGreaterEqual(
                sum(_row_height_px(saved_wave, row) for row in range(2, 18)),
                REPORT_IMAGE_DISPLAY_SIZE_PX[1],
            )

    def test_dpt_template_supports_more_than_four_condition_slots(self):
        from dpt_extractor.export.mcu2506_layout import COL_OFF
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "expanded_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A10")
            ws.merge_cells("B5:B10")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            wave = wb.create_sheet("U相_双脉冲波形")
            conditions = (
                "750V_1050A",
                "750V_805A",
                "750V_50A",
                "600V_285A",
                "750V_1200A",
            )
            for idx, condition in enumerate(conditions):
                base = 1 + idx * 53
                wave.merge_cells(start_row=base, start_column=1, end_row=base + 16, end_column=8)
                wave.cell(base, 1, "UH_RT")
                wave.merge_cells(start_row=base + 17, start_column=1, end_row=base + 33, end_column=8)
                wave.cell(base + 17, 1, condition)
                wave.merge_cells(start_row=base + 34, start_column=1, end_row=base + 50, end_column=8)
                wave.cell(base + 34, 1, "总概览图")
            wb.save(report)

            for row, current in zip(range(5, 9), (1050, 805, 50, 285)):
                ws.cell(row, 4, 750 if current != 285 else 600)
                ws.cell(row, 5, current)
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "RT" / "UH_750V_1200A_000.tss"),
                profile_code="UH",
                turn_off=TurnOffResult(delta_vce=555.0),
            )
            summary = write_report_template(result, report)
            saved = load_workbook(report)

            self.assertEqual(summary.data_row, 9)
            self.assertEqual(summary.waveform_anchor_row, 213)
            self.assertEqual(saved["U相_双脉冲数据"].cell(9, COL_OFF["delta_vce"]).value, 555.0)

    def test_dpt_template_distinguishes_gate_resistance_condition_groups(self):
        from dpt_extractor.export.mcu2506_layout import (
            COL_CONDITION,
            COL_CURRENT,
            COL_OFF,
            COL_VOLTAGE,
        )
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "condition_groups.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            for start, rg_on, rg_off, cg in (
                (5, 3.3, 3.6, 10),
                (9, 5.6, 7.5, 22),
            ):
                ws.merge_cells(start_row=start, start_column=1, end_row=start + 3, end_column=1)
                ws.merge_cells(start_row=start, start_column=2, end_row=start + 3, end_column=2)
                ws.merge_cells(start_row=start, start_column=3, end_row=start + 3, end_column=3)
                ws.cell(start, 1, "UH")
                ws.cell(start, 2, "25℃")
                ws.cell(
                    start,
                    COL_CONDITION,
                    f"Rg_on = {rg_on} ohm\nRg_off = {rg_off} ohm\nCg = {cg} nf",
                )
                ws.cell(start, COL_VOLTAGE, 750)
                ws.cell(start, COL_CURRENT, 805)
            wb.save(report)

            rows = []
            for filename, marker in (
                ("Rg_on3.3_Rg_off3.6_Cg10_UH_750V_805A_000.tss", 33.0),
                ("UH_750V_805A_Rg_on5.6_Rg_off7.5_Cg22_000.tss", 56.0),
            ):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / filename),
                        profile_code="UH",
                        turn_off=TurnOffResult(delta_vce=marker),
                    ),
                    report,
                )
                rows.append(summary.data_row)

            saved = load_workbook(report)["U相_双脉冲数据"]
            self.assertEqual(rows, [5, 9])
            self.assertEqual(saved.cell(5, COL_OFF["delta_vce"]).value, 33.0)
            self.assertEqual(saved.cell(9, COL_OFF["delta_vce"]).value, 56.0)
            self.assertEqual(saved.cell(5, COL_CURRENT).value, 805)
            self.assertEqual(saved.cell(9, COL_CURRENT).value, 805)

    def test_dpt_template_syncs_condition_cell_from_filename(self):
        from dpt_extractor.export.mcu2506_layout import COL_CONDITION, COL_OFF
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "sync_condition.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A8")
            ws.merge_cells("B5:B8")
            ws.merge_cells("C5:C8")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(
                5,
                COL_CONDITION,
                "备注保留\nRg_on =  ohm\nRg_off =  ohm\nCg =  nf",
            )
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(
                        Path("samples")
                        / "RT"
                        / "UH_750V_805A_Rg_on3.233_Rg_off3.586_Cg10_000.tss"
                    ),
                    profile_code="UH",
                    turn_off=TurnOffResult(delta_vce=12.0),
                ),
                report,
            )

            saved = load_workbook(report)["U相_双脉冲数据"]
            condition = saved.cell(5, COL_CONDITION).value
            self.assertEqual(summary.data_row, 5)
            self.assertIn("备注保留", condition)
            self.assertIn("Rg_on = 3.233 ohm", condition)
            self.assertIn("Rg_off = 3.586 ohm", condition)
            self.assertIn("Cg = 10 nf", condition)
            self.assertEqual(condition.count("Rg_on"), 1)
            self.assertEqual(condition.count("Rg_off"), 1)
            self.assertEqual(condition.count("Cg"), 1)
            self.assertEqual(saved.cell(5, COL_OFF["delta_vce"]).value, 12.0)

    def test_dpt_template_multi_pulse_rows_do_not_overwrite_next_condition(self):
        from dpt_extractor.export.mcu2506_layout import COL_CURRENT, COL_OFF, COL_ON, COL_VOLTAGE
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, TurnOffResult, TurnOnResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "multi_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "U相_双脉冲数据"
            ws.merge_cells("A5:A7")
            ws.merge_cells("B5:B7")
            ws["A5"] = "UH"
            ws["B5"] = "25℃"
            ws.cell(5, COL_VOLTAGE, 750)
            ws.cell(5, COL_CURRENT, 1048)
            ws.cell(6, COL_VOLTAGE, 750)
            ws.cell(6, COL_CURRENT, 806)
            wb.save(report)

            rows = [
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1048A_000.tss"),
                    profile_code="UH",
                    detected_pulse_count=3,
                    off_pulse_index=1,
                    on_pulse_index=2,
                    turn_off=TurnOffResult(delta_vce=101.0),
                    turn_on=TurnOnResult(delta_vce=201.0),
                ),
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_1048A_000.tss"),
                    profile_code="UH",
                    detected_pulse_count=3,
                    off_pulse_index=2,
                    on_pulse_index=3,
                    turn_off=TurnOffResult(delta_vce=102.0),
                    turn_on=TurnOnResult(delta_vce=202.0),
                ),
            ]
            summary = write_report_template(rows, report)
            saved = load_workbook(report)["U相_双脉冲数据"]

            self.assertEqual(summary.data_row, 5)
            self.assertEqual(summary.data_row_end, 6)
            self.assertEqual(summary.data_rows_written, 2)
            self.assertEqual(saved.cell(5, COL_OFF["delta_vce"]).value, 101.0)
            self.assertEqual(saved.cell(5, COL_ON["delta_vce"]).value, 201.0)
            self.assertEqual(saved.cell(6, COL_OFF["delta_vce"]).value, 102.0)
            self.assertEqual(saved.cell(6, COL_ON["delta_vce"]).value, 202.0)
            self.assertEqual(saved.cell(7, COL_CURRENT).value, 806)

    def test_short_template_uses_labeled_picture_blocks_not_row_offsets(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "short_report.xlsx"
            image = td_path / "short.png"
            _write_tiny_png(image)

            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws.merge_cells("A5:A6")
            ws["A5"] = 25
            ws["B5"] = "UH"
            ws["B6"] = "UL"
            pic = wb.create_sheet("短路测试图片")
            for base, phase_temp in ((1, "UH_RT"), (12, "UL_RT")):
                pic.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 3,
                    end_column=5,
                )
                pic.cell(base, 1, phase_temp)
                pic.cell(base + 4, 1, "Vce")
                pic.cell(base + 4, 2, "Imax")
                pic.merge_cells(
                    start_row=base,
                    start_column=6,
                    end_row=base,
                    end_column=8,
                )
                pic.cell(base, 6, "短路电流Imax 短路时间Tsc 短路能量Esc_本管")
                pic.merge_cells(
                    start_row=base + 1,
                    start_column=6,
                    end_row=base + 9,
                    end_column=8,
                )
            wb.save(report)

            rows_and_anchors = []
            for phase, filename, current in (
                ("UL", "UL_750V_000.tss", 2222.0),
                ("UH", "UH_750V_000.tss", 1111.0),
            ):
                summary = write_report_template(
                    ExtractResult(
                        source_path=str(Path("samples") / "RT" / filename),
                        profile_code=phase,
                        vdc_set=750.0,
                        short_circuit_mode=True,
                        short_circuit=ShortCircuitResult(ic_max=current),
                    ),
                    report,
                    images={("短路过程", "短路电流Imax"): image},
                )
                rows_and_anchors.append((summary.data_row, summary.waveform_anchor_row))

            saved = load_workbook(report)
            self.assertEqual(rows_and_anchors, [(6, 12), (5, 1)])
            self.assertEqual(saved["短路测试"].cell(5, 5).value, 1111)
            self.assertEqual(saved["短路测试"].cell(6, 5).value, 2222)
            self.assertEqual(saved["短路测试图片"].cell(1, 1).value, "UH_25℃")
            self.assertEqual(saved["短路测试图片"].cell(12, 1).value, "UL_25℃")
            self.assertEqual(saved["短路测试图片"].cell(6, 1).value, "750V")
            self.assertEqual(saved["短路测试图片"].cell(6, 2).value, "1111A")
            self.assertEqual(saved["短路测试图片"].cell(17, 1).value, "750V")
            self.assertEqual(saved["短路测试图片"].cell(17, 2).value, "2222A")
            self.assertEqual(len(saved["短路测试图片"]._images), 2)

    def test_short_template_manual_page_conditions_override_path(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            report = Path(td) / "short_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            for start_row, temp in ((5, "RT"), (7, "LT")):
                ws.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row + 1,
                    end_column=1,
                )
                ws.cell(start_row, 1, temp)
                ws.cell(start_row, 2, "UH")
                ws.cell(start_row + 1, 2, "UL")
            pic = wb.create_sheet("短路测试图片")
            for base, phase_temp in ((1, "UH_RT"), (12, "UL_LT")):
                pic.merge_cells(
                    start_row=base,
                    start_column=1,
                    end_row=base + 3,
                    end_column=5,
                )
                pic.cell(base, 1, phase_temp)
            wb.save(report)

            summary = write_report_template(
                ExtractResult(
                    source_path=str(Path("samples") / "RT" / "UH_750V_000.tss"),
                    profile_code="UH",
                    short_circuit_mode=True,
                    short_circuit=ShortCircuitResult(ic_max=1234.0),
                ),
                report,
                temperature_labels={"RT": "26.5℃", "HT": "175℃", "LT": "-55℃"},
                temperature_code="LT",
                phase_code="UL",
            )

            saved = load_workbook(report)
            self.assertEqual(summary.data_row, 8)
            self.assertEqual(summary.waveform_anchor_row, 12)
            self.assertIsNone(saved["短路测试"].cell(5, 5).value)
            self.assertEqual(saved["短路测试"].cell(8, 5).value, 1234)
            self.assertEqual(saved["短路测试"].cell(7, 1).value, "-55℃")
            self.assertEqual(saved["短路测试图片"].cell(1, 1).value, "UH_26.5℃")
            self.assertEqual(saved["短路测试图片"].cell(12, 1).value, "UL_-55℃")

    def test_short_same_code_keeps_distinct_actual_temperatures(self):
        from openpyxl.drawing.image import Image as XLImage

        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "short_report.xlsx"
            image = td_path / "preserved.png"
            _write_tiny_png(image)
            wb = Workbook()
            data = wb.active
            data.title = "短路测试"
            for start_row, temperature in ((5, "25℃"), (7, "-40℃")):
                data.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row + 1,
                    end_column=1,
                )
                data.cell(start_row, 1, temperature)
                data.cell(start_row, 2, "UH")
                data.cell(start_row + 1, 2, "UL")
            picture = wb.create_sheet("短路测试图片")
            picture.merge_cells("A1:E4")
            picture["A1"] = "UL_-40℃"
            preserved = wb.create_sheet("保留内容")
            preserved["A1"] = "=2+2"
            preserved.add_image(XLImage(str(image)), "C3")
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "UL_750V_000.tss"),
                profile_code="UL",
                short_circuit_mode=True,
                short_circuit=ShortCircuitResult(ic_max=1111.0),
            )
            first = write_report_template(
                result,
                report,
                temperature_labels={"RT": "25℃", "HT": "150℃", "LT": "-55℃"},
                temperature_code="LT",
                phase_code="UL",
            )
            result.short_circuit.ic_max = 2222.0
            second = write_report_template(
                result,
                report,
                temperature_labels={"RT": "25℃", "HT": "150℃", "LT": "-60℃"},
                temperature_code="LT",
                phase_code="UL",
            )

            saved = load_workbook(report, data_only=False)
            self.assertEqual(first.data_row, 10)
            self.assertEqual(second.data_row, 12)
            self.assertEqual(
                saved.sheetnames,
                ["短路测试", "短路测试图片", "保留内容"],
            )
            self.assertEqual(saved["短路测试"]["A5"].value, "25℃")
            self.assertEqual(saved["短路测试"]["A7"].value, "-40℃")
            self.assertEqual(saved["短路测试"]["A9"].value, "-55℃")
            self.assertEqual(saved["短路测试"]["A11"].value, "-60℃")
            self.assertEqual(saved["短路测试图片"]["A1"].value, "UL_-40℃")
            self.assertEqual(saved["短路测试图片"]["A5"].value, "UL_-55℃")
            self.assertEqual(saved["短路测试图片"]["A9"].value, "UL_-60℃")
            self.assertIsNone(saved["短路测试"].cell(6, 5).value)
            self.assertEqual(saved["短路测试"].cell(10, 5).value, 1111)
            self.assertEqual(saved["短路测试"].cell(12, 5).value, 2222)
            self.assertEqual(saved["保留内容"]["A1"].value, "=2+2")
            self.assertEqual(len(saved["保留内容"]._images), 1)
            anchor = saved["保留内容"]._images[0].anchor._from
            self.assertEqual((anchor.row, anchor.col), (2, 2))
            for code in ("RT", "HT", "LT"):
                identity = saved.defined_names[f"_DPT_TEMP_IDENTITY_{code}"]
                self.assertTrue(identity.hidden)

    def test_short_template_writes_row_and_global_image_slot(self):
        from dpt_extractor.export.report_template import write_report_template
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            report = td_path / "short_report.xlsx"
            image = td_path / "short.png"
            _write_tiny_png(image)

            wb = Workbook()
            ws = wb.active
            ws.title = "短路测试"
            ws.merge_cells("A5:A6")
            ws["A5"] = 25
            ws["B5"] = "UH"
            ws["B6"] = "UL"
            ws.merge_cells("A7:A8")
            ws["A7"] = 25
            ws["B7"] = "VH"
            ws["B8"] = "VL"
            pic = wb.create_sheet("短路测试图片")
            pic.merge_cells("A1:E4")
            pic["A1"] = "VH_RT"
            pic["A5"] = "Vce"
            pic["B5"] = "Imax"
            pic.merge_cells("F1:H1")
            pic["F1"] = "短路电流Imax 短路时间Tsc 短路能量Esc_本管"
            pic.merge_cells("F2:H10")
            wb.save(report)

            result = ExtractResult(
                source_path=str(Path("samples") / "RT" / "VH_750V_000.tss"),
                profile_code="VH",
                vdc_set=488.123,
                short_circuit_mode=True,
                short_circuit=ShortCircuitResult(ic_max=3210.5),
            )
            summary = write_report_template(
                result,
                report,
                images={("短路过程", "短路电流Imax"): image},
            )

            saved = load_workbook(report)
            self.assertEqual(summary.data_sheet, "短路测试")
            self.assertEqual(summary.data_row, 7)
            self.assertEqual(saved["短路测试"].cell(7, 4).value, 750)
            self.assertEqual(saved["短路测试"].cell(7, 5).value, 3210.5)
            self.assertEqual(saved["短路测试图片"].cell(6, 1).value, "750V")
            self.assertEqual(saved["短路测试图片"].cell(6, 2).value, "3210.5A")
            self.assertEqual(len(saved["短路测试图片"]._images), 1)


if __name__ == "__main__":
    unittest.main()
