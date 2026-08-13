"""End-to-end regressions for the production report task chain.

These tests intentionally exercise the queued QTimer screenshot path instead of
the older synchronous helper.  Report generation is one operation spanning the
page snapshot, screenshots, and Excel write, so each stage must keep the state
captured when the user clicked ``写入报告``.
"""

from __future__ import annotations

from copy import deepcopy
import os
from pathlib import Path
import tempfile
import time
import unittest
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PyQt6.QtCore import QEvent, QEventLoop, QObject
from PyQt6.QtGui import QImage
from PyQt6.QtWidgets import QApplication, QFileDialog, QMessageBox

from dpt_extractor.config.loader import load_config
from dpt_extractor.export.report_template import (
    DPT_REPORT_IMAGE_PARAMS,
    ReportWriteSummary,
)
from dpt_extractor.gui.main_window import (
    MainWindow,
    REPORT_PLOT_CAPTURE_SIZE,
    REPORT_PROGRESS_TOTAL,
    _ReportPrepareTask,
    _ReportWriteTask,
)
from dpt_extractor.io.waveform_loader import load_waveform
from dpt_extractor.models.bridge_profile import (
    guess_profile_from_path,
    make_profile,
)
from dpt_extractor.models.channel_mapping import (
    apply_mapping,
    infer_mapping_from_bundle,
)
from dpt_extractor.models.results import ExtractResult
from dpt_extractor.pipeline.extract import extract_all


ROOT = Path(__file__).resolve().parents[2]
SONGZHENXI_20260717_UH = (
    ROOT
    / "示例文件"
    / "songzhenxi"
    / "KSU2577"
    / "07CF2C1000 20260717"
    / "SMC"
    / "HT"
    / "UH_750V_1048A_000.tss"
)


class _GeometryEventCounter(QObject):
    def __init__(self) -> None:
        super().__init__()
        self.resize_count = 0
        self.move_count = 0

    def eventFilter(self, watched, event) -> bool:  # noqa: N802
        if event.type() == QEvent.Type.Resize:
            self.resize_count += 1
        elif event.type() == QEvent.Type.Move:
            self.move_count += 1
        return False


class _RecordingPool:
    def __init__(self) -> None:
        self.tasks: list[object] = []

    def start(self, task: object) -> None:
        self.tasks.append(task)


class _ImmediatePool:
    def start(self, task: object) -> None:
        task.run()  # type: ignore[attr-defined]


class TestReportEndToEndRegressions(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def _load_real_window(self) -> tuple[MainWindow, ExtractResult]:
        if not SONGZHENXI_20260717_UH.is_file():
            self.skipTest(f"songzhenxi 样例缺失: {SONGZHENXI_20260717_UH}")

        cfg = load_config()
        bundle = load_waveform(SONGZHENXI_20260717_UH)
        guessed = guess_profile_from_path(SONGZHENXI_20260717_UH.name)
        profile = make_profile(guessed.phase, guessed.bridge)
        inferred = infer_mapping_from_bundle(bundle, guessed.bridge)
        if inferred is not None:
            profile = apply_mapping(profile, inferred)
        result = extract_all(bundle, profile, cfg)

        win = MainWindow()
        win.bundle = bundle
        win.profile = profile
        win.result = result
        win.cfg = cfg
        win._set_profile_combos(profile)
        win._set_temperature_code("HT")
        win.result_table.set_result(result)
        win.wave_plot.plot_waveforms(bundle, profile, result)
        win.resize(1100, 760)
        win.show()
        self.app.processEvents()
        return win, result

    def _drain_events_until(self, predicate, *, timeout_s: float = 20.0) -> None:
        deadline = time.monotonic() + timeout_s
        while not predicate() and time.monotonic() < deadline:
            self.app.processEvents(QEventLoop.ProcessEventsFlag.AllEvents, 50)
        self.assertTrue(predicate(), "QTimer 报告链在超时前未完成")

    def test_capture_initialization_failure_restores_page_view_and_tempdir(self) -> None:
        for failure_stage in ("capture-page", "image-params"):
            with self.subTest(failure_stage=failure_stage):
                win, result = self._load_real_window()
                original_bundle = win.bundle
                original_profile = win.profile
                original_result = win.result
                vb = win.wave_plot.plot.getPlotItem().getViewBox()
                vb.setRange(xRange=(2.0, 3.0), yRange=(-1.0, 1.0), padding=0.0)
                self.app.processEvents()
                before_x, before_y = vb.viewRange()

                tempdir = tempfile.TemporaryDirectory()
                temp_path = Path(tempdir.name)
                original_apply = win._apply_report_page_state
                apply_calls = 0

                def apply_then_fail_first(page) -> None:
                    nonlocal apply_calls
                    apply_calls += 1
                    original_apply(page)
                    if apply_calls == 1:
                        raise RuntimeError("injected capture-page failure")

                try:
                    if failure_stage == "capture-page":
                        context = patch.object(
                            win,
                            "_apply_report_page_state",
                            side_effect=apply_then_fail_first,
                        )
                    else:
                        context = patch.object(
                            win,
                            "_report_image_params",
                            side_effect=RuntimeError("injected image-params failure"),
                        )
                    with context:
                        with self.assertRaisesRegex(
                            RuntimeError,
                            f"injected {failure_stage} failure",
                        ):
                            win._start_report_capture_sequence(tempdir, [result])

                    after_x, after_y = vb.viewRange()
                    for actual, expected in zip(after_x, before_x):
                        self.assertAlmostEqual(float(actual), float(expected), places=12)
                    for actual, expected in zip(after_y, before_y):
                        self.assertAlmostEqual(float(actual), float(expected), places=12)
                    self.assertIs(win.bundle, original_bundle)
                    self.assertIs(win.profile, original_profile)
                    self.assertIs(win.result, original_result)
                    self.assertIsNone(win._report_capture_state)
                    self.assertFalse(temp_path.exists())
                finally:
                    tempdir.cleanup()
                    win.close()
                    self.app.processEvents()

    def test_capture_completion_releases_local_zoom_space_before_excel(self) -> None:
        win, result = self._load_real_window()
        completed: dict[str, object] = {}

        def record_write_start(tempdir, images, results, **kwargs) -> None:
            completed.update(
                tempdir=tempdir,
                images=dict(images),
                results=list(results),
                kwargs=dict(kwargs),
            )

        tempdir = tempfile.TemporaryDirectory()
        try:
            self.assertTrue(win._try_begin_report_operation())
            win._report_request_id += 1
            request_id = win._report_request_id
            with patch.object(
                win,
                "_start_report_write_task",
                side_effect=record_write_start,
            ):
                win._start_report_capture_sequence(
                    tempdir,
                    [result],
                    request_id=request_id,
                )
                state = win._report_capture_state
                self.assertIsNotNone(state)
                assert state is not None
                restored_plot_top = win.wave_plot.plot.geometry().top()

                # Reproduce the last parameter screenshot layout: overview and
                # its scale bar occupy fixed rows above the main waveform.
                full_x = win.wave_plot._full_x_range
                self.assertIsNotNone(full_x)
                assert full_x is not None
                win.wave_plot._apply_x_us_per_div(
                    0.2,
                    center_us=0.5 * (float(full_x[0]) + float(full_x[1])),
                )
                panel_layout = win.wave_plot._waveform_panel.layout()
                self.assertIsNotNone(panel_layout)
                assert panel_layout is not None
                panel_layout.activate()
                self.assertFalse(win.wave_plot._overview_plot.isHidden())
                self.assertFalse(win.wave_plot._scope_scale_bar.isHidden())
                occupied_plot_top = win.wave_plot.plot.geometry().top()
                panel_top = panel_layout.contentsMargins().top()
                self.assertGreater(occupied_plot_top, panel_top)

                # Complete capture synchronously.  Excel must not start in the
                # same callback that hides the temporary rows.
                state.index = len(state.params)
                win._capture_next_report_image()
                self.assertFalse(completed)
                self.assertTrue(win.wave_plot._overview_plot.isHidden())
                self.assertTrue(win.wave_plot._scope_scale_bar.isHidden())
                self.assertEqual(
                    win.wave_plot.plot.geometry().top(),
                    restored_plot_top,
                )

                self._drain_events_until(lambda: bool(completed))

            self.assertEqual(completed["kwargs"]["request_id"], request_id)
        finally:
            win._report_request_id += 1
            win._report_capture_state = None
            win._release_report_operation()
            tempdir.cleanup()
            win.close()
            self.app.processEvents()

    def test_production_qtimer_capture_writes_all_19_real_pngs_without_geometry_change(
        self,
    ) -> None:
        win, result = self._load_real_window()
        result_before = deepcopy(result)
        self.assertEqual(win._report_image_params(), DPT_REPORT_IMAGE_PARAMS)
        self.assertEqual(len(DPT_REPORT_IMAGE_PARAMS), 19)

        window_events = _GeometryEventCounter()
        waveform_events = _GeometryEventCounter()
        win.installEventFilter(window_events)
        win.wave_plot.installEventFilter(waveform_events)

        before_window_geometry = win.geometry()
        before_window_state = win.windowState()
        before_waveform_geometry = win.wave_plot.geometry()
        before_minimum = win.wave_plot.minimumSize()
        before_maximum = win.wave_plot.maximumSize()
        before_policy = win.wave_plot.sizePolicy()
        before_x, before_y = (
            win.wave_plot.plot.getPlotItem().getViewBox().viewRange()
        )

        completed: dict[str, object] = {}

        def record_write_start(tempdir, images, results, **kwargs) -> None:
            completed.update(
                tempdir=tempdir,
                images=dict(images),
                results=list(results),
                kwargs=dict(kwargs),
            )

        tempdir = tempfile.TemporaryDirectory()
        try:
            self.assertTrue(win._try_begin_report_operation())
            win._report_request_id += 1
            request_id = win._report_request_id
            with patch.object(
                win,
                "_start_report_write_task",
                side_effect=record_write_start,
            ):
                win._start_report_capture_sequence(
                    tempdir,
                    [result],
                    request_id=request_id,
                    temperature_code="HT",
                    temperature_labels=win._temperature_display_labels(),
                    phase_code="UH",
                    image_result_index=0,
                )
                self._drain_events_until(lambda: bool(completed))

            images = completed["images"]
            self.assertIsInstance(images, dict)
            assert isinstance(images, dict)
            self.assertEqual(tuple(images), DPT_REPORT_IMAGE_PARAMS)
            self.assertEqual(len({Path(path) for path in images.values()}), 19)
            for param, path_value in images.items():
                path = Path(path_value)
                with self.subTest(param=param):
                    self.assertTrue(path.is_file())
                    self.assertGreater(path.stat().st_size, 0)
                    image = QImage(str(path))
                    self.assertFalse(image.isNull())
                    self.assertEqual(image.size(), REPORT_PLOT_CAPTURE_SIZE)
                    preview = image.scaled(64, 48)
                    colors = {
                        preview.pixelColor(x, y).rgba()
                        for y in range(preview.height())
                        for x in range(preview.width())
                    }
                    self.assertGreater(len(colors), 4, "截图不应是空白纯色图")

            after_x, after_y = (
                win.wave_plot.plot.getPlotItem().getViewBox().viewRange()
            )
            self.assertEqual(win.geometry(), before_window_geometry)
            self.assertEqual(win.windowState(), before_window_state)
            self.assertEqual(win.wave_plot.geometry(), before_waveform_geometry)
            self.assertEqual(win.wave_plot.minimumSize(), before_minimum)
            self.assertEqual(win.wave_plot.maximumSize(), before_maximum)
            self.assertEqual(
                win.wave_plot.sizePolicy().horizontalPolicy(),
                before_policy.horizontalPolicy(),
            )
            self.assertEqual(
                win.wave_plot.sizePolicy().verticalPolicy(),
                before_policy.verticalPolicy(),
            )
            self.assertEqual(window_events.resize_count, 0)
            self.assertEqual(window_events.move_count, 0)
            self.assertEqual(waveform_events.resize_count, 0)
            self.assertEqual(waveform_events.move_count, 0)
            self.assertAlmostEqual(after_x[0], before_x[0], places=6)
            self.assertAlmostEqual(after_x[1], before_x[1], places=6)
            self.assertAlmostEqual(after_y[0], before_y[0], places=6)
            self.assertAlmostEqual(after_y[1], before_y[1], places=6)
            self.assertEqual(win.result, result_before)
        finally:
            win._report_capture_state = None
            win._release_report_operation()
            tempdir.cleanup()
            win.close()

    def test_page_snapshot_survives_ui_changes_through_prepare_capture_and_write(
        self,
    ) -> None:
        win = MainWindow()
        pool = _RecordingPool()
        original_pool = win._load_pool
        bridged_tempdirs: list[tempfile.TemporaryDirectory] = []
        try:
            win._load_pool = pool  # type: ignore[assignment]
            win.result = ExtractResult(
                source_path=str(Path("operator") / "U_L" / "case.tss"),
                profile_code="UL",
                detected_pulse_count=2,
            )
            win.result.turn_off.eoff = 314.159
            win._set_profile_combos(make_profile("U", "lower"))
            win._set_temperature_code("LT")

            with tempfile.TemporaryDirectory() as tmp:
                output = Path(tmp) / "report.xlsx"
                output.write_bytes(b"existing report placeholder")
                win._report_output_path = output

                def bridge_capture(tempdir, results, **kwargs) -> None:
                    bridged_tempdirs.append(tempdir)
                    win._start_report_write_task(
                        tempdir,
                        {},
                        results,
                        request_id=kwargs["request_id"],
                        temperature_code=kwargs["temperature_code"],
                        temperature_labels=kwargs["temperature_labels"],
                        phase_code=kwargs["phase_code"],
                        image_result_index=kwargs["image_result_index"],
                    )

                with patch.object(
                    win,
                    "_start_report_capture_sequence",
                    side_effect=bridge_capture,
                ):
                    win._write_report_template()
                    self.assertEqual(len(pool.tasks), 1)
                    prepare = pool.tasks[0]
                    self.assertIsInstance(prepare, _ReportPrepareTask)
                    assert isinstance(prepare, _ReportPrepareTask)
                    self.assertEqual(prepare.profile.code, "UL")
                    self.assertEqual(prepare.temperature_code, "LT")
                    self.assertEqual(prepare.phase_code, "UL")
                    self.assertEqual(prepare.current_result.turn_off.eoff, 314.159)

                    # Simulate any later page change.  The operation guard normally
                    # prevents mouse input, but the report contract must still be a
                    # snapshot rather than a collection of live widget references.
                    win._set_profile_combos(make_profile("V", "upper"))
                    win._set_temperature_code("RT")
                    assert win.result is not None
                    win.result.profile_code = "VH"
                    win.result.turn_off.eoff = -1.0

                    prepare.run()
                    self._drain_events_until(lambda: len(pool.tasks) == 2)

                write = pool.tasks[1]
                self.assertIsInstance(write, _ReportWriteTask)
                assert isinstance(write, _ReportWriteTask)
                self.assertEqual(write.phase_code, "UL")
                self.assertEqual(write.temperature_code, "LT")
                self.assertEqual(write.temperature_labels["LT"], "-40℃")
                self.assertEqual(write.image_result_index, 0)
                self.assertIsInstance(write.result, ExtractResult)
                assert isinstance(write.result, ExtractResult)
                self.assertEqual(write.result.profile_code, "UL")
                self.assertEqual(write.result.turn_off.eoff, 314.159)

                self.assertEqual(win._current_report_phase_code(), "VH")
                self.assertEqual(win._current_temperature_code(), "RT")
                self.assertEqual(win.result.turn_off.eoff, -1.0)
        finally:
            for tempdir in bridged_tempdirs:
                tempdir.cleanup()
            win._report_prepare_tasks.clear()
            win._report_tasks.clear()
            win._release_report_operation()
            win._load_pool = original_pool
            win.close()

    def test_report_snapshot_keeps_inactive_manual_eon_after_nonreset_recalculation(
        self,
    ) -> None:
        """A preserved cursor cache must also remain materialized in report data."""

        win, _result = self._load_real_window()
        pool = _RecordingPool()
        original_pool = win._load_pool
        try:
            original_eon = float(win.result.turn_on.eon)
            untouched_eoff = float(win.result.turn_off.eoff)

            win._on_value_clicked("开通", "Eon")
            plot = win.wave_plot
            self.assertIsNotNone(plot._cursor_a)
            assert plot._cursor_a is not None
            plot._cursor_a.setPos(float(plot._cursor_a.value()) + 0.031)
            self.app.processEvents()

            adjusted_eon = float(win.result.turn_on.eon)
            self.assertNotAlmostEqual(adjusted_eon, original_eon, places=6)
            self.assertIn(("开通", "Eon"), win._manual_energy)

            # A non-reset recalculation is reachable through range, pulse and
            # channel operations.  Keep Eon inactive so merely restoring the
            # selected card cannot hide a stale result/manual-cache split.
            win._on_value_clicked("关断过程", "Eoff")
            win._recalculate(reset_manual=False)
            self.assertIn(("开通", "Eon"), win._manual_energy)

            win._load_pool = pool  # type: ignore[assignment]
            win._start_report_prepare_task()
            self.assertEqual(len(pool.tasks), 1)
            prepare = pool.tasks[0]
            self.assertIsInstance(prepare, _ReportPrepareTask)
            assert isinstance(prepare, _ReportPrepareTask)

            self.assertAlmostEqual(
                prepare.current_result.turn_on.eon,
                adjusted_eon,
                places=9,
            )
            self.assertAlmostEqual(
                prepare.current_result.turn_off.eoff,
                untouched_eoff,
                places=9,
            )

            # Exercise the production workbook writer with the request-frozen
            # object, not just the in-memory prepare boundary.
            from openpyxl import load_workbook

            from dpt_extractor.export.mcu2506_layout import (
                COL_OFF,
                COL_ON,
                DATA_ROW,
                build_mcu2506_workbook,
            )
            from dpt_extractor.export.report_template import write_report_template

            frozen = prepare.current_result
            phase_code = str(prepare.phase_code or frozen.profile_code).upper()
            temp_code = str(prepare.temperature_code or "HT").upper()
            temp_label = prepare.temperature_labels[temp_code]
            with tempfile.TemporaryDirectory() as td:
                report = Path(td) / "manual_cursor_report.xlsx"
                workbook = build_mcu2506_workbook(frozen)
                worksheet = workbook.active
                worksheet.title = f"{phase_code[0]}相_双脉冲数据"
                worksheet.merge_cells(
                    start_row=DATA_ROW,
                    start_column=1,
                    end_row=DATA_ROW + 3,
                    end_column=1,
                )
                worksheet.merge_cells(
                    start_row=DATA_ROW,
                    start_column=2,
                    end_row=DATA_ROW + 3,
                    end_column=2,
                )
                worksheet.cell(DATA_ROW, 1, phase_code)
                worksheet.cell(DATA_ROW, 2, temp_label)
                workbook.save(report)

                summary = write_report_template(
                    frozen,
                    report,
                    temperature_code=temp_code,
                    temperature_labels=prepare.temperature_labels,
                    phase_code=phase_code,
                    report_conditions=prepare.report_conditions,
                )
                saved = load_workbook(report, data_only=True)[worksheet.title]
                self.assertEqual(summary.data_row, DATA_ROW)
                self.assertEqual(
                    saved.cell(summary.data_row, COL_ON["eon"]).value,
                    round(adjusted_eon, 3),
                )
                self.assertEqual(
                    saved.cell(summary.data_row, COL_OFF["eoff"]).value,
                    round(untouched_eoff, 3),
                )
        finally:
            win._report_prepare_tasks.clear()
            win._load_pool = original_pool
            win.close()

    def test_manual_measurements_do_not_leak_between_pulse_pairs(self) -> None:
        win, original = self._load_real_window()
        try:
            key = ("开通", "Eon")
            win.result.turn_on.eon = 123.456
            win._manual_energy[key] = (1.0, 2.0, 3.0, 4.0)
            win._touch_manual_waveform_source()

            next_pair = deepcopy(original)
            next_pair.detected_pulse_count = 3
            next_pair.off_pulse_index = 2
            next_pair.on_pulse_index = 3
            next_pair.turn_on.eon = 456.789
            with patch(
                "dpt_extractor.gui.main_window.run_extraction",
                return_value=next_pair,
            ):
                win._recalculate(reset_manual=False)

            self.assertAlmostEqual(win.result.turn_on.eon, 456.789, places=9)
            self.assertEqual(win._manual_energy, {})
            self.assertEqual(win._manual_waveform_source, "")
            self.assertIsNone(win._manual_pulse_pair)

            first_pair = deepcopy(original)
            first_pair.detected_pulse_count = 3
            first_pair.turn_on.eon = 789.123
            with patch(
                "dpt_extractor.gui.main_window.run_extraction",
                return_value=first_pair,
            ):
                win._recalculate(reset_manual=False)

            self.assertAlmostEqual(win.result.turn_on.eon, 789.123, places=9)
            self.assertEqual(win._manual_energy, {})
        finally:
            win.close()

    def test_manual_didt_restores_value_and_report_availability_together(self) -> None:
        win, original = self._load_real_window()
        try:
            keys = {
                "off": ("关断过程", "di/dt"),
                "on": ("开通", "di/dt"),
                "rr": ("反向恢复", "di/dt"),
            }
            for key in keys.values():
                win._manual_didt[key] = (1.0, 2.0, 3.0, 4.0)
            win._touch_manual_waveform_source()

            previous = deepcopy(original)
            previous.turn_off.didt = 11.0
            previous.turn_on.didt = 12.0
            previous.reverse_recovery.didt_irr = 13.0
            previous.unavailable_metrics.update(
                {
                    keys["off"],
                    ("关断过程", "Ls_off"),
                    keys["rr"],
                }
            )
            previous.unavailable_metrics.discard(keys["on"])
            previous.unavailable_metrics.discard(("开通", "Ls_on"))

            current = deepcopy(original)
            current.turn_off.didt = 101.0
            current.turn_on.didt = 102.0
            current.reverse_recovery.didt_irr = 103.0
            current.unavailable_metrics.discard(keys["off"])
            current.unavailable_metrics.discard(("关断过程", "Ls_off"))
            current.unavailable_metrics.discard(keys["rr"])
            current.unavailable_metrics.update(
                {keys["on"], ("开通", "Ls_on")}
            )
            win.result = current

            win._restore_manual_result_values_after_recalculation(previous)

            self.assertEqual(win.result.turn_off.didt, 11.0)
            self.assertEqual(win.result.turn_on.didt, 12.0)
            self.assertEqual(win.result.reverse_recovery.didt_irr, 13.0)
            self.assertTrue(win.result.is_metric_unavailable(*keys["off"]))
            self.assertTrue(
                win.result.is_metric_unavailable("关断过程", "Ls_off")
            )
            self.assertTrue(win.result.is_metric_unavailable(*keys["rr"]))
            self.assertFalse(win.result.is_metric_unavailable(*keys["on"]))
            self.assertFalse(win.result.is_metric_unavailable("开通", "Ls_on"))
            self.assertEqual(win.result.turn_off.ls_off, 0.0)
            self.assertEqual(win.result.turn_on.ls_on, previous.turn_on.ls_on)
        finally:
            win.close()

    def test_real_capture_uses_frozen_result_and_restores_later_page(self) -> None:
        win, original_result = self._load_real_window()
        snapshot_result = deepcopy(original_result)
        snapshot_result.turn_off.eoff = 314.159
        prepare = _ReportPrepareTask(
            1,
            win.bundle,
            win.profile,
            win.cfg,
            snapshot_result,
            temperature_code="LT",
            temperature_labels={"LT": "-40℃"},
            phase_code="UH",
            slope_ranges=win._slope_ranges,
            manual_state=win._snapshot_report_manual_state(),
        )

        later_result = deepcopy(original_result)
        later_result.profile_code = "VH"
        later_result.turn_off.eoff = -1.0
        win.result = later_result
        win.result_table.set_result(later_result)
        later_table_text = next(
            win.result_table.table.item(row, 4).text()
            for row, meta in enumerate(win.result_table._row_meta)
            if meta == ("关断过程", "Eoff")
        )
        assert win.bundle is not None
        win.wave_plot.plot_waveforms(win.bundle, win.profile, later_result)
        win._on_value_clicked("反向恢复", "di/dt")
        later_active_metric = win.result_table._active_metric
        later_active_slope = win._active_slope_param
        hidden_key = next(
            key
            for key in win.wave_plot._trace_items
            if key != win.wave_plot._readout_channel()
        )
        win.wave_plot._toggle_channel_visibility(hidden_key)
        win.wave_plot._remember_user_x_scale(0.35)
        later_display_state = win.wave_plot.snapshot_report_display_state()

        observed: list[tuple[str, str, str, float, str]] = []
        completed: dict[str, object] = {}
        original_click = win._on_value_clicked

        def observe_snapshot(section: str, name: str) -> None:
            assert win.result is not None
            observed.append(
                (
                    section,
                    name,
                    win.result.profile_code,
                    win.result.turn_off.eoff,
                    win.bundle.meta.source_path if win.bundle is not None else "",
                )
            )
            original_click(section, name)

        def record_write_start(tempdir, images, results, **kwargs) -> None:
            completed.update(
                tempdir=tempdir,
                images=dict(images),
                results=list(results),
                kwargs=dict(kwargs),
            )

        tempdir = tempfile.TemporaryDirectory()
        try:
            self.assertTrue(win._try_begin_report_operation())
            win._report_request_id += 1
            with (
                patch.object(
                    win,
                    "_report_image_params",
                    return_value=(("关断过程", "Eoff"),),
                ),
                patch.object(win, "_on_value_clicked", side_effect=observe_snapshot),
                patch.object(
                    win,
                    "_start_report_write_task",
                    side_effect=record_write_start,
                ),
            ):
                win._start_report_capture_sequence(
                    tempdir,
                    [snapshot_result],
                    request_id=win._report_request_id,
                    temperature_code="LT",
                    temperature_labels={"LT": "-40℃"},
                    phase_code="UH",
                    image_result_index=0,
                    capture_bundle=prepare.bundle,
                    capture_profile=prepare.profile,
                    capture_cfg=prepare.cfg,
                    capture_result=prepare.current_result,
                    capture_slope_ranges=prepare.slope_ranges,
                    capture_manual_state=prepare.manual_state,
                )
                self._drain_events_until(lambda: bool(completed))

            self.assertEqual(len(observed), 3)
            self.assertEqual(observed[0][:2], ("反向恢复", "di/dt"))
            self.assertEqual(observed[1][:2], ("关断过程", "Eoff"))
            for snapshot_observation in observed[:2]:
                self.assertEqual(snapshot_observation[2], snapshot_result.profile_code)
                self.assertEqual(snapshot_observation[3], 314.159)
                self.assertEqual(snapshot_observation[4], snapshot_result.source_path)
            self.assertEqual(observed[2][:2], ("反向恢复", "di/dt"))
            self.assertEqual(observed[2][2], "VH")
            self.assertEqual(observed[2][3], -1.0)
            images = completed["images"]
            self.assertIsInstance(images, dict)
            assert isinstance(images, dict)
            image_path = Path(images[("关断过程", "Eoff")])
            self.assertTrue(image_path.is_file())
            self.assertGreater(image_path.stat().st_size, 0)

            self.assertIs(win.result, later_result)
            self.assertEqual(win.result.profile_code, "VH")
            self.assertEqual(win.result.turn_off.eoff, -1.0)
            self.assertEqual(win.result_table._active_metric, later_active_metric)
            self.assertEqual(win._active_slope_param, later_active_slope)
            self.assertEqual(
                win.wave_plot.snapshot_report_display_state(),
                later_display_state,
            )
            self.assertFalse(win.wave_plot._trace_items[hidden_key].isVisible())
            restored_table_text = next(
                win.result_table.table.item(row, 4).text()
                for row, meta in enumerate(win.result_table._row_meta)
                if meta == ("关断过程", "Eoff")
            )
            self.assertEqual(restored_table_text, later_table_text)
        finally:
            win._report_capture_state = None
            win._release_report_operation()
            tempdir.cleanup()
            win.close()

    def test_template_selection_rejects_current_output_file(self) -> None:
        win = MainWindow()
        try:
            with tempfile.TemporaryDirectory() as tmp:
                same = Path(tmp) / "same.xlsx"
                same.write_bytes(b"xlsx")
                original_template = Path(tmp) / "original.xlsx"
                win._report_template_source_path = original_template
                win._report_output_path = same
                with (
                    patch.object(
                        QFileDialog,
                        "getOpenFileName",
                        return_value=(str(same), "Excel 报告 (*.xlsx)"),
                    ),
                    patch(
                        "dpt_extractor.gui.main_window.set_report_template_source_path"
                    ) as persist,
                    patch.object(QMessageBox, "warning") as warning,
                ):
                    win._select_report_template()
                self.assertEqual(win._report_template_source_path, original_template)
                persist.assert_not_called()
                warning.assert_called_once()
        finally:
            win.close()

    def test_existing_output_equal_to_template_is_rejected_before_write(self) -> None:
        win = MainWindow()
        try:
            with tempfile.TemporaryDirectory() as tmp:
                same = Path(tmp) / "same.xlsx"
                same.write_bytes(b"xlsx")
                win._report_template_source_path = same
                win._report_output_path = same
                with (
                    patch.object(QMessageBox, "warning") as warning,
                    patch(
                        "dpt_extractor.gui.main_window.copy_report_template"
                    ) as copy_template,
                ):
                    self.assertFalse(win._ensure_report_output_file())
                warning.assert_called_once()
                copy_template.assert_not_called()
                self.assertEqual(same.read_bytes(), b"xlsx")
        finally:
            win.close()

    def test_permission_error_releases_guard_and_immediate_retry_succeeds(self) -> None:
        win = MainWindow()
        original_pool = win._load_pool
        win._load_pool = _ImmediatePool()  # type: ignore[assignment]
        summary = ReportWriteSummary(
            report_path=Path("report.xlsx"),
            data_sheet="U相_双脉冲数据",
            data_row=13,
        )
        calls = 0

        def fail_then_succeed(*_args, **_kwargs):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise PermissionError("workbook is locked")
            return summary

        try:
            with tempfile.TemporaryDirectory() as tmp:
                win._report_output_path = Path(tmp) / "report.xlsx"
                with (
                    patch(
                        "dpt_extractor.gui.main_window.write_report_template",
                        side_effect=fail_then_succeed,
                    ),
                    patch(
                        "dpt_extractor.gui.main_window.set_report_output_path"
                    ),
                    patch("dpt_extractor.gui.main_window.set_last_export_path"),
                    patch(
                        "dpt_extractor.gui.main_window.QMessageBox.critical"
                    ) as critical,
                    patch(
                        "dpt_extractor.gui.main_window.QMessageBox.information"
                    ) as information,
                ):
                    self.assertTrue(win._try_begin_report_operation())
                    win._begin_report_progress(
                        REPORT_PROGRESS_TOTAL,
                        "正在打开并写入 Excel...",
                    )
                    win._start_report_write_task(
                        tempfile.TemporaryDirectory(dir=tmp),
                        {},
                        [ExtractResult()],
                    )
                    self.assertEqual(calls, 1)
                    self.assertFalse(win._report_operation_active)
                    self.assertTrue(win.btn_write_report.isEnabled())
                    self.assertEqual(win.report_progress.detail_text(), "失败")
                    critical.assert_called_once()
                    self.assertIn("Excel", critical.call_args.args[2])

                    # No restart or stale-task cleanup should be necessary before
                    # retrying as soon as the user closes the locked workbook.
                    self.assertTrue(win._try_begin_report_operation())
                    win._begin_report_progress(
                        REPORT_PROGRESS_TOTAL,
                        "正在打开并写入 Excel...",
                    )
                    win._start_report_write_task(
                        tempfile.TemporaryDirectory(dir=tmp),
                        {},
                        [ExtractResult()],
                    )
                    self.assertEqual(calls, 2)
                    self.assertFalse(win._report_operation_active)
                    self.assertTrue(win.btn_write_report.isEnabled())
                    self.assertEqual(win.report_progress.percent_text(), "100.0%")
                    self.assertEqual(win.report_progress.eta_text(), "0 ms")
                    information.assert_called_once()
        finally:
            win._report_tasks.clear()
            win._release_report_operation()
            win._load_pool = original_pool
            win.close()


if __name__ == "__main__":
    unittest.main()
