from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from dpt_extractor.tests.sample_paths import SAMPLE_ROOT

DL_UH = SAMPLE_ROOT / "tss格式" / "KSU2506" / "DCU" / "DL" / "LT" / "UH_480V_000.tss"
DL_UL = SAMPLE_ROOT / "tss格式" / "KSU2506" / "DCU" / "DL" / "LT" / "UL_480V_000.tss"
DDD_UH = (
    SAMPLE_ROOT
    / "tss格式"
    / "KSU2577"
    / "SSM1R7PB12B3DTFMMSPP25M4CF0016"
    / "DDD"
    / "HT"
    / "UH_750V_000.tss"
)
NED34_SHORT_VH_750 = (
    SAMPLE_ROOT
    / "likangkang"
    / "NED34jixian"
    / "short"
    / "750v-vh-short-25c_000.tss"
)
DDD_RT_VH = (
    SAMPLE_ROOT
    / "tss格式"
    / "KSU2577"
    / "SSM1R7PB12B3DTFMMSPP25M4CF0016"
    / "DDD"
    / "RT"
    / "VH_750V_000.tss"
)
SONG_DL_LT_UL = (
    SAMPLE_ROOT
    / "songzhenxi"
    / "KSU2577"
    / "07CF2C1000 20260506"
    / "DL"
    / "LT"
    / "UL_750V_000.tss"
)
SONG_DL_LT_VL = SONG_DL_LT_UL.with_name("VL_750V_000.tss")
SONG_DDD_LT_VH = (
    SAMPLE_ROOT
    / "songzhenxi"
    / "KSU2577"
    / "SSM1R7PB12B3DTFMMSPP25M4CF0016"
    / "DDD"
    / "LT"
    / "VH_750V_000.tss"
)
SONG_DDD_HT_UH = SONG_DDD_LT_VH.parent.parent / "HT" / "UH_750V_000.tss"
SONG_DDD_LT_UL = SONG_DDD_LT_VH.with_name("UL_750V_000.tss")
SONG_LCG_DDD_RT_UL = (
    SAMPLE_ROOT
    / "songzhenxi"
    / "KSU2577"
    / "LCG660FF120I3A2-G1LEP202510090002"
    / "DDD"
    / "RT"
    / "UL_750V_000.tss"
)
SONG_LCG_DDD_RT_UH = SONG_LCG_DDD_RT_UL.with_name("UH_750V_000.tss")


class TestShortCircuitLabelMapping(unittest.TestCase):
    def test_infers_non_default_short_circuit_labels(self):
        import numpy as np

        from dpt_extractor.models.channel_mapping import infer_short_circuit_mapping_from_bundle
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle

        meta = TekMetadata(
            channel_labels={
                "CH4": "H-Vge",
                "CH5": "H-Vce",
                "CH2": "Ic",
                "CH1": "L-Vce",
                "CH6": "L-Vge",
            }
        )
        bundle = WaveformBundle(
            t=np.linspace(0.0, 1e-6, 8),
            channels={f"CH{i}": np.zeros(8) for i in range(1, 7)},
            meta=meta,
        )

        mapping = infer_short_circuit_mapping_from_bundle(bundle, "upper")
        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual(mapping.vge, "CH4")
        self.assertEqual(mapping.vce, "CH5")
        self.assertEqual(mapping.ic, "CH2")
        self.assertEqual(mapping.v_diode, "CH1")
        self.assertEqual(mapping.vge_other, "CH6")
        self.assertFalse(mapping.ic_from_sum_irr_il)

    def test_short_circuit_mapping_allows_missing_other_channels(self):
        import numpy as np

        from dpt_extractor.models.channel_mapping import infer_short_circuit_mapping_from_bundle
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle

        meta = TekMetadata(
            channel_labels={
                "CH4": "H-Vge",
                "CH5": "H-Vce",
                "CH2": "Ic",
            }
        )
        bundle = WaveformBundle(
            t=np.linspace(0.0, 1e-6, 8),
            channels={
                "CH2": np.zeros(8),
                "CH4": np.zeros(8),
                "CH5": np.zeros(8),
            },
            meta=meta,
        )

        mapping = infer_short_circuit_mapping_from_bundle(bundle, "upper")

        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual(mapping.vge, "CH4")
        self.assertEqual(mapping.vce, "CH5")
        self.assertEqual(mapping.ic, "CH2")
        self.assertEqual(mapping.v_diode, "")
        self.assertEqual(mapping.vge_other, "")
        self.assertEqual(mapping.vdesat, "")
        self.assertFalse(mapping.ic_from_sum_irr_il)

    def test_short_circuit_energy_math_channel_uses_display_inversion(self):
        import numpy as np

        from dpt_extractor.models.bridge_profile import as_short_circuit_profile, make_profile
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_energy_peak_value,
            short_circuit_energy_value,
        )

        raw_energy = np.array([-3.0, -2.0, -1.0], dtype=np.float64)
        profile = as_short_circuit_profile(make_profile("U", "upper"))
        bundle = WaveformBundle(
            t=np.arange(raw_energy.size, dtype=np.float64),
            channels={
                "CH2": np.array([10.0, 10.0, 10.0], dtype=np.float64),
                "CH3": np.array([1.0, 2.0, 3.0], dtype=np.float64),
                "MATH1": raw_energy,
            },
            meta=TekMetadata(channel_display_inversions={"MATH1"}),
        )

        value, source = short_circuit_energy_value(
            bundle,
            profile,
            0,
            2,
            math_channel="MATH1",
        )
        peak, peak_source = short_circuit_energy_peak_value(
            bundle,
            profile,
            0,
            2,
            math_channel="MATH1",
        )

        self.assertEqual(source, "MATH1")
        self.assertAlmostEqual(value, 40.0)
        self.assertEqual(peak_source, "MATH1")
        self.assertAlmostEqual(peak, 3.0)


class TestShortCircuitSyntheticRegressions(unittest.TestCase):
    @staticmethod
    def _short_bundle(*, flat_gate: bool = False, nan_vce: bool = False):
        import numpy as np

        from dpt_extractor.models.bridge_profile import make_short_circuit_profile
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle

        dt = 10e-9
        count = 200
        t = np.arange(count, dtype=np.float64) * dt
        gate = np.full(count, 15.0 if flat_gate else 0.0, dtype=np.float64)
        if not flat_gate:
            gate[50:130] = 15.0
        current = np.zeros(count, dtype=np.float64)
        current[50:130] = 100.0
        vce = np.full(count, 600.0, dtype=np.float64)
        if nan_vce:
            vce[80] = np.nan
        bundle = WaveformBundle(
            t=t,
            channels={
                "CH1": gate,
                "CH2": vce,
                "CH3": current,
                "CH5": np.full(count, 300.0, dtype=np.float64),
                "CH6": np.zeros(count, dtype=np.float64),
            },
            meta=TekMetadata(sample_interval=dt),
        )
        return bundle, make_short_circuit_profile("U", "upper")

    def test_tsc_symmetric_custom_range_normalization(self):
        from dpt_extractor.models.results import (
            format_short_circuit_tsc_symmetric_range,
            short_circuit_tsc_symmetric_percent,
        )
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_tsc_range_percentages,
        )

        self.assertEqual(
            format_short_circuit_tsc_symmetric_range(25.0),
            "25%-25%",
        )
        self.assertEqual(
            format_short_circuit_tsc_symmetric_range(12.5),
            "12.5%-12.5%",
        )
        self.assertEqual(short_circuit_tsc_symmetric_percent("25%-25%"), 25.0)
        self.assertIsNone(short_circuit_tsc_symmetric_percent("20%-30%"))
        self.assertIsNone(short_circuit_tsc_symmetric_percent("100%-100%"))
        self.assertEqual(
            short_circuit_tsc_range_percentages("25%-25%"),
            (25.0, 25.0, "25%-25%"),
        )
        self.assertEqual(
            short_circuit_tsc_range_percentages("0.0%-0.0%"),
            (0.0, 0.0, "0%-0%"),
        )
        self.assertEqual(
            short_circuit_tsc_range_percentages("10.0%-10.0%"),
            (10.0, 10.0, "10%-10%"),
        )
        self.assertEqual(
            short_circuit_tsc_range_percentages("20%-30%"),
            (0.0, 0.0, "0%-0%"),
        )

    def test_raw_fall_crossing_uses_smooth_anchor_and_rejects_two_point_glitch(
        self,
    ):
        import numpy as np

        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_cursors,
        )

        dt = 1e-9
        t = np.arange(2000, dtype=np.float64) * dt
        current = np.zeros(2000, dtype=np.float64)
        current[500:1300] = 100.0
        current[900:902] = -20.0

        cursors = short_circuit_current_cursors(
            t,
            current,
            500,
            1300,
            dt,
            smooth_ns=50.0,
        )

        self.assertIsNotNone(cursors)
        assert cursors is not None
        self.assertGreater(cursors.t_b_s, float(t[1200]))
        self.assertAlmostEqual(cursors.t_b_s, float(t[1300]), places=15)
        self.assertAlmostEqual(
            float(np.interp(cursors.t_b_s, t, current)),
            cursors.hb_a,
            places=12,
        )

    def test_missing_raw_ten_percent_pair_returns_none_and_marks_tsc_unavailable(
        self,
    ):
        import numpy as np
        from unittest.mock import patch

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_10
        from dpt_extractor.pipeline import short_circuit_extract as short_extract

        bundle, profile = self._short_bundle()
        raw_crossings = short_extract._raw_pulse_base_crossings

        def reject_percent_crossings(*args, **kwargs):
            level = float(args[2])
            if abs(level) > 1.0:
                return None
            return raw_crossings(*args, **kwargs)

        with patch.object(
            short_extract,
            "_raw_pulse_base_crossings",
            side_effect=reject_percent_crossings,
        ):
            cursors = short_extract.short_circuit_current_percent_cursors(
                bundle.t,
                bundle.get(profile.ic),
                50,
                130,
                bundle.dt,
                percent=10.0,
                smooth_ns=40.0,
            )
        self.assertIsNone(cursors)

        cfg = load_config()
        cfg.short_circuit_tsc_range = SHORT_CIRCUIT_TSC_RANGE_10
        with patch.object(
            short_extract,
            "short_circuit_current_percent_cursors",
            return_value=None,
        ):
            result = short_extract.extract_short_circuit(bundle, profile, cfg)

        self.assertTrue(
            result.is_metric_unavailable("短路过程", "短路时间Tsc")
        )
        self.assertEqual(result.short_circuit.tsc, 0.0)
        self.assertIsNone(result.short_circuit.tsc_start_us)
        self.assertIsNone(result.short_circuit.tsc_end_us)
        self.assertFalse(result.is_metric_unavailable("短路过程", "短路电流Imax"))
        self.assertTrue(np.isfinite(result.short_circuit.ic_max))

    def test_energy_uses_exact_interpolated_endpoints_and_nan_fails_closed(self):
        import numpy as np

        from dpt_extractor.models.bridge_profile import make_short_circuit_profile
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_energy_value,
        )

        profile = make_short_circuit_profile("U", "upper")
        t = np.asarray([0.0, 1.0, 2.0, 3.0], dtype=np.float64)
        channels = {
            "CH2": np.full(4, 3.0, dtype=np.float64),
            "CH3": np.full(4, 2.0, dtype=np.float64),
        }
        bundle = WaveformBundle(
            t=t,
            channels=channels,
            meta=TekMetadata(sample_interval=1.0),
        )
        energy, _source = short_circuit_energy_value(
            bundle,
            profile,
            0,
            3,
            t_a_s=0.25,
            t_b_s=2.75,
        )
        self.assertAlmostEqual(energy, 15.0, places=12)

        invalid_interval, _source = short_circuit_energy_value(
            bundle,
            profile,
            1,
            1,
        )
        self.assertTrue(np.isnan(invalid_interval))

        invalid_channels = dict(channels)
        invalid_channels["CH2"] = np.asarray([3.0, np.nan, 3.0, 3.0])
        invalid_bundle = WaveformBundle(
            t=t,
            channels=invalid_channels,
            meta=TekMetadata(sample_interval=1.0),
        )
        invalid_energy, _source = short_circuit_energy_value(
            invalid_bundle,
            profile,
            0,
            3,
            t_a_s=0.25,
            t_b_s=2.75,
        )
        self.assertTrue(np.isnan(invalid_energy))

    def test_invalid_energy_and_missing_gate_crossings_use_current_vpeak_window(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.pipeline.short_circuit_extract import (
            extract_short_circuit,
            short_circuit_current_cursors,
            short_circuit_vpeak_cursors,
        )

        invalid_bundle, profile = self._short_bundle(nan_vce=True)
        invalid_result = extract_short_circuit(invalid_bundle, profile, load_config())
        self.assertTrue(
            invalid_result.is_metric_unavailable("短路过程", "短路能量Esc_本管")
        )
        self.assertTrue(np.isnan(invalid_result.short_circuit.esc_dut))

        flat_gate_bundle, profile = self._short_bundle(flat_gate=True)
        flat_gate_bundle.get(profile.vce)[20] = 1200.0
        flat_gate_bundle.get(profile.vce)[80] = 850.0
        flat_gate_bundle.get(profile.v_diode)[20] = 900.0
        flat_gate_bundle.get(profile.v_diode)[90] = 450.0
        flat_gate_result = extract_short_circuit(
            flat_gate_bundle,
            profile,
            load_config(),
        )
        current_cursors = short_circuit_current_cursors(
            flat_gate_bundle.t,
            flat_gate_bundle.get(profile.ic),
            0,
            flat_gate_bundle.n - 1,
            flat_gate_bundle.dt,
        )
        self.assertIsNotNone(current_cursors)
        assert current_cursors is not None
        dut_cursors = short_circuit_vpeak_cursors(
            flat_gate_bundle.t,
            flat_gate_bundle.get(profile.vge),
            flat_gate_bundle.get(profile.vce),
            0,
            flat_gate_bundle.n - 1,
            flat_gate_bundle.dt,
            current_cursors=current_cursors,
        )
        self.assertIsNotNone(dut_cursors)
        assert dut_cursors is not None
        self.assertEqual(dut_cursors.boundary_source, "ic")
        self.assertAlmostEqual(dut_cursors.t_a_s, current_cursors.t_a_s, places=15)
        self.assertAlmostEqual(dut_cursors.t_b_s, current_cursors.t_b_s, places=15)
        self.assertEqual(dut_cursors.ha_a, 850.0)
        self.assertEqual(flat_gate_result.short_circuit.vpeak_dut, 850.0)
        self.assertEqual(flat_gate_result.short_circuit.vpeak_other, 450.0)
        self.assertFalse(
            flat_gate_result.is_metric_unavailable("短路过程", "应力Vpeak_本管")
        )
        self.assertFalse(
            flat_gate_result.is_metric_unavailable("短路过程", "应力Vpeak_对管")
        )

    def test_flat_gate_vpeak_gui_binds_to_current_ab_window(self):
        from unittest.mock import patch

        from PyQt6.QtCore import QSettings
        from PyQt6.QtWidgets import QApplication

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.short_circuit_extract import extract_short_circuit

        app = QApplication.instance() or QApplication(sys.argv)
        bundle, profile = self._short_bundle(flat_gate=True)
        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        result = extract_short_circuit(bundle, profile, cfg)

        with tempfile.TemporaryDirectory() as tmp:
            settings_path = Path(tmp) / "DPTExtractor.ini"

            def settings_factory() -> QSettings:
                return QSettings(str(settings_path), QSettings.Format.IniFormat)

            with patch(
                "dpt_extractor.gui.main_window._app_settings",
                settings_factory,
            ):
                win = MainWindow()
                try:
                    win.cfg = cfg
                    win._apply_test_mode_ui()
                    win.bundle = bundle
                    win.profile = profile
                    win.result = result
                    win.result_table.set_result(result)
                    win.wave_plot.plot_waveforms(bundle, profile, result)
                    app.processEvents()

                    current = win._short_circuit_ic_default_context()
                    self.assertIsNotNone(current)
                    assert current is not None
                    for name, voltage_channel in (
                        ("应力Vpeak_本管", win.profile.vce),
                        ("应力Vpeak_对管", win.profile.v_diode),
                    ):
                        with self.subTest(name=name):
                            context = win._short_circuit_vpeak_default_context(
                                voltage_channel,
                                gate_channel=win.profile.vge,
                            )
                            self.assertIsNotNone(context)
                            assert context is not None
                            self.assertEqual(context.boundary_source, "ic")
                            self.assertAlmostEqual(
                                context.t_a_s, current.t_a_s, places=15
                            )
                            self.assertAlmostEqual(
                                context.t_b_s, current.t_b_s, places=15
                            )
                            self.assertEqual(
                                win._cursor_endpoint_channels_for_param(
                                    "短路过程", name
                                ),
                                ("ic", "ic"),
                            )

                            win._enable_generic_parameter_interaction(
                                "短路过程", name
                            )
                            app.processEvents()
                            self.assertEqual(win.wave_plot._interval_a_channel, "ic")
                            self.assertEqual(win.wave_plot._interval_b_channel, "ic")
                            self.assertEqual(win.wave_plot._interval_hb_channel, "ic")
                finally:
                    win.close()

    def test_empty_short_waveform_marks_every_metric_unavailable(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.models.bridge_profile import make_short_circuit_profile
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle
        from dpt_extractor.pipeline.short_circuit_extract import extract_short_circuit

        bundle = WaveformBundle(
            t=np.asarray([], dtype=np.float64),
            channels={},
            meta=TekMetadata(sample_interval=1e-9),
        )
        result = extract_short_circuit(
            bundle,
            make_short_circuit_profile("U", "upper"),
            load_config(),
        )

        for name in (
            "短路电流Imax",
            "短路时间Tsc",
            "短路能量Esc_本管",
            "应力Vpeak_本管",
            "短路能量Esc_对管",
            "应力Vpeak_对管",
            "Desat动作时间",
        ):
            self.assertTrue(result.is_metric_unavailable("短路过程", name), name)


@unittest.skipUnless(
    SONG_LCG_DDD_RT_UH.exists() and SONG_LCG_DDD_RT_UL.exists(),
    "songzhenxi short-circuit upper/lower samples missing",
)
class TestShortCircuitCustomTscRange(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        from PyQt6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication(sys.argv)

    def setUp(self):
        from dpt_extractor.gui.main_window import (
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            _app_settings,
        )
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_DEFAULT

        self._tsc_settings = _app_settings()
        self._old_tsc_range = self._tsc_settings.value(
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            None,
        )
        self._tsc_settings.setValue(
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            SHORT_CIRCUIT_TSC_RANGE_DEFAULT,
        )

    def tearDown(self):
        from dpt_extractor.gui.main_window import (
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
        )

        if self._old_tsc_range is None:
            self._tsc_settings.remove(SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY)
        else:
            self._tsc_settings.setValue(
                SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
                self._old_tsc_range,
            )

    def test_custom_range_covers_upper_and_lower_without_linkage(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_percent_cursors,
        )

        for path, expected_profile in (
            (SONG_LCG_DDD_RT_UH, "UH"),
            (SONG_LCG_DDD_RT_UL, "UL"),
        ):
            with self.subTest(path=path.name):
                bundle = load_waveform(path)
                profile = guess_profile_from_path(path)

                cfg_default = load_config()
                cfg_default.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                default_result = run_extraction(bundle, profile, cfg_default)

                cfg_custom = load_config()
                cfg_custom.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                cfg_custom.short_circuit_tsc_range = "25%-25%"
                custom_result = run_extraction(bundle, profile, cfg_custom)
                custom = custom_result.short_circuit

                self.assertEqual(custom_result.profile_code, expected_profile)
                self.assertEqual(custom.tsc_range, "25%-25%")
                self.assertGreater(custom.tsc, 0.0)
                self.assertLess(custom.tsc, default_result.short_circuit.tsc)
                self.assertIsNotNone(custom.tsc_start_us)
                self.assertIsNotNone(custom.tsc_end_us)
                assert custom_result.segments is not None
                i0, i1 = custom_result.segments.turn_off
                current = bundle_total_current(
                    bundle,
                    as_short_circuit_profile(profile),
                )
                cursors = short_circuit_current_percent_cursors(
                    bundle.t,
                    current,
                    i0,
                    i1,
                    bundle.dt,
                    smooth_ns=cfg_custom.smoothing.detect_window_ns,
                    percent=25.0,
                )
                if cursors is None:
                    cursors = short_circuit_current_percent_cursors(
                        bundle.t,
                        current,
                        0,
                        len(bundle.t) - 1,
                        bundle.dt,
                        smooth_ns=cfg_custom.smoothing.detect_window_ns,
                        percent=25.0,
                    )
                self.assertIsNotNone(cursors)
                assert cursors is not None
                self.assertAlmostEqual(
                    custom.tsc_start_us,
                    cursors.t_a_s * 1e6,
                    delta=1e-6,
                )
                self.assertAlmostEqual(
                    custom.tsc_end_us,
                    cursors.t_b_s * 1e6,
                    delta=1e-6,
                )
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_a_s, bundle.t, current)),
                    cursors.hb_a,
                    delta=1e-3,
                )
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_b_s, bundle.t, current)),
                    cursors.hb_a,
                    delta=1e-3,
                )
                for attr in (
                    "ic_max",
                    "esc_dut",
                    "esc_other",
                    "vpeak_dut",
                    "vpeak_other",
                    "desat_time",
                ):
                    expected_value = getattr(default_result.short_circuit, attr)
                    actual_value = getattr(custom, attr)
                    if expected_value is None:
                        self.assertIsNone(actual_value, attr)
                        continue
                    expected = float(expected_value)
                    actual = float(actual_value)
                    if np.isnan(expected):
                        self.assertTrue(np.isnan(actual), attr)
                    else:
                        self.assertAlmostEqual(
                            actual,
                            expected,
                            delta=1e-6,
                            msg=attr,
                        )

    def test_custom_range_persists_across_restart_and_file_load(self):
        from dpt_extractor.gui.main_window import (
            MainWindow,
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
        )
        from dpt_extractor.models.test_mode import TestMode

        custom_range = "25%-25%"
        win = MainWindow()
        try:
            win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
            win._apply_test_mode_ui()
            win._load_file(str(SONG_LCG_DDD_RT_UH), background=False)
            self.app.processEvents()
            tsc_key = ("短路过程", "短路时间Tsc")
            win._manual_intervals[tsc_key] = (1.0, 2.0)
            win._manual_short_current[tsc_key] = (1.0, 2.0, 3.0, 4.0)
            win._on_short_circuit_tsc_range_changed(custom_range)
            self.app.processEvents()
            self.assertNotIn(tsc_key, win._manual_intervals)
            self.assertNotIn(tsc_key, win._manual_short_current)
            self.assertEqual(
                self._tsc_settings.value(SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY),
                custom_range,
            )
        finally:
            win.close()

        win2 = MainWindow()
        try:
            self.assertEqual(win2.cfg.short_circuit_tsc_range, custom_range)
            win2.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
            win2._apply_test_mode_ui()
            win2._load_file(str(SONG_LCG_DDD_RT_UH), background=False)
            self.app.processEvents()
            self.assertEqual(win2.cfg.short_circuit_tsc_range, custom_range)
            assert win2.result is not None
            self.assertEqual(win2.result.short_circuit.tsc_range, custom_range)
            self.assertEqual(
                win2._load_cfg_for_new_file().short_circuit_tsc_range,
                custom_range,
            )
        finally:
            win2.close()

    def test_range_dialog_uses_one_symmetric_custom_value(self):
        from dpt_extractor.gui.result_table import ShortCircuitTscRangeDialog
        from dpt_extractor.models.results import (
            SHORT_CIRCUIT_TSC_RANGE_10,
            SHORT_CIRCUIT_TSC_RANGE_CUSTOM,
        )

        dialog = ShortCircuitTscRangeDialog(current="25%-25%")
        try:
            self.assertEqual(
                dialog.range_selector.currentText(),
                SHORT_CIRCUIT_TSC_RANGE_CUSTOM,
            )
            self.assertAlmostEqual(dialog.custom_percent.value(), 25.0)
            self.assertFalse(dialog.custom_percent.isHidden())
            self.assertEqual(dialog.range_label(), "25%-25%")

            dialog.range_selector.setCurrentText(SHORT_CIRCUIT_TSC_RANGE_10)
            self.assertTrue(dialog.custom_percent.isHidden())
            self.assertEqual(dialog.range_label(), SHORT_CIRCUIT_TSC_RANGE_10)
        finally:
            dialog.close()


@unittest.skipUnless(DL_UH.exists() and DL_UL.exists(), "short-circuit DL samples missing")
class TestShortCircuitExtract(unittest.TestCase):
    def _extract(self, path: Path):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        bundle = load_waveform(path)
        profile = guess_profile_from_path(path)
        return bundle, run_extraction(bundle, profile, cfg)

    def test_extracts_upper_short_circuit_window_and_values(self):
        bundle, result = self._extract(DL_UH)
        sc = result.short_circuit
        i0, i1 = result.segments.turn_off
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_cursors,
            short_circuit_energy_value,
            short_circuit_vpeak_cursors,
        )

        cfg = load_config()
        profile = as_short_circuit_profile(guess_profile_from_path(DL_UH))
        cursors = short_circuit_current_cursors(
            bundle.t,
            bundle_total_current(bundle, profile),
            i0,
            i1,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
        )
        self.assertIsNotNone(cursors)
        assert cursors is not None
        dut_vce_cursors = short_circuit_vpeak_cursors(
            bundle.t,
            bundle.get(profile.vge),
            bundle.get(profile.vce),
            i0,
            i1,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
        )
        other_vce_cursors = short_circuit_vpeak_cursors(
            bundle.t,
            bundle.get(profile.vge),
            bundle.get(profile.v_diode),
            i0,
            i1,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
        )
        self.assertIsNotNone(dut_vce_cursors)
        self.assertIsNotNone(other_vce_cursors)
        assert dut_vce_cursors is not None
        assert other_vce_cursors is not None
        self.assertEqual(dut_vce_cursors.boundary_source, "vge")
        self.assertEqual(other_vce_cursors.boundary_source, "vge")
        esc_dut_expected, _ = short_circuit_energy_value(
            bundle,
            profile,
            cursors.i0,
            cursors.i1,
            other=False,
            t_a_s=cursors.t_a_s,
            t_b_s=cursors.t_b_s,
        )
        esc_other_expected, _ = short_circuit_energy_value(
            bundle,
            profile,
            cursors.i0,
            cursors.i1,
            other=True,
            t_a_s=cursors.t_a_s,
            t_b_s=cursors.t_b_s,
        )

        self.assertTrue(result.short_circuit_mode)
        self.assertEqual(result.profile_code, "UH")
        self.assertGreater(sc.ic_max, 3000.0)
        self.assertGreater(sc.esc_dut, 1.0)
        self.assertGreater(sc.vpeak_dut, sc.vpeak_other)
        self.assertGreater(sc.tsc, 1.0)
        self.assertLess(sc.tsc, 5.0)
        self.assertAlmostEqual(
            sc.ic_max,
            float(bundle.channels["CH3"][cursors.i0 : cursors.i1 + 1].max()),
            delta=1e-6,
        )
        self.assertAlmostEqual(
            sc.tsc,
            float((cursors.t_b_s - cursors.t_a_s) * 1e6),
            delta=1e-6,
        )
        self.assertIsNotNone(sc.tsc_start_us)
        self.assertIsNotNone(sc.tsc_end_us)
        assert sc.tsc_start_us is not None and sc.tsc_end_us is not None
        self.assertAlmostEqual(sc.tsc_start_us, cursors.t_a_s * 1e6, delta=1e-6)
        self.assertAlmostEqual(sc.tsc_end_us, cursors.t_b_s * 1e6, delta=1e-6)
        self.assertAlmostEqual(sc.tsc, sc.tsc_end_us - sc.tsc_start_us, delta=1e-6)
        self.assertEqual(sc.tsc_range, "0%-0%")
        self.assertAlmostEqual(sc.esc_dut, esc_dut_expected, delta=1e-6)
        self.assertAlmostEqual(sc.esc_other, esc_other_expected, delta=1e-6)
        self.assertAlmostEqual(sc.vpeak_dut, dut_vce_cursors.ha_a, delta=1e-6)
        self.assertAlmostEqual(sc.vpeak_other, other_vce_cursors.ha_a, delta=1e-6)

    def test_non_default_mapping_without_other_channels_keeps_dut_metrics(self):
        from openpyxl import load_workbook

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.short_circuit_layout import (
            COL_DESAT,
            COL_ESC_DUT,
            COL_ESC_OTHER,
            COL_ICMAX,
            COL_PHASE,
            COL_TEMP,
            COL_VPEAK_DUT,
            COL_VPEAK_OTHER,
            DATA_START_ROW,
            export_short_circuit,
        )
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.channel_mapping import (
            apply_mapping,
            infer_short_circuit_mapping_from_bundle,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import TekMetadata, WaveformBundle
        from dpt_extractor.pipeline.run_extract import run_extraction

        source = load_waveform(DL_UH)
        labels = {
            "CH4": "H-Vge",
            "CH5": "H-Vce",
            "CH2": "Ic",
        }
        bundle = WaveformBundle(
            t=source.t,
            channels={
                "CH4": source.get("CH1").copy(),
                "CH5": source.get("CH2").copy(),
                "CH2": source.get("CH3").copy(),
            },
            meta=TekMetadata(
                sample_interval=source.meta.sample_interval,
                record_length=source.meta.record_length,
                zero_index=source.meta.zero_index,
                source_path=str(DL_UH),
                channel_labels=labels,
            ),
        )
        mapping = infer_short_circuit_mapping_from_bundle(bundle, "upper")
        self.assertIsNotNone(mapping)
        assert mapping is not None
        profile = apply_mapping(
            as_short_circuit_profile(guess_profile_from_path(DL_UH)),
            mapping,
        )
        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value

        result = run_extraction(bundle, profile, cfg)
        sc = result.short_circuit

        self.assertTrue(result.short_circuit_mode)
        self.assertGreater(sc.ic_max, 3000.0)
        self.assertGreater(sc.tsc, 1.0)
        self.assertGreater(sc.esc_dut, 1.0)
        self.assertGreater(sc.vpeak_dut, 500.0)
        self.assertFalse(result.is_metric_unavailable("短路过程", "短路电流Imax"))
        self.assertFalse(result.is_metric_unavailable("短路过程", "短路时间Tsc"))
        self.assertFalse(result.is_metric_unavailable("短路过程", "短路能量Esc_本管"))
        self.assertFalse(result.is_metric_unavailable("短路过程", "应力Vpeak_本管"))
        self.assertTrue(result.is_metric_unavailable("短路过程", "短路能量Esc_对管"))
        self.assertTrue(result.is_metric_unavailable("短路过程", "应力Vpeak_对管"))
        self.assertTrue(result.is_metric_unavailable("短路过程", "Desat动作时间"))

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "short_missing_other.xlsx"
            export_short_circuit(result, out)
            ws = load_workbook(out).active
            target_row = next(
                row
                for row in range(DATA_START_ROW, ws.max_row + 1)
                if ws.cell(row, COL_TEMP).value == "-40℃"
                and ws.cell(row, COL_PHASE).value == "UH"
            )
            self.assertIsNotNone(ws.cell(target_row, COL_ICMAX).value)
            self.assertIsNotNone(ws.cell(target_row, COL_ESC_DUT).value)
            self.assertIsNotNone(ws.cell(target_row, COL_VPEAK_DUT).value)
            self.assertIsNone(ws.cell(target_row, COL_ESC_OTHER).value)
            self.assertIsNone(ws.cell(target_row, COL_VPEAK_OTHER).value)
            self.assertIsNone(ws.cell(target_row, COL_DESAT).value)

    def test_short_circuit_tsc_can_use_10_percent_current_range(self):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import as_short_circuit_profile, guess_profile_from_path
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_10
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_percent_cursors,
        )

        bundle = load_waveform(DL_UH)
        profile = guess_profile_from_path(DL_UH)
        cfg_default = load_config()
        cfg_default.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        default_result = run_extraction(bundle, profile, cfg_default)

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        cfg.short_circuit_tsc_range = SHORT_CIRCUIT_TSC_RANGE_10
        result = run_extraction(bundle, profile, cfg)
        sc = result.short_circuit
        assert result.segments is not None
        i0, i1 = result.segments.turn_off
        sc_profile = as_short_circuit_profile(profile)
        cursors = short_circuit_current_percent_cursors(
            bundle.t,
            bundle_total_current(bundle, sc_profile),
            i0,
            i1,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
            percent=10.0,
        )

        self.assertIsNotNone(cursors)
        assert cursors is not None
        self.assertEqual(sc.tsc_range, SHORT_CIRCUIT_TSC_RANGE_10)
        self.assertAlmostEqual(sc.tsc_start_us, cursors.t_a_s * 1e6, delta=1e-6)
        self.assertAlmostEqual(sc.tsc_end_us, cursors.t_b_s * 1e6, delta=1e-6)
        self.assertAlmostEqual(sc.tsc, (cursors.t_b_s - cursors.t_a_s) * 1e6, delta=1e-6)
        self.assertLess(sc.tsc, default_result.short_circuit.tsc)
        self.assertAlmostEqual(sc.esc_dut, default_result.short_circuit.esc_dut, delta=1e-6)
        self.assertAlmostEqual(sc.esc_other, default_result.short_circuit.esc_other, delta=1e-6)

    def test_lower_short_circuit_tsc_10_percent_keeps_energy_and_profile(self):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_10
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction

        bundle = load_waveform(DL_UL)
        profile = guess_profile_from_path(DL_UL)
        cfg_default = load_config()
        cfg_default.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        default_result = run_extraction(bundle, profile, cfg_default)

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        cfg.short_circuit_tsc_range = SHORT_CIRCUIT_TSC_RANGE_10
        result = run_extraction(bundle, profile, cfg)
        sc = result.short_circuit

        self.assertEqual(result.profile_code, "UL")
        self.assertEqual(sc.tsc_range, SHORT_CIRCUIT_TSC_RANGE_10)
        self.assertGreater(sc.tsc, 0.0)
        self.assertLess(sc.tsc, default_result.short_circuit.tsc)
        self.assertAlmostEqual(sc.ic_max, default_result.short_circuit.ic_max, delta=1e-6)
        self.assertAlmostEqual(sc.esc_dut, default_result.short_circuit.esc_dut, delta=1e-6)
        self.assertAlmostEqual(sc.esc_other, default_result.short_circuit.esc_other, delta=1e-6)

    def test_custom_symmetric_tsc_range_covers_upper_and_lower_without_linkage(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import as_short_circuit_profile, guess_profile_from_path
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_percent_cursors,
        )

        for path, expected_profile in ((DL_UH, "UH"), (DL_UL, "UL")):
            with self.subTest(path=path.name):
                bundle = load_waveform(path)
                profile = guess_profile_from_path(path)

                cfg_default = load_config()
                cfg_default.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                default_result = run_extraction(bundle, profile, cfg_default)

                cfg_custom = load_config()
                cfg_custom.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                cfg_custom.short_circuit_tsc_range = "25%-25%"
                custom_result = run_extraction(bundle, profile, cfg_custom)
                custom = custom_result.short_circuit

                self.assertEqual(custom_result.profile_code, expected_profile)
                self.assertEqual(custom.tsc_range, "25%-25%")
                self.assertGreater(custom.tsc, 0.0)
                self.assertLess(custom.tsc, default_result.short_circuit.tsc)
                self.assertIsNotNone(custom.tsc_start_us)
                self.assertIsNotNone(custom.tsc_end_us)
                assert custom_result.segments is not None
                i0, i1 = custom_result.segments.turn_off
                current = bundle_total_current(
                    bundle,
                    as_short_circuit_profile(profile),
                )
                cursors = short_circuit_current_percent_cursors(
                    bundle.t,
                    current,
                    i0,
                    i1,
                    bundle.dt,
                    smooth_ns=cfg_custom.smoothing.detect_window_ns,
                    percent=25.0,
                )
                self.assertIsNotNone(cursors)
                assert cursors is not None
                self.assertAlmostEqual(
                    custom.tsc_start_us,
                    cursors.t_a_s * 1e6,
                    delta=1e-6,
                )
                self.assertAlmostEqual(
                    custom.tsc_end_us,
                    cursors.t_b_s * 1e6,
                    delta=1e-6,
                )
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_a_s, bundle.t, current)),
                    cursors.hb_a,
                    delta=1e-3,
                )
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_b_s, bundle.t, current)),
                    cursors.hb_a,
                    delta=1e-3,
                )
                for attr in (
                    "ic_max",
                    "esc_dut",
                    "esc_other",
                    "vpeak_dut",
                    "vpeak_other",
                    "desat_time",
                ):
                    expected_value = getattr(default_result.short_circuit, attr)
                    actual_value = getattr(custom, attr)
                    if expected_value is None:
                        self.assertIsNone(actual_value, attr)
                        continue
                    expected = float(expected_value)
                    actual = float(actual_value)
                    if np.isnan(expected):
                        self.assertTrue(np.isnan(actual), attr)
                    else:
                        self.assertAlmostEqual(actual, expected, delta=1e-6, msg=attr)

    def test_extracts_lower_short_circuit_uses_lower_vce_as_dut(self):
        _bundle, result = self._extract(DL_UL)
        sc = result.short_circuit

        self.assertTrue(result.short_circuit_mode)
        self.assertEqual(result.profile_code, "UL")
        self.assertGreater(sc.ic_max, 3000.0)
        self.assertGreater(sc.esc_dut, 1.0)
        self.assertGreater(sc.vpeak_dut, sc.vpeak_other)
        self.assertGreater(sc.tsc, 1.0)

    def test_vpeak_other_uses_dut_vge_window_for_upper_and_lower(self):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import short_circuit_vpeak_cursors

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        for path in (DL_UH, DL_UL):
            bundle = load_waveform(path)
            profile = as_short_circuit_profile(guess_profile_from_path(path))
            result = run_extraction(bundle, profile, cfg)
            assert result.segments is not None
            i0, i1 = result.segments.turn_off
            cursors = short_circuit_vpeak_cursors(
                bundle.t,
                bundle.get(profile.vge),
                bundle.get(profile.v_diode),
                i0,
                i1,
                bundle.dt,
                smooth_ns=cfg.smoothing.detect_window_ns,
            )

            self.assertIsNotNone(cursors, path.name)
            assert cursors is not None
            self.assertAlmostEqual(
                result.short_circuit.vpeak_other,
                cursors.ha_a,
                delta=1e-6,
                msg=path.name,
            )

@unittest.skipUnless(
    all(path.exists() for path in (SONG_DL_LT_UL, SONG_DL_LT_VL, SONG_DDD_LT_VH)),
    "slow-gate short-circuit samples missing",
)
class TestSlowGateVpeakRawCrossings(unittest.TestCase):
    def test_slow_gate_vpeak_window_uses_raw_vge_base_crossings(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import short_circuit_vpeak_cursors

        expected_dut_peaks = {
            SONG_DL_LT_UL: 1013.9375,
            SONG_DL_LT_VL: 1013.3125,
            SONG_DDD_LT_VH: 1026.0625,
        }
        for path, expected_peak in expected_dut_peaks.items():
            with self.subTest(path=path):
                cfg = load_config()
                cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                bundle = load_waveform(path)
                profile = as_short_circuit_profile(guess_profile_from_path(path))
                result = run_extraction(bundle, profile, cfg)
                assert result.segments is not None
                gate0, gate1 = result.segments.turn_off
                vge = np.asarray(bundle.get(profile.vge), dtype=np.float64)

                dut = short_circuit_vpeak_cursors(
                    bundle.t,
                    vge,
                    bundle.get(profile.vce),
                    gate0,
                    gate1,
                    bundle.dt,
                    smooth_ns=cfg.smoothing.detect_window_ns,
                )
                other = short_circuit_vpeak_cursors(
                    bundle.t,
                    vge,
                    bundle.get(profile.v_diode),
                    gate0,
                    gate1,
                    bundle.dt,
                    smooth_ns=cfg.smoothing.detect_window_ns,
                )
                self.assertIsNotNone(dut)
                self.assertIsNotNone(other)
                assert dut is not None and other is not None

                for cursors in (dut, other):
                    self.assertAlmostEqual(
                        float(np.interp(cursors.t_a_s, bundle.t, vge)),
                        cursors.hb_a,
                        places=9,
                    )
                    self.assertAlmostEqual(
                        float(np.interp(cursors.t_b_s, bundle.t, vge)),
                        cursors.hb_a,
                        places=9,
                    )
                    self.assertAlmostEqual(dut.t_a_s, other.t_a_s, places=15)
                    self.assertAlmostEqual(dut.t_b_s, other.t_b_s, places=15)
                    self.assertAlmostEqual(dut.hb_a, other.hb_a, places=12)

                self.assertAlmostEqual(dut.ha_a, expected_peak, places=6)
                self.assertAlmostEqual(
                    result.short_circuit.vpeak_dut,
                    expected_peak,
                    places=6,
                )
                self.assertAlmostEqual(
                    result.short_circuit.vpeak_other,
                    other.ha_a,
                    places=6,
                )

    @unittest.skipUnless(
        all(path.exists() for path in (SONG_DDD_HT_UH, SONG_DDD_LT_UL)),
        "raw-current short-circuit samples missing",
    )
    def test_short_current_window_uses_raw_ic_base_crossings(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.channel_mapping import (
            apply_mapping,
            infer_short_circuit_mapping_from_bundle,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import short_circuit_current_cursors

        expected = {
            # B is the raw Base intersection nearest the smoothed physical
            # fall anchor.  The earlier first raw intersections are ringing
            # dropouts while the local smoothed current is still 83.38 A and
            # 68.85 A above Base, respectively.
            SONG_DDD_HT_UH: (7697.8125, 2.7110327871688495),
            SONG_DDD_LT_UL: (7540.625, 2.3329034965247444),
        }
        for path, (expected_peak, expected_tsc_us) in expected.items():
            with self.subTest(path=path):
                cfg = load_config()
                cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
                bundle = load_waveform(path)
                base_profile = guess_profile_from_path(path)
                mapping = infer_short_circuit_mapping_from_bundle(
                    bundle, base_profile.bridge
                )
                profile = as_short_circuit_profile(base_profile)
                if mapping is not None:
                    profile = apply_mapping(profile, mapping)
                result = run_extraction(bundle, profile, cfg)
                assert result.segments is not None
                cursors = short_circuit_current_cursors(
                    bundle.t,
                    bundle_total_current(bundle, profile),
                    *result.segments.turn_off,
                    bundle.dt,
                    smooth_ns=cfg.smoothing.detect_window_ns,
                )
                self.assertIsNotNone(cursors)
                assert cursors is not None
                ic = np.asarray(bundle_total_current(bundle, profile), dtype=np.float64)
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_a_s, bundle.t, ic)),
                    cursors.hb_a,
                    places=8,
                )
                self.assertAlmostEqual(
                    float(np.interp(cursors.t_b_s, bundle.t, ic)),
                    cursors.hb_a,
                    places=8,
                )
                self.assertAlmostEqual(result.short_circuit.ic_max, expected_peak, places=6)
                self.assertAlmostEqual(
                    result.short_circuit.tsc,
                    expected_tsc_us,
                    places=9,
                )

    @unittest.skipUnless(
        SONG_DDD_HT_UH.exists(),
        "10-percent short-circuit sample missing",
    )
    def test_tsc_10_percent_uses_raw_current_threshold_crossings(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.channel_mapping import (
            apply_mapping,
            infer_short_circuit_mapping_from_bundle,
        )
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_10
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_percent_cursors,
        )

        bundle = load_waveform(SONG_DDD_HT_UH)
        base_profile = guess_profile_from_path(SONG_DDD_HT_UH)
        mapping = infer_short_circuit_mapping_from_bundle(bundle, base_profile.bridge)
        profile = as_short_circuit_profile(base_profile)
        if mapping is not None:
            profile = apply_mapping(profile, mapping)
        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        cfg.short_circuit_tsc_range = SHORT_CIRCUIT_TSC_RANGE_10
        result = run_extraction(bundle, profile, cfg)
        assert result.segments is not None
        ic = np.asarray(bundle_total_current(bundle, profile), dtype=np.float64)
        cursors = short_circuit_current_percent_cursors(
            bundle.t,
            ic,
            *result.segments.turn_off,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
            percent=10.0,
        )

        self.assertIsNotNone(cursors)
        assert cursors is not None
        self.assertAlmostEqual(
            float(np.interp(cursors.t_a_s, bundle.t, ic)),
            cursors.hb_a,
            places=8,
        )
        self.assertAlmostEqual(
            float(np.interp(cursors.t_b_s, bundle.t, ic)),
            cursors.hb_a,
            places=8,
        )
        self.assertAlmostEqual(
            result.short_circuit.tsc_start_us,
            cursors.t_a_s * 1e6,
            places=9,
        )
        self.assertAlmostEqual(
            result.short_circuit.tsc_end_us,
            cursors.t_b_s * 1e6,
            places=9,
        )

    @unittest.skipUnless(
        SONG_LCG_DDD_RT_UL.exists(),
        "flat-gate short-circuit sample missing",
    )
    def test_vpeak_without_vge_base_crossings_uses_current_ab_window(self):
        import numpy as np

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import (
            as_short_circuit_profile,
            guess_profile_from_path,
        )
        from dpt_extractor.models.channel_mapping import (
            apply_mapping,
            infer_short_circuit_mapping_from_bundle,
        )
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction
        from dpt_extractor.models.waveform import bundle_total_current
        from dpt_extractor.pipeline.short_circuit_extract import (
            short_circuit_current_cursors,
        )

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        bundle = load_waveform(SONG_LCG_DDD_RT_UL)
        base_profile = guess_profile_from_path(SONG_LCG_DDD_RT_UL)
        mapping = infer_short_circuit_mapping_from_bundle(bundle, base_profile.bridge)
        profile = as_short_circuit_profile(base_profile)
        if mapping is not None:
            profile = apply_mapping(profile, mapping)
        result = run_extraction(bundle, profile, cfg)
        assert result.segments is not None
        gate0, gate1 = result.segments.turn_off
        current_cursors = short_circuit_current_cursors(
            bundle.t,
            bundle_total_current(bundle, profile),
            gate0,
            gate1,
            bundle.dt,
            smooth_ns=cfg.smoothing.detect_window_ns,
        )
        if current_cursors is None:
            current_cursors = short_circuit_current_cursors(
                bundle.t,
                bundle_total_current(bundle, profile),
                0,
                bundle.n - 1,
                bundle.dt,
                smooth_ns=cfg.smoothing.detect_window_ns,
            )
        self.assertIsNotNone(current_cursors)
        assert current_cursors is not None
        expected_dut = float(
            np.nanmax(bundle.get(profile.vce)[current_cursors.i0 : current_cursors.i1 + 1])
        )
        expected_other = float(
            np.nanmax(
                bundle.get(profile.v_diode)[current_cursors.i0 : current_cursors.i1 + 1]
            )
        )
        self.assertFalse(
            result.is_metric_unavailable("短路过程", "应力Vpeak_本管")
        )
        self.assertFalse(
            result.is_metric_unavailable("短路过程", "应力Vpeak_对管")
        )
        self.assertAlmostEqual(result.short_circuit.vpeak_dut, expected_dut, places=9)
        self.assertAlmostEqual(result.short_circuit.vpeak_other, expected_other, places=9)


@unittest.skipUnless(DDD_UH.exists(), "short-circuit DDD sample missing")
class TestShortCircuitMathChannels(unittest.TestCase):
    def test_uses_matching_math_channels_when_available(self):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        bundle = load_waveform(DDD_UH)
        result = run_extraction(bundle, guess_profile_from_path(DDD_UH), cfg)

        self.assertEqual(result.short_circuit.energy_dut_channel, "MATH1")
        self.assertEqual(result.short_circuit.energy_other_channel, "MATH2")
        self.assertGreater(result.short_circuit.esc_dut, result.short_circuit.esc_other)


class TestValidateTssSampleClassification(unittest.TestCase):
    def test_short_token_paths_use_short_circuit_validation(self):
        from scripts.validate_tss_samples import _is_short_circuit_sample

        self.assertTrue(
            _is_short_circuit_sample(Path("case/short/750v-vh-short-25c_000.tss"))
        )
        self.assertTrue(
            _is_short_circuit_sample(Path("case/uh/750v-uh-short-25c_000.tss"))
        )
        self.assertFalse(
            _is_short_circuit_sample(
                Path("case/uh/UH_400V_1070A_Rgon1.515R_Rgoff6.346R_000.tss")
            )
        )

    @unittest.skipUnless(NED34_SHORT_VH_750.exists(), "NED34 short sample missing")
    def test_short_token_sample_uses_mapped_short_validation(self):
        from scripts.validate_tss_samples import _validate_sample

        result = _validate_sample(NED34_SHORT_VH_750)

        self.assertEqual(result.kind, "SC")
        self.assertEqual(result.status, "OK")
        self.assertIn("map=label", result.detail)


@unittest.skipUnless(DDD_UH.exists(), "short-circuit DDD sample missing")
class TestValidateTssSamplesScript(unittest.TestCase):
    def test_ddd_voltage_only_sample_uses_short_circuit_validation(self):
        from scripts.validate_tss_samples import (
            _is_short_circuit_sample,
            _validate_sample,
        )

        self.assertTrue(_is_short_circuit_sample(DDD_UH))
        result = _validate_sample(DDD_UH)

        self.assertEqual(result.kind, "SC")
        self.assertEqual(result.status, "OK")
        self.assertIn("Imax=", result.detail)
        self.assertIn("Tsc=", result.detail)


class TestShortCircuitTemplateLayout(unittest.TestCase):
    def test_header_merge_and_alignment(self):
        from openpyxl import load_workbook

        from dpt_extractor.export.short_circuit_layout import (
            COL_PHASE,
            COL_TEMP,
            COL_TYPE,
            COL_VDC,
            DATA_START_ROW,
            HEADER_NAME_ROW,
            HEADER_UNIT_ROW,
            LAST_COL,
            TEMPLATE_ROWS,
            build_short_circuit_workbook,
        )

        ws = build_short_circuit_workbook().active
        merged = {m.coord for m in ws.merged_cells.ranges}
        last_data_row = DATA_START_ROW + len(TEMPLATE_ROWS) - 1

        self.assertIn("A3:A4", merged)
        self.assertIn("B3:B4", merged)
        self.assertIn("C3:C4", merged)
        self.assertEqual(ws.max_row, last_data_row)

        for row in range(HEADER_NAME_ROW, DATA_START_ROW + len(TEMPLATE_ROWS)):
            for col in range(1, LAST_COL + 1):
                cell = ws.cell(row, col)
                self.assertEqual(cell.alignment.horizontal, "center")
                self.assertEqual(cell.alignment.vertical, "center")
                self.assertFalse(cell.alignment.wrap_text)

        self.assertEqual(ws.cell(HEADER_NAME_ROW, COL_TEMP).value, "Temp")
        self.assertEqual(ws.cell(HEADER_NAME_ROW, COL_PHASE).value, "测试相")
        self.assertEqual(ws.cell(HEADER_NAME_ROW, COL_TYPE).value, "短路类型")
        self.assertIsNone(ws.cell(HEADER_UNIT_ROW, COL_TEMP).value)
        self.assertIsNone(ws.cell(HEADER_UNIT_ROW, COL_PHASE).value)
        self.assertIsNone(ws.cell(HEADER_UNIT_ROW, COL_TYPE).value)
        expected_temp_labels = {
            5: "25℃",
            7: "150℃",
            9: "-40℃",
            11: "25℃",
            13: "150℃",
            15: "-40℃",
            17: "25℃",
            19: "150℃",
            21: "-40℃",
        }
        for row in range(DATA_START_ROW, DATA_START_ROW + len(TEMPLATE_ROWS)):
            if row in expected_temp_labels:
                self.assertEqual(ws.cell(row, COL_TEMP).value, expected_temp_labels[row])
            else:
                self.assertIsNone(ws.cell(row, COL_TEMP).value)
            self.assertIsNone(ws.cell(row, COL_PHASE).value)
            self.assertIsNone(ws.cell(row, COL_TYPE).value)
            self.assertIsNone(ws.cell(row, COL_VDC).value)

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "short_layout.xlsx"
            ws.parent.save(out)
            saved_ws = load_workbook(out).active

        self.assertIn("A3:A4", {m.coord for m in saved_ws.merged_cells.ranges})
        self.assertEqual(saved_ws.max_row, last_data_row)
        for row in range(DATA_START_ROW, DATA_START_ROW + len(TEMPLATE_ROWS)):
            if row in expected_temp_labels:
                self.assertEqual(saved_ws.cell(row, COL_TEMP).value, expected_temp_labels[row])
            else:
                self.assertIsNone(saved_ws.cell(row, COL_TEMP).value)
            self.assertIsNone(saved_ws.cell(row, COL_PHASE).value)
            self.assertIsNone(saved_ws.cell(row, COL_TYPE).value)
            self.assertIsNone(saved_ws.cell(row, COL_VDC).value)
        for coord in ("A3", "B3", "C3", "D3", "D4", "A5", "E5"):
            alignment = saved_ws[coord].alignment
            self.assertEqual(alignment.horizontal, "center")
            self.assertEqual(alignment.vertical, "center")
            self.assertFalse(alignment.wrap_text)

    def test_export_infers_phase_from_source_path_without_writing_temp(self):
        from openpyxl import load_workbook

        from dpt_extractor.export.short_circuit_layout import (
            COL_ICMAX,
            COL_PHASE,
            COL_TEMP,
            COL_VDC,
            DATA_START_ROW,
            TEMPLATE_ROWS,
            export_short_circuit,
        )
        from dpt_extractor.models.results import ExtractResult, ShortCircuitResult

        result = ExtractResult(
            short_circuit=ShortCircuitResult(ic_max=12.3),
            short_circuit_mode=True,
            source_path=str(Path("sample") / "RT" / "VL_750V_000.tss"),
        )

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "short_path_infer.xlsx"
            export_short_circuit(result, out)
            ws = load_workbook(out, data_only=True).active

        rows = [
            row
            for row in range(DATA_START_ROW, DATA_START_ROW + len(TEMPLATE_ROWS))
            if ws.cell(row, COL_PHASE).value == "VL"
            and ws.cell(row, COL_ICMAX).value is not None
        ]
        self.assertEqual(rows, [12])
        self.assertEqual(ws.cell(11, COL_TEMP).value, "25℃")
        self.assertIsNone(ws.cell(rows[0], COL_TEMP).value)
        self.assertEqual(ws.cell(rows[0], COL_VDC).value, 750)


@unittest.skipUnless(DL_UH.exists(), "short-circuit DL sample missing")
class TestShortCircuitExcel(unittest.TestCase):
    def test_export_short_circuit_template(self):
        from openpyxl import load_workbook

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.short_circuit_layout import (
            COL_ICMAX,
            COL_PHASE,
            COL_TEMP,
            DATA_START_ROW,
            COL_TSC,
            COL_TYPE,
            COL_VDC,
            TEMPLATE_ROWS,
        )
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.models.test_mode import TestMode
        from dpt_extractor.pipeline.run_extract import run_extraction

        cfg = load_config()
        cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        bundle = load_waveform(DL_UH)
        result = run_extraction(bundle, guess_profile_from_path(DL_UH), cfg)

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "short.xlsx"
            export_to_excel(result, out)
            ws = load_workbook(out, data_only=True).active

        self.assertEqual(ws.title, "短路测试")
        self.assertEqual(ws.cell(1, 1).value, "短路测试")
        self.assertEqual(ws.cell(3, COL_ICMAX).value, "短路电流Imax")
        rows = [
            row
            for row in range(DATA_START_ROW, DATA_START_ROW + len(TEMPLATE_ROWS))
            if ws.cell(row, COL_PHASE).value == "UH"
            and ws.cell(row, COL_ICMAX).value is not None
        ]
        self.assertEqual(rows, [9])
        row = rows[0]
        self.assertGreater(float(ws.cell(row, COL_ICMAX).value), 3000.0)
        self.assertGreater(float(ws.cell(row, COL_TSC).value), 1.0)
        self.assertEqual(ws.cell(row, COL_TEMP).value, "-40℃")
        self.assertIsNone(ws.cell(row, COL_TYPE).value)
        self.assertEqual(ws.cell(row, COL_VDC).value, 480)


@unittest.skipUnless(DL_UH.exists(), "short-circuit DL sample missing")
class TestShortCircuitGuiInteraction(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        from PyQt6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication(sys.argv)

    def setUp(self):
        from dpt_extractor.gui.main_window import (
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            _app_settings,
        )
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_DEFAULT

        self._tsc_settings = _app_settings()
        self._old_tsc_range = self._tsc_settings.value(
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            None,
        )
        self._tsc_settings.setValue(
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
            SHORT_CIRCUIT_TSC_RANGE_DEFAULT,
        )

    def tearDown(self):
        from dpt_extractor.gui.main_window import SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY

        if self._old_tsc_range is None:
            self._tsc_settings.remove(SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY)
        else:
            self._tsc_settings.setValue(
                SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
                self._old_tsc_range,
            )

    def test_short_circuit_parameter_click_keeps_current_view(self):
        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win._apply_test_mode_ui()
        win._load_file(str(DL_UH), background=False)
        self.app.processEvents()

        calls: list[tuple[float, float]] = []

        def _spy_focus(t0_us: float, t1_us: float) -> None:
            calls.append((float(t0_us), float(t1_us)))

        win.wave_plot.focus_interval_us = _spy_focus  # type: ignore[method-assign]
        win._enable_generic_parameter_interaction("短路过程", "短路电流Imax")
        self.app.processEvents()

        self.assertEqual(calls, [])
        assert win.result is not None and win.result.segments is not None
        t = win.bundle.t
        gate0, gate1 = win.result.segments.turn_off
        cursors = win._short_circuit_ic_default_cursors()
        self.assertIsNotNone(cursors)
        assert cursors is not None
        t_a_us, t_b_us, hb, ha = cursors
        self.assertGreater(t_a_us, float(t[gate0] * 1e6))
        self.assertGreater(t_b_us, t_a_us)
        self.assertLess(t_b_us - t_a_us, float((t[gate1] - t[gate0]) * 1e6))
        self.assertAlmostEqual(
            float(win.wave_plot._cursor_a.value()), float(t_a_us), places=6
        )
        self.assertAlmostEqual(
            float(win.wave_plot._cursor_b.value()), float(t_b_us), places=6
        )
        assert win.wave_plot._h_cursor_a is not None
        assert win.wave_plot._h_cursor_b is not None
        ha_line = win.wave_plot._from_disp("ic", float(win.wave_plot._h_cursor_a.value()))
        hb_line = win.wave_plot._from_disp("ic", float(win.wave_plot._h_cursor_b.value()))
        self.assertAlmostEqual(float(ha_line), float(ha), places=3)
        self.assertAlmostEqual(float(hb_line), float(hb), places=3)
        self.assertTrue(win.wave_plot._interval_max_hline_enabled)
        self.assertAlmostEqual(win.result.short_circuit.tsc, t_b_us - t_a_us, places=6)

        win._enable_generic_parameter_interaction("短路过程", "短路时间Tsc")
        self.app.processEvents()
        self.assertEqual(calls, [])
        self.assertAlmostEqual(
            float(win.wave_plot._cursor_a.value()), float(t_a_us), places=6
        )
        self.assertAlmostEqual(
            float(win.wave_plot._cursor_b.value()), float(t_b_us), places=6
        )
        self.assertTrue(win.wave_plot._interval_max_hline_enabled)
        win.close()

    def test_short_circuit_tsc_range_persists_across_restart_and_file_load(self):
        from dpt_extractor.gui.main_window import (
            MainWindow,
            SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY,
        )
        from dpt_extractor.models.results import SHORT_CIRCUIT_TSC_RANGE_10
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win._apply_test_mode_ui()
        win._load_file(str(DL_UH), background=False)
        self.app.processEvents()
        tsc_key = ("短路过程", "短路时间Tsc")
        win._manual_intervals[tsc_key] = (1.0, 2.0)
        win._manual_short_current[tsc_key] = (1.0, 2.0, 3.0, 4.0)
        win._on_short_circuit_tsc_range_changed(SHORT_CIRCUIT_TSC_RANGE_10)
        self.app.processEvents()
        self.assertNotIn(tsc_key, win._manual_intervals)
        self.assertNotIn(tsc_key, win._manual_short_current)
        self.assertEqual(
            self._tsc_settings.value(SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY),
            SHORT_CIRCUIT_TSC_RANGE_10,
        )
        custom_range = "25%-25%"
        win._manual_intervals[tsc_key] = (1.0, 2.0)
        win._manual_short_current[tsc_key] = (1.0, 2.0, 3.0, 4.0)
        win._on_short_circuit_tsc_range_changed(custom_range)
        self.app.processEvents()
        self.assertNotIn(tsc_key, win._manual_intervals)
        self.assertNotIn(tsc_key, win._manual_short_current)
        self.assertEqual(
            self._tsc_settings.value(SHORT_CIRCUIT_TSC_RANGE_SETTINGS_KEY),
            custom_range,
        )
        win.close()

        win2 = MainWindow()
        self.assertEqual(win2.cfg.short_circuit_tsc_range, custom_range)
        win2.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win2._apply_test_mode_ui()
        win2._load_file(str(DL_UH), background=False)
        self.app.processEvents()
        self.assertEqual(win2.cfg.short_circuit_tsc_range, custom_range)
        assert win2.result is not None
        self.assertEqual(win2.result.short_circuit.tsc_range, custom_range)
        self.assertEqual(
            win2._load_cfg_for_new_file().short_circuit_tsc_range,
            custom_range,
        )
        win2.close()

    def test_short_circuit_tsc_range_dialog_uses_one_symmetric_custom_value(self):
        from dpt_extractor.gui.result_table import ShortCircuitTscRangeDialog
        from dpt_extractor.models.results import (
            SHORT_CIRCUIT_TSC_RANGE_10,
            SHORT_CIRCUIT_TSC_RANGE_CUSTOM,
        )

        dialog = ShortCircuitTscRangeDialog(current="25%-25%")
        try:
            self.assertEqual(
                dialog.range_selector.currentText(),
                SHORT_CIRCUIT_TSC_RANGE_CUSTOM,
            )
            self.assertAlmostEqual(dialog.custom_percent.value(), 25.0)
            self.assertFalse(dialog.custom_percent.isHidden())
            self.assertEqual(dialog.range_label(), "25%-25%")

            dialog.range_selector.setCurrentText(SHORT_CIRCUIT_TSC_RANGE_10)
            self.assertTrue(dialog.custom_percent.isHidden())
            self.assertEqual(dialog.range_label(), SHORT_CIRCUIT_TSC_RANGE_10)
        finally:
            dialog.close()

    def test_initial_short_circuit_cursors_use_tsc_window_without_edge_markers(self):
        if not DDD_RT_VH.exists():
            self.skipTest("screenshot-matching short-circuit sample missing")

        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win._apply_test_mode_ui()
        win._load_file(str(DDD_RT_VH), background=False)
        self.app.processEvents()

        assert win.result is not None and win.result.segments is not None
        sc = win.result.short_circuit
        self.assertIsNotNone(sc.tsc_start_us)
        self.assertIsNotNone(sc.tsc_end_us)
        assert sc.tsc_start_us is not None and sc.tsc_end_us is not None

        labels: set[str] = set()
        for item in win.wave_plot.plot.getPlotItem().items:
            label = getattr(getattr(item, "label", None), "format", None)
            if isinstance(label, str):
                labels.add(label)

        self.assertNotIn("短路开始", labels)
        self.assertNotIn("短路结束", labels)
        gate0, gate1 = win.result.segments.turn_off
        self.assertGreater(abs(sc.tsc_start_us - float(win.bundle.t[gate0] * 1e6)), 0.01)
        self.assertGreater(abs(sc.tsc_end_us - float(win.bundle.t[gate1] * 1e6)), 0.05)
        assert win.wave_plot._cursor_a is not None and win.wave_plot._cursor_b is not None
        self.assertAlmostEqual(float(win.wave_plot._cursor_a.value()), sc.tsc_start_us, places=6)
        self.assertAlmostEqual(float(win.wave_plot._cursor_b.value()), sc.tsc_end_us, places=6)
        win.close()

    def test_short_circuit_energy_and_vpeak_default_windows(self):
        import numpy as np

        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win._apply_test_mode_ui()
        win._load_file(str(DL_UH), background=False)
        self.app.processEvents()

        calls: list[tuple[float, float]] = []

        def _spy_focus(t0_us: float, t1_us: float) -> None:
            calls.append((float(t0_us), float(t1_us)))

        win.wave_plot.focus_interval_us = _spy_focus  # type: ignore[method-assign]
        ic_cursors = win._short_circuit_ic_default_cursors()
        dut_vce_cursors = win._short_circuit_vpeak_default_cursors(win.profile.vce)
        other_vce_cursors = win._short_circuit_vpeak_default_cursors(
            win.profile.v_diode,
            gate_channel=win.profile.vge,
        )
        self.assertIsNotNone(ic_cursors)
        self.assertIsNotNone(dut_vce_cursors)
        self.assertIsNotNone(other_vce_cursors)
        assert ic_cursors is not None
        assert dut_vce_cursors is not None
        assert other_vce_cursors is not None

        def _assert_ab(cursors: tuple[float, float, float, float]) -> None:
            t_a_us, t_b_us, _hb, _ha = cursors
            self.assertAlmostEqual(
                float(win.wave_plot._cursor_a.value()), float(t_a_us), places=6
            )
            self.assertAlmostEqual(
                float(win.wave_plot._cursor_b.value()), float(t_b_us), places=6
            )

        for name in ("短路能量Esc_本管", "短路能量Esc_对管"):
            win._enable_generic_parameter_interaction("短路过程", name)
            self.app.processEvents()
            _assert_ab(ic_cursors)
            _ta, _tb, hb, _ha = ic_cursors
            assert win.wave_plot._h_cursor_a is not None
            assert win.wave_plot._h_cursor_b is not None
            hb_line = win.wave_plot._from_disp(
                "ic", float(win.wave_plot._h_cursor_b.value())
            )
            self.assertAlmostEqual(float(hb_line), float(hb), places=3)
            i0 = int(np.searchsorted(win.bundle.t, min(_ta, _tb) * 1e-6, side="left"))
            i1 = int(np.searchsorted(win.bundle.t, max(_ta, _tb) * 1e-6, side="left"))
            marker = win._short_circuit_energy_peak_marker(name, i0, i1)
            if marker is not None:
                peak, peak_channel = marker
                ha_line = win.wave_plot._from_disp(
                    peak_channel, float(win.wave_plot._h_cursor_a.value())
                )
                self.assertAlmostEqual(float(ha_line), float(peak), places=3)

        for name, cursors, channel, base_channel in (
            ("应力Vpeak_本管", dut_vce_cursors, "vce", "vge"),
            ("应力Vpeak_对管", other_vce_cursors, "v_diode", "vge"),
        ):
            win._enable_generic_parameter_interaction("短路过程", name)
            self.app.processEvents()
            _assert_ab(cursors)
            _ta, _tb, hb, ha = cursors
            assert win.wave_plot._h_cursor_a is not None
            assert win.wave_plot._h_cursor_b is not None
            ha_line = win.wave_plot._from_disp(
                channel, float(win.wave_plot._h_cursor_a.value())
            )
            hb_line = win.wave_plot._from_disp(
                base_channel, float(win.wave_plot._h_cursor_b.value())
            )
            self.assertAlmostEqual(float(ha_line), float(ha), places=3)
            self.assertAlmostEqual(float(hb_line), float(hb), places=3)

        self.assertEqual(calls, [])
        win.close()

    def test_short_circuit_dut_and_other_rows_have_different_colors(self):
        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.gui.theme import SECTION_SHORT_DUT, SECTION_SHORT_OTHER
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.SHORT_CIRCUIT.value
        win._apply_test_mode_ui()
        win._load_file(str(DL_UH), background=False)
        self.app.processEvents()

        row_for = {
            name: row
            for row, (_section, name) in enumerate(win.result_table._row_meta)
        }
        dut_row = row_for["短路能量Esc_本管"]
        other_row = row_for["短路能量Esc_对管"]
        dut_color = win.result_table.table.item(dut_row, 1).background().color().name()
        other_color = win.result_table.table.item(other_row, 1).background().color().name()

        self.assertEqual(dut_color, SECTION_SHORT_DUT.lower())
        self.assertEqual(other_color, SECTION_SHORT_OTHER.lower())
        self.assertNotEqual(dut_color, other_color)
        win.close()


@unittest.skipUnless(DDD_UH.exists(), "short-circuit DDD sample missing")
class TestCrossModeWaveformLoading(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        from PyQt6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication(sys.argv)

    def test_dpt_mode_keeps_short_circuit_waveform_when_parameters_fail(self):
        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.models.test_mode import TestMode

        win = MainWindow()
        win.cfg.test_mode.mode = TestMode.DPT.value
        win._apply_test_mode_ui()
        win._load_file(str(DDD_UH), background=False)
        self.app.processEvents()

        self.assertIsNotNone(win.bundle)
        self.assertIsNone(win.result)
        self.assertGreater(len(win.wave_plot._trace_items), 0)
        self.assertIsNotNone(win.wave_plot._cursor_a)
        self.assertIsNotNone(win.wave_plot._cursor_b)
        self.assertNotIn("[CH4]", win.wave_plot._readout_label.text())
        self.assertIn("参数未计算", win.result_table.summary.text())
        win.close()
