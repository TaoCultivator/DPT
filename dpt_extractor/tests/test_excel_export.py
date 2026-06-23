"""Excel 导出测试。"""
from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from dpt_extractor.tests.sample_paths import sample_tss

WH = sample_tss("WH_480V_800A_000.tss")
UH = sample_tss("UH_750V_1050A_000.tss")


@unittest.skipUnless(WH.exists() and UH.exists(), "WH/UH sample missing")
class TestExcelExport(unittest.TestCase):
    def test_default_export_path_uses_source_stem(self):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import default_export_path
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.pipeline.extract import extract_all

        bundle = load_waveform(UH)
        profile = guess_profile_from_path(UH.name)
        result = extract_all(bundle, profile, load_config())
        result.source_path = str(UH)
        self.assertEqual(default_export_path(result), UH.with_suffix(".xlsx"))

    def test_generated_workbook_layout_and_data_row(self):
        from openpyxl import load_workbook

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.mcu2506_layout import (
            DATA_ROW,
            HEADER_NAME_ROW,
            MERGE_INFO,
            MERGE_OFF,
            MERGE_SUMMARY,
            SHEET_ZOOM_PERCENT,
        )
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.pipeline.extract import extract_all

        bundle = load_waveform(WH)
        profile = guess_profile_from_path(WH.name)
        result = extract_all(bundle, profile, load_config())
        result.source_path = str(WH)

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "WH_480V_800A_000_ALL.xlsx"
            export_to_excel(result, out)
            wb = load_workbook(out, data_only=True)
            ws = wb.active
            self.assertIn("双脉冲", ws.title)
            self.assertEqual(ws.cell(HEADER_NAME_ROW, 1).value, "测试相")
            self.assertEqual(ws.cell(1, 1).value, "信息")
            self.assertEqual(ws.cell(1, MERGE_OFF[1]).value, "关断过程")
            self.assertEqual(ws.cell(1, MERGE_SUMMARY[1]).value, "汇总")
            merged = {m.coord for m in ws.merged_cells.ranges}
            self.assertIn("A1:E2", merged)
            self.assertIn("F1:P2", merged)
            self.assertIn("AI1:AK2", merged)
            self.assertIn("AC2:AH2", merged)
            self.assertEqual(ws.sheet_view.zoomScale, SHEET_ZOOM_PERCENT)
            self.assertIsNotNone(ws.cell(DATA_ROW, 2).border.left.style)
            self.assertEqual(ws.cell(2, 17).value, "开通")
            self.assertEqual(ws.cell(2, 29).value, "反向恢复")
            self.assertEqual(ws.cell(DATA_ROW, 1).value, "WH")
            self.assertAlmostEqual(float(ws.cell(DATA_ROW, 4).value), 480.0, delta=1)
            self.assertAlmostEqual(float(ws.cell(DATA_ROW, 5).value), 800.0, delta=1)
            self.assertGreater(float(ws.cell(DATA_ROW, 16).value), 0)
            self.assertGreater(float(ws.cell(DATA_ROW, 28).value), 0)
            self.assertGreater(float(ws.cell(DATA_ROW, 34).value), 0)

    def test_uh_data_on_row_5(self):
        from openpyxl import load_workbook

        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.mcu2506_layout import DATA_ROW
        from dpt_extractor.io.waveform_loader import load_waveform
        from dpt_extractor.models.bridge_profile import guess_profile_from_path
        from dpt_extractor.pipeline.extract import extract_all

        bundle = load_waveform(UH)
        profile = guess_profile_from_path(UH.name)
        result = extract_all(bundle, profile, load_config())
        result.source_path = str(UH)

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            export_to_excel(result, out)
            ws = load_workbook(out, data_only=True).active
            self.assertEqual(ws.cell(DATA_ROW, 1).value, "UH")
            self.assertAlmostEqual(float(ws.cell(DATA_ROW, 4).value), 750.0, delta=5)
            self.assertAlmostEqual(float(ws.cell(DATA_ROW, 5).value), 1050.0, delta=50)
            self.assertIsNotNone(ws.cell(DATA_ROW, 6).value)

    def test_dpt_multi_pulse_export_writes_consecutive_rows(self):
        from openpyxl import load_workbook

        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.mcu2506_layout import COL_OFF, COL_ON, DATA_ROW
        from dpt_extractor.models.results import ExtractResult, TurnOffResult, TurnOnResult

        rows = [
            ExtractResult(
                profile_code="UH",
                source_path=str(UH),
                detected_pulse_count=3,
                off_pulse_index=1,
                on_pulse_index=2,
                turn_off=TurnOffResult(delta_vce=101.0),
                turn_on=TurnOnResult(delta_vce=201.0),
            ),
            ExtractResult(
                profile_code="UH",
                source_path=str(UH),
                detected_pulse_count=3,
                off_pulse_index=2,
                on_pulse_index=3,
                turn_off=TurnOffResult(delta_vce=102.0),
                turn_on=TurnOnResult(delta_vce=202.0),
            ),
        ]

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "multi.xlsx"
            export_to_excel(rows, out)
            ws = load_workbook(out, data_only=True).active

            self.assertEqual(ws.cell(DATA_ROW, COL_OFF["delta_vce"]).value, 101.0)
            self.assertEqual(ws.cell(DATA_ROW, COL_ON["delta_vce"]).value, 201.0)
            self.assertEqual(ws.cell(DATA_ROW + 1, COL_OFF["delta_vce"]).value, 102.0)
            self.assertEqual(ws.cell(DATA_ROW + 1, COL_ON["delta_vce"]).value, 202.0)
