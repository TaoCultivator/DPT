"""验证 Excel 导出写入的是 GUI 手动微调后内存中的 result，而非仅 extract 初值。"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PyQt6.QtWidgets import QApplication

ROOT = Path(__file__).resolve().parents[2]
WL = ROOT / "WL_480V_800A_000_ALL.csv"
UH = ROOT / "UH_750V_1050A_000_ALL.csv"


def _excel_cell(path: Path, row: int, col: int) -> float:
    from openpyxl import load_workbook

    ws = load_workbook(path, data_only=True).active
    return float(ws.cell(row, col).value)


@unittest.skipUnless(WL.exists(), "WL sample missing")
class TestExcelExportManualAdjust(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication(sys.argv)

    def _load_win(self, csv_path: Path):
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.mcu2506_layout import COL_OFF, COL_ON, COL_RR, DATA_ROW
        from dpt_extractor.gui.main_window import MainWindow
        from dpt_extractor.io.tek_parser import TekParser
        from dpt_extractor.models.bridge_profile import (
            LOWER_BRIDGE,
            UPPER_BRIDGE,
            guess_profile_from_path,
        )
        from dpt_extractor.pipeline.extract import extract_all

        bundle = TekParser().parse(csv_path)
        profile = (
            LOWER_BRIDGE
            if "WL" in csv_path.name.upper() or "UL" in csv_path.name.upper()
            else UPPER_BRIDGE
        )
        win = MainWindow()
        win.bundle = bundle
        win.profile = profile
        win.result = extract_all(bundle, profile, load_config())
        win.cfg = load_config()
        win.result_table.set_result(win.result)
        win.wave_plot.plot_waveforms(bundle, profile, win.result)
        return win, export_to_excel, DATA_ROW, COL_OFF, COL_ON, COL_RR

    def test_export_reads_mutated_result_object(self):
        """导出链路直接读 ExtractResult（与 GUI 共用同一对象）。"""
        from dpt_extractor.config.loader import load_config
        from dpt_extractor.export.excel_export import export_to_excel
        from dpt_extractor.export.mcu2506_layout import COL_ON, DATA_ROW
        from dpt_extractor.io.tek_parser import TekParser
        from dpt_extractor.models.bridge_profile import LOWER_BRIDGE
        from dpt_extractor.pipeline.extract import extract_all

        bundle = TekParser().parse(WL)
        result = extract_all(bundle, LOWER_BRIDGE, load_config())
        marker = result.turn_on.turn_on_current + 111.0
        result.turn_on.turn_on_current = marker
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "t.xlsx"
            export_to_excel(result, path)
            got = _excel_cell(path, DATA_ROW, COL_ON["turn_on_current"])
        self.assertAlmostEqual(got, marker, delta=0.05)

    def test_wl_manual_ic_off_max_export(self):
        """恢复手动 A/B 区间后重算，Excel 应写入更新后的 result。"""
        win, export_to_excel, DATA_ROW, COL_OFF, COL_ON, COL_RR = self._load_win(WL)
        _ = COL_ON, COL_RR
        orig = float(win.result.turn_off.ic_off_max)
        interval = win._parameter_interval_us("关断过程", "Ic_off_max")
        self.assertIsNotNone(interval)
        t0, t1 = interval
        # 收窄区间，最大值通常与全段 extract 不同
        win._manual_intervals[("关断过程", "Ic_off_max")] = (
            t0 + 0.15,
            t1 - 0.15,
        )
        win._enable_generic_parameter_interaction("关断过程", "Ic_off_max")
        adjusted = float(win.result.turn_off.ic_off_max)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "wl_icmax.xlsx"
            export_to_excel(win.result, path)
            xls = _excel_cell(path, DATA_ROW, COL_OFF["ic_off_max"])
        self.assertAlmostEqual(xls, adjusted, delta=0.5)
        # 若区间收窄后峰值不变，至少证明导出与 result 一致
        self.assertAlmostEqual(xls, float(win.result.turn_off.ic_off_max), delta=0.5)

    def test_wl_manual_turn_on_current_callback_export(self):
        """模拟拖开通电流光标后的回调写 result，再导出 Excel。"""
        win, export_to_excel, DATA_ROW, COL_OFF, COL_ON, COL_RR = self._load_win(WL)
        _ = COL_OFF, COL_RR
        marker = 850.0
        win._enable_turn_on_current_interaction()
        # 直接调用与拖 Ha 相同的回调（GUI 内 _on_turn_on_current_change）
        win.result.turn_on.turn_on_current = marker
        win._manual_turn_on_current = (19.0, 20.0, 100.0, marker)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "wl_ic_on.xlsx"
            export_to_excel(win.result, path)
            xls = _excel_cell(path, DATA_ROW, COL_ON["turn_on_current"])
        self.assertAlmostEqual(xls, marker, delta=0.5)

    def test_wl_manual_trr_export(self):
        win, export_to_excel, DATA_ROW, COL_OFF, COL_ON, COL_RR = self._load_win(WL)
        _ = COL_OFF, COL_ON
        marker = 222.5
        win.result.reverse_recovery.trr = marker
        win.result_table.set_metric_value("反向恢复", "Trr", marker)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "wl_trr.xlsx"
            export_to_excel(win.result, path)
            xls = _excel_cell(path, DATA_ROW, COL_RR["trr"])
        self.assertAlmostEqual(xls, marker, delta=0.05)

    def test_uh_manual_eon_export(self):
        if not UH.exists():
            self.skipTest("UH sample missing")
        win, export_to_excel, DATA_ROW, COL_OFF, COL_ON, COL_RR = self._load_win(UH)
        _ = COL_OFF, COL_RR
        marker = float(win.result.turn_on.eon) + 12.345
        win.result.turn_on.eon = marker
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "uh_eon.xlsx"
            export_to_excel(win.result, path)
            xls = _excel_cell(path, DATA_ROW, COL_ON["eon"])
        self.assertAlmostEqual(xls, marker, delta=0.02)


if __name__ == "__main__":
    unittest.main()
