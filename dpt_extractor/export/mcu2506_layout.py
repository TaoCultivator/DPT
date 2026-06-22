"""MCU2506 双脉冲数据表：按规范列序生成工作簿并写入单次测试结果。"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from dpt_extractor.models.results import ExtractResult
from dpt_extractor.utils.filename import parse_setpoints_from_filename

# 列号（1-based，A–AK 共 37 列）
COL_PHASE = 1
COL_TEMP = 2
COL_CONDITION = 3
COL_VOLTAGE = 4
COL_CURRENT = 5

COL_OFF = {
    "delta_vce": 6,
    "ic_off_max": 7,
    "vce_off_max": 8,
    "dvdt": 9,
    "didt": 10,
    "ls_off": 11,
    "toff": 12,
    "td_off": 13,
    "tf": 14,
    "crosstalk": 15,
    "eoff": 16,
}

COL_ON = {
    "delta_vce": 17,
    "ic_on_max": 18,
    "vce_on_max": 19,
    "turn_on_current": 20,
    "dvdt": 21,
    "didt": 22,
    "ls_on": 23,
    "ton": 24,
    "td_on": 25,
    "tr": 26,
    "crosstalk": 27,
    "eon": 28,
}

COL_RR = {
    "irr": 29,
    "trr": 30,
    "vrr": 31,
    "dvdt": 32,
    "didt": 33,
    "err": 34,
}

COL_TAIL = {
    "deadtime": 35,
    "etotal": 36,
    "picture": 37,
}

LAST_COL = COL_TAIL["picture"]
SHEET_ZOOM_PERCENT = 55

# 第 1 行大分区合并（按粗边界分区）
MERGE_INFO = (1, 1, 2, 5)          # A1:E2  信息
MERGE_OFF = (1, 6, 2, 16)          # F1:P2  关断过程
MERGE_ON_PROCESS = (1, 17, 1, 34)  # Q1:AH1 开通过程（含反向恢复）
MERGE_SUMMARY = (1, 35, 2, 37)     # AI1:AK2 汇总

HEADER_NAME_ROW = 3
HEADER_UNIT_ROW = 4
DATA_ROW = 5

FILL_INFO = PatternFill("solid", fgColor="4472C4")
FILL_OFF = PatternFill("solid", fgColor="F5A623")
FILL_ON = PatternFill("solid", fgColor="7ED321")
FILL_RR = PatternFill("solid", fgColor="4A90E2")
FILL_ENERGY = PatternFill("solid", fgColor="FFFF00")
# 汇总表头底色：按用户给定紫色色块设置
FILL_SUMMARY_HDR = PatternFill("solid", fgColor="8064A2")
FILL_TAIL = PatternFill("solid", fgColor="FCE4D6")
FILL_OFF_HDR = PatternFill("solid", fgColor="FCE4D6")
FILL_ON_HDR = PatternFill("solid", fgColor="E2EFDA")
FILL_RR_HDR = PatternFill("solid", fgColor="DDEBF7")
FILL_INFO_HDR = PatternFill("solid", fgColor="D9E1F2")

THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# (列号, 参数名, 单位, 表头底色)
COLUMNS: list[tuple[int, str, str, PatternFill]] = [
    (COL_PHASE, "测试相", "", FILL_INFO_HDR),
    (COL_TEMP, "Temp", "°C", FILL_INFO_HDR),
    (COL_CONDITION, "测试条件", "", FILL_INFO_HDR),
    (COL_VOLTAGE, "Recorded Voltage", "V", FILL_INFO_HDR),
    (COL_CURRENT, "Recorded Current", "A", FILL_INFO_HDR),
    (COL_OFF["delta_vce"], "ΔVce", "V", FILL_OFF_HDR),
    (COL_OFF["ic_off_max"], "Ic_off_max", "A", FILL_OFF_HDR),
    (COL_OFF["vce_off_max"], "Vce_off_max", "V", FILL_OFF_HDR),
    (COL_OFF["dvdt"], "dv/dt", "V/ns", FILL_OFF_HDR),
    (COL_OFF["didt"], "di/dt", "A/ns", FILL_OFF_HDR),
    (COL_OFF["ls_off"], "Ls_off", "nH", FILL_OFF_HDR),
    (COL_OFF["toff"], "Toff", "ns", FILL_OFF_HDR),
    (COL_OFF["td_off"], "Td_off", "ns", FILL_OFF_HDR),
    (COL_OFF["tf"], "Tf", "ns", FILL_OFF_HDR),
    (COL_OFF["crosstalk"], "串扰电压", "V", FILL_OFF_HDR),
    (COL_OFF["eoff"], "Eoff", "mJ", FILL_ENERGY),
    (COL_ON["delta_vce"], "ΔVce", "V", FILL_ON_HDR),
    (COL_ON["ic_on_max"], "Ic_on_max", "A", FILL_ON_HDR),
    (COL_ON["vce_on_max"], "Vce_on_max", "V", FILL_ON_HDR),
    (COL_ON["turn_on_current"], "开通电流", "A", FILL_ON_HDR),
    (COL_ON["dvdt"], "dv/dt", "V/ns", FILL_ON_HDR),
    (COL_ON["didt"], "di/dt", "A/ns", FILL_ON_HDR),
    (COL_ON["ls_on"], "Ls_on", "nH", FILL_ON_HDR),
    (COL_ON["ton"], "Ton", "ns", FILL_ON_HDR),
    (COL_ON["td_on"], "Td_on", "ns", FILL_ON_HDR),
    (COL_ON["tr"], "Tr", "ns", FILL_ON_HDR),
    (COL_ON["crosstalk"], "串扰电压", "V", FILL_ON_HDR),
    (COL_ON["eon"], "Eon", "mJ", FILL_ENERGY),
    (COL_RR["irr"], "Irr", "A", FILL_RR_HDR),
    (COL_RR["trr"], "Trr", "ns", FILL_RR_HDR),
    (COL_RR["vrr"], "Vrr", "V", FILL_RR_HDR),
    (COL_RR["dvdt"], "Dvdt_max", "V/ns", FILL_RR_HDR),
    (COL_RR["didt"], "Didt_Irr", "A/ns", FILL_RR_HDR),
    (COL_RR["err"], "Err", "mJ", FILL_ENERGY),
    (COL_TAIL["deadtime"], "Deadtime", "ns", FILL_TAIL),
    (COL_TAIL["etotal"], "Etotal（all）", "mJ", FILL_ENERGY),
    (COL_TAIL["picture"], "Waveform", "Picture number", FILL_TAIL),
]


def _num(v: float | None, digits: int = 4) -> float | None:
    if v is None:
        return None
    if v != v or abs(v) < 1e-30:
        return None
    return round(float(v), digits)


def _crosstalk_str(vmax: float, vmin: float) -> str | None:
    if vmax == 0.0 and vmin == 0.0:
        return None
    return f"{vmax:.2f}/{vmin:.2f}"


def _phase_label(result: ExtractResult) -> str:
    if result.profile_code:
        return str(result.profile_code)
    if result.phase:
        return str(result.phase)
    return ""


def _match_setpoints(result: ExtractResult) -> tuple[float | None, float | None]:
    fn_v, fn_i = parse_setpoints_from_filename(result.source_path)
    vdc = fn_v
    idc = fn_i
    if vdc is None:
        vdc = result.vdc_set if result.vdc_set is not None else result.vdc
    if idc is None or idc <= 0:
        idc = result.idc_set if result.idc_set is not None else result.turn_off.ic_off_max
    return vdc, idc


def _sheet_title(result: ExtractResult) -> str:
    code = _phase_label(result) or "DPT"
    return f"{code}_双脉冲数据"[:31]


def _style_cell(
    ws: Worksheet,
    row: int,
    col: int,
    text: str,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    font_color: str = "000000",
) -> None:
    c = ws.cell(row, col, text)
    c.alignment = CENTER
    c.border = BORDER
    if fill is not None:
        c.fill = fill
    if bold:
        c.font = Font(bold=True, color=font_color)


def _set_value(ws: Worksheet, row: int, col: int, value: Any) -> None:
    if value is None:
        return
    if isinstance(value, str) and not value.strip():
        return
    cell = ws.cell(row, col, value)
    cell.alignment = CENTER
    cell.border = BORDER


def _set_metric_value(
    ws: Worksheet,
    row: int,
    col: int,
    result: ExtractResult,
    section: str,
    name: str,
    value: Any,
) -> None:
    if result.is_metric_unavailable(section, name):
        ws.cell(row, col).value = None
        return
    _set_value(ws, row, col, value)


def _merge_and_label(
    ws: Worksheet,
    r0: int,
    c0: int,
    r1: int,
    c1: int,
    text: str,
    *,
    fill: PatternFill,
    bold: bool = True,
    font_color: str = "FFFFFF",
) -> None:
    ws.merge_cells(
        start_row=r0, start_column=c0, end_row=r1, end_column=c1,
    )
    _style_cell(
        ws, r0, c0, text, fill=fill, bold=bold, font_color=font_color,
    )


def _apply_range_borders(
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    """表头 + 数据区全格线（含空单元格）。"""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            if cell.alignment is None:
                cell.alignment = CENTER


def _apply_sheet_view(ws: Worksheet) -> None:
    ws.sheet_view.zoomScale = SHEET_ZOOM_PERCENT
    ws.sheet_view.showGridLines = True


def build_mcu2506_workbook(result: ExtractResult | None = None) -> WorkbookType:
    """生成与 MCU2506 规范一致的双脉冲数据表（表头 1–4 行，数据从第 5 行起）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(result) if result else "双脉冲数据"

    # 第 1 行：大分区
    _merge_and_label(ws, *MERGE_INFO, "信息", fill=FILL_INFO)
    _merge_and_label(ws, *MERGE_OFF, "关断过程", fill=FILL_OFF)
    _merge_and_label(ws, *MERGE_ON_PROCESS, "开通过程", fill=FILL_ON)
    _merge_and_label(ws, *MERGE_SUMMARY, "汇总", fill=FILL_SUMMARY_HDR, font_color="FFFFFF")

    # 第 2 行：开通 / 反向恢复
    ws.merge_cells(start_row=2, start_column=17, end_row=2, end_column=28)
    ws.merge_cells(start_row=2, start_column=29, end_row=2, end_column=34)
    _style_cell(ws, 2, 17, "开通", fill=FILL_ON, bold=True)
    _style_cell(ws, 2, 29, "反向恢复", fill=FILL_RR, bold=True)

    # 第 3–4 行：列名 + 单位
    for col, name, unit, fill in COLUMNS:
        _style_cell(ws, HEADER_NAME_ROW, col, name, fill=fill, bold=True)
        _style_cell(ws, HEADER_UNIT_ROW, col, unit, fill=fill)

    for letter, w in (("A", 8), ("B", 8), ("C", 22), ("D", 11), ("E", 11)):
        ws.column_dimensions[letter].width = w
    for col in range(6, 38):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.row_dimensions[HEADER_NAME_ROW].height = 36
    ws.row_dimensions[HEADER_UNIT_ROW].height = 22

    _apply_range_borders(ws, 1, DATA_ROW, 1, LAST_COL)
    _apply_sheet_view(ws)

    return wb


def fill_data_row(ws: Worksheet, row: int, result: ExtractResult) -> None:
    """将单次提取结果写入指定行。"""
    off = result.turn_off
    on = result.turn_on
    rr = result.reverse_recovery
    vdc, idc = _match_setpoints(result)

    _set_value(ws, row, COL_PHASE, _phase_label(result))
    _set_value(ws, row, COL_VOLTAGE, _num(vdc, 1))
    _set_value(ws, row, COL_CURRENT, _num(idc, 1))

    _set_metric_value(ws, row, COL_OFF["delta_vce"], result, "关断过程", "ΔVce", _num(off.delta_vce))
    _set_metric_value(
        ws, row, COL_OFF["ic_off_max"], result, "关断过程", "Ic_off_max", _num(off.ic_off_max),
    )
    _set_metric_value(
        ws, row, COL_OFF["vce_off_max"], result, "关断过程", "Vce_off_max", _num(off.vce_off_max),
    )
    _set_metric_value(ws, row, COL_OFF["dvdt"], result, "关断过程", "dv/dt", _num(off.dvdt))
    _set_metric_value(ws, row, COL_OFF["didt"], result, "关断过程", "di/dt", _num(off.didt))
    _set_metric_value(ws, row, COL_OFF["ls_off"], result, "关断过程", "Ls_off", _num(off.ls_off))
    _set_metric_value(ws, row, COL_OFF["toff"], result, "关断过程", "Toff", _num(off.toff))
    _set_metric_value(ws, row, COL_OFF["td_off"], result, "关断过程", "Td_off", _num(off.td_off))
    _set_metric_value(ws, row, COL_OFF["tf"], result, "关断过程", "Tf", _num(off.tf))
    _set_metric_value(
        ws,
        row,
        COL_OFF["crosstalk"],
        result,
        "关断过程",
        "串扰电压",
        _crosstalk_str(off.crosstalk_vmax, off.crosstalk_vmin),
    )
    _set_metric_value(ws, row, COL_OFF["eoff"], result, "关断过程", "Eoff", _num(off.eoff, 3))

    _set_metric_value(ws, row, COL_ON["delta_vce"], result, "开通", "ΔVce", _num(on.delta_vce))
    _set_metric_value(ws, row, COL_ON["ic_on_max"], result, "开通", "Ic_on_max", _num(on.ic_on_max))
    _set_metric_value(ws, row, COL_ON["vce_on_max"], result, "开通", "Vce_on_max", _num(on.vce_on_max))
    _set_metric_value(
        ws, row, COL_ON["turn_on_current"], result, "开通", "开通电流", _num(on.turn_on_current),
    )
    _set_metric_value(ws, row, COL_ON["dvdt"], result, "开通", "dv/dt", _num(on.dvdt))
    _set_metric_value(ws, row, COL_ON["didt"], result, "开通", "di/dt", _num(on.didt))
    _set_metric_value(ws, row, COL_ON["ls_on"], result, "开通", "Ls_on", _num(on.ls_on))
    _set_metric_value(ws, row, COL_ON["ton"], result, "开通", "Ton", _num(on.ton))
    _set_metric_value(ws, row, COL_ON["td_on"], result, "开通", "Td_on", _num(on.td_on))
    _set_metric_value(ws, row, COL_ON["tr"], result, "开通", "Tr", _num(on.tr))
    _set_metric_value(
        ws,
        row,
        COL_ON["crosstalk"],
        result,
        "开通",
        "串扰电压",
        _crosstalk_str(on.crosstalk_vmax, on.crosstalk_vmin),
    )
    _set_metric_value(ws, row, COL_ON["eon"], result, "开通", "Eon", _num(on.eon, 3))

    _set_metric_value(ws, row, COL_RR["irr"], result, "反向恢复", "Irr", _num(rr.irr))
    _set_metric_value(ws, row, COL_RR["trr"], result, "反向恢复", "Trr", _num(rr.trr))
    _set_metric_value(ws, row, COL_RR["vrr"], result, "反向恢复", "Vrr", _num(rr.vrr))
    _set_metric_value(ws, row, COL_RR["dvdt"], result, "反向恢复", "dv/dt", _num(rr.dvdt_max))
    _set_metric_value(ws, row, COL_RR["didt"], result, "反向恢复", "di/dt", _num(rr.didt_irr))
    _set_metric_value(ws, row, COL_RR["err"], result, "反向恢复", "Err", _num(rr.err, 3))

    etotal = off.eoff + on.eon + rr.err
    if any(
        result.is_metric_unavailable(*key)
        for key in (("关断过程", "Eoff"), ("开通", "Eon"), ("反向恢复", "Err"))
    ):
        ws.cell(row, COL_TAIL["etotal"]).value = None
    elif etotal > 0:
        _set_value(ws, row, COL_TAIL["etotal"], _num(etotal, 3))


def export_mcu2506(result: ExtractResult, path: str | Path) -> None:
    """按规范列序生成工作簿，并将本次测试数据写入第 5 行。"""
    wb = build_mcu2506_workbook(result)
    ws = wb.active
    fill_data_row(ws, DATA_ROW, result)
    _apply_range_borders(ws, 1, DATA_ROW, 1, LAST_COL)
    _apply_sheet_view(ws)
    wb.save(Path(path))
