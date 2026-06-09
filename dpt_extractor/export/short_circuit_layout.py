"""短路测试数据表导出。"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from dpt_extractor.models.results import ExtractResult

HEADER_NAME_ROW = 3
HEADER_UNIT_ROW = 4
DATA_START_ROW = 5
LAST_COL = 12
SHEET_ZOOM_PERCENT = 85
PHASE_CODES = ("UH", "UL", "VH", "VL", "WH", "WL")
TEMP_LABELS = ("RT", "HT", "LT")

COL_TEMP = 1
COL_PHASE = 2
COL_TYPE = 3
COL_VDC = 4
COL_ICMAX = 5
COL_TSC = 6
COL_ESC_DUT = 7
COL_VPEAK_DUT = 8
COL_ESC_OTHER = 9
COL_VPEAK_OTHER = 10
COL_DESAT = 11
COL_WAVEFORM = 12

FILL_TITLE = PatternFill("solid", fgColor="ED7D31")
FILL_INFO = PatternFill("solid", fgColor="F4B183")
FILL_DUT = PatternFill("solid", fgColor="A9D18E")
FILL_OTHER = PatternFill("solid", fgColor="8EAADB")
FILL_TAIL = PatternFill("solid", fgColor="F4B183")

THIN = Side(style="thin", color="404040")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

COLUMNS: tuple[tuple[int, str, str, PatternFill], ...] = (
    (COL_TEMP, "Temp", "", FILL_INFO),
    (COL_PHASE, "测试相", "", FILL_INFO),
    (COL_TYPE, "短路类型", "", FILL_INFO),
    (COL_VDC, "母线电压Udc", "V", FILL_INFO),
    (COL_ICMAX, "短路电流Imax", "A", FILL_DUT),
    (COL_TSC, "短路时间Tsc", "us", FILL_DUT),
    (COL_ESC_DUT, "短路能量Esc_本管", "J", FILL_DUT),
    (COL_VPEAK_DUT, "应力Vpeak_本管", "V", FILL_DUT),
    (COL_ESC_OTHER, "短路能量Esc_对管", "J", FILL_OTHER),
    (COL_VPEAK_OTHER, "应力Vpeak_对管", "V", FILL_OTHER),
    (COL_DESAT, "Desat动作时间", "us", FILL_DUT),
    (COL_WAVEFORM, "Waveform", "Picture number", FILL_TAIL),
)

TEMPLATE_ROWS: tuple[tuple[str, str], ...] = (
    ("RT", "UH"),
    ("RT", "UL"),
    ("HT", "UH"),
    ("HT", "UL"),
    ("LT", "UH"),
    ("LT", "UL"),
    ("RT", "VH"),
    ("RT", "VL"),
    ("HT", "VH"),
    ("HT", "VL"),
    ("LT", "VH"),
    ("LT", "VL"),
    ("RT", "WH"),
    ("RT", "WL"),
    ("HT", "WH"),
    ("HT", "WL"),
    ("LT", "WH"),
    ("LT", "WL"),
)


def _num(v: float | None, digits: int = 4) -> float | None:
    if v is None:
        return None
    if v != v:
        return None
    return round(float(v), digits)


def _set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any = None,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    font_color: str = "000000",
) -> None:
    c = ws.cell(row, col, value)
    c.alignment = CENTER
    c.border = BORDER
    if fill is not None:
        c.fill = fill
    if bold:
        c.font = Font(bold=True, color=font_color)


def _apply_range_borders(ws: Worksheet, min_row: int, max_row: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(1, LAST_COL + 1):
            c = ws.cell(row, col)
            c.border = BORDER
            c.alignment = CENTER


def _infer_temp_label_from_path(path: str) -> str | None:
    for part in reversed(Path(path).parts):
        text = Path(part).stem.upper()
        for label in TEMP_LABELS:
            if re.search(rf"(?<![A-Z0-9]){label}(?![A-Z0-9])", text):
                return label
    return None


def _infer_phase_from_path(path: str) -> str | None:
    for part in reversed(Path(path).parts):
        text = Path(part).stem.upper()
        for code in PHASE_CODES:
            if re.search(rf"(?<![A-Z0-9]){code}(?![A-Z0-9])", text):
                return code
    return None


def _infer_voltage_from_filename(path: str) -> float | None:
    match = re.search(
        r"(?<![A-Z0-9])(\d+(?:\.\d+)?)V(?![A-Z])",
        Path(path).stem.upper(),
    )
    if not match:
        return None
    return float(match.group(1))


def _result_phase_code(result: ExtractResult) -> str:
    return (
        _infer_phase_from_path(result.source_path)
        or result.profile_code
        or result.phase
        or ""
    ).upper()


def _target_row(result: ExtractResult) -> int:
    code = _result_phase_code(result)
    temp = _infer_temp_label_from_path(result.source_path)
    first_code_row: int | None = None
    for offset, (row_temp, row_code) in enumerate(TEMPLATE_ROWS):
        row = DATA_START_ROW + offset
        if not row_code:
            continue
        if row_code == code and first_code_row is None:
            first_code_row = row
        if row_code == code and temp is not None and row_temp == temp:
            return row
    if first_code_row is not None:
        return first_code_row
    return DATA_START_ROW


def build_short_circuit_workbook(result: ExtractResult | None = None) -> WorkbookType:
    wb = Workbook()
    ws = wb.active
    ws.title = "短路测试"

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=LAST_COL)
    _set_cell(ws, 1, 1, "短路测试", fill=FILL_TITLE, bold=True)
    ws.cell(1, 1).font = Font(bold=True, size=18, color="000000")

    for col, name, unit, fill in COLUMNS:
        _set_cell(ws, HEADER_NAME_ROW, col, name, fill=fill, bold=True)
        _set_cell(ws, HEADER_UNIT_ROW, col, unit, fill=fill)
    for col in (COL_TEMP, COL_PHASE, COL_TYPE):
        ws.merge_cells(
            start_row=HEADER_NAME_ROW,
            start_column=col,
            end_row=HEADER_UNIT_ROW,
            end_column=col,
        )

    for offset, (temp, code) in enumerate(TEMPLATE_ROWS):
        row = DATA_START_ROW + offset
        fill_info = FILL_INFO
        for col in range(1, LAST_COL + 1):
            fill = FILL_INFO
            if COL_ICMAX <= col <= COL_VPEAK_DUT or col == COL_DESAT:
                fill = FILL_DUT
            elif COL_ESC_OTHER <= col <= COL_VPEAK_OTHER:
                fill = FILL_OTHER
            elif col == COL_WAVEFORM:
                fill = FILL_TAIL
            _set_cell(ws, row, col, None, fill=fill)
        if offset % 2 == 0:
            _set_cell(ws, row, COL_TEMP, temp, fill=fill_info)
    for row in range(DATA_START_ROW, DATA_START_ROW + len(TEMPLATE_ROWS), 2):
        row_code = TEMPLATE_ROWS[row - DATA_START_ROW][1]
        if row_code:
            ws.merge_cells(
                start_row=row,
                start_column=COL_TEMP,
                end_row=row + 1,
                end_column=COL_TEMP,
            )

    widths = {
        "A": 9,
        "B": 9,
        "C": 11,
        "D": 15,
        "E": 17,
        "F": 15,
        "G": 19,
        "H": 18,
        "I": 19,
        "J": 18,
        "K": 17,
        "L": 15,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for col in range(1, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].bestFit = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[HEADER_NAME_ROW].height = 28
    ws.row_dimensions[HEADER_UNIT_ROW].height = 22
    ws.sheet_view.zoomScale = SHEET_ZOOM_PERCENT
    _apply_range_borders(ws, 1, DATA_START_ROW + len(TEMPLATE_ROWS) - 1)
    return wb


def fill_short_circuit_row(ws: Worksheet, row: int, result: ExtractResult) -> None:
    sc = result.short_circuit
    _set_cell(ws, row, COL_PHASE, _result_phase_code(result), fill=FILL_INFO)
    _set_cell(
        ws,
        row,
        COL_VDC,
        _num(_infer_voltage_from_filename(result.source_path), 1),
        fill=FILL_INFO,
    )
    _set_cell(ws, row, COL_ICMAX, _num(sc.ic_max, 3), fill=FILL_DUT)
    _set_cell(ws, row, COL_TSC, _num(sc.tsc, 4), fill=FILL_DUT)
    _set_cell(ws, row, COL_ESC_DUT, _num(sc.esc_dut, 4), fill=FILL_DUT)
    _set_cell(ws, row, COL_VPEAK_DUT, _num(sc.vpeak_dut, 3), fill=FILL_DUT)
    _set_cell(ws, row, COL_ESC_OTHER, _num(sc.esc_other, 4), fill=FILL_OTHER)
    _set_cell(ws, row, COL_VPEAK_OTHER, _num(sc.vpeak_other, 3), fill=FILL_OTHER)
    _set_cell(ws, row, COL_DESAT, _num(sc.desat_time, 4), fill=FILL_DUT)


def export_short_circuit(result: ExtractResult, path: str | Path) -> None:
    wb = build_short_circuit_workbook(result)
    ws = wb.active
    fill_short_circuit_row(ws, _target_row(result), result)
    _apply_range_borders(ws, 1, DATA_START_ROW + len(TEMPLATE_ROWS) - 1)
    wb.save(Path(path))
