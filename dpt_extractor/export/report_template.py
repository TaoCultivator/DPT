from __future__ import annotations

from collections.abc import Sequence
from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Callable, Mapping

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from dpt_extractor.export.mcu2506_layout import (
    COL_CONDITION,
    COL_CURRENT,
    COL_VOLTAGE,
    DATA_ROW,
    COL_OFF,
    COL_ON,
    COL_RR,
    COL_TAIL,
    LAST_COL,
    COLUMNS,
    _crosstalk_str,
    _match_setpoints,
    _num,
    HEADER_NAME_ROW,
    HEADER_UNIT_ROW,
)
from dpt_extractor.models.results import ExtractResult, power_metric_name
from dpt_extractor.utils.filename import parse_setpoints_from_filename


PHASE_CODES = ("UH", "UL", "VH", "VL", "WH", "WL")
TEMP_LABELS = {
    "RT": ("25℃", 25),
    "HT": ("150℃", 150),
    "LT": ("-40℃", -40),
}
TemperatureLabels = Mapping[str, str | int | float]

_TEMPERATURE_IDENTITY_MAGIC = "DPT temperature identities v1"
_TEMPERATURE_IDENTITY_NAME_PREFIX = "_DPT_TEMP_IDENTITY_"


class _TemperatureIdentityLabels(dict[str, str | int | float]):
    """Current labels plus the last labels persisted in this report file."""

    def __init__(
        self,
        current: TemperatureLabels | None = None,
        *,
        previous: Mapping[str, object] | None = None,
    ) -> None:
        super().__init__(current or {})
        self.previous = {
            str(code).strip().upper(): value
            for code, value in (previous or {}).items()
            if str(code).strip().upper() in TEMP_LABELS
        }

DPT_OVERVIEW_IMAGE_PARAM = ("总概览图", "全局")
DPT_DATA_GROUP_TEMPLATE_ROWS = 4

DPT_REPORT_IMAGE_PARAMS: tuple[tuple[str, str], ...] = (
    DPT_OVERVIEW_IMAGE_PARAM,
    ("关断过程", "ΔVce"),
    ("关断过程", "dv/dt"),
    ("关断过程", "di/dt"),
    ("关断过程", "Tf"),
    ("关断过程", "Td_off"),
    ("关断过程", "Eoff"),
    ("开通", "ΔVce"),
    ("开通", "dv/dt"),
    ("开通", "di/dt"),
    ("开通", "Tr"),
    ("开通", "Td_on"),
    ("开通", "Eon"),
    ("反向恢复", "Irr"),
    ("反向恢复", "Trr"),
    ("反向恢复", "Vrr"),
    ("反向恢复", "dv/dt"),
    ("反向恢复", "di/dt"),
    ("反向恢复", "Err"),
)

_SINGLE_PULSE_SKIPPED_IMAGE_SECTIONS = {"开通", "反向恢复"}

SHORT_REPORT_IMAGE_PARAMS: tuple[tuple[str, str], ...] = (
    ("短路过程", "短路电流Imax"),
    ("短路过程", "应力Vpeak_本管"),
    ("短路过程", "短路能量Esc_对管"),
    ("短路过程", "应力Vpeak_对管"),
    ("短路过程", "Desat动作时间"),
)

_DPT_IMAGE_HEADERS: dict[tuple[str, str], tuple[str, str]] = {
    ("关断过程", "ΔVce"): ("off", "△Vce"),
    ("关断过程", "dv/dt"): ("off", "dv/dt"),
    ("关断过程", "di/dt"): ("off", "di/dt"),
    ("关断过程", "Tf"): ("off", "Tf"),
    ("关断过程", "Td_off"): ("off", "Td_off"),
    ("关断过程", "Eoff"): ("off", "Eoff"),
    ("开通", "ΔVce"): ("on", "△Vce"),
    ("开通", "dv/dt"): ("on", "dv/dt"),
    ("开通", "di/dt"): ("on", "di/dt"),
    ("开通", "Tr"): ("on", "Tr"),
    ("开通", "Td_on"): ("on", "Td_on"),
    ("开通", "Eon"): ("on", "Eon"),
    ("反向恢复", "Irr"): ("rr", "Irr"),
    ("反向恢复", "Trr"): ("rr", "Trr"),
    ("反向恢复", "Vrr"): ("rr", "Vrr"),
    ("反向恢复", "dv/dt"): ("rr", "Dvdt_max"),
    ("反向恢复", "di/dt"): ("rr", "Didt_Irr"),
    ("反向恢复", "Err"): ("rr", "Err"),
}

_SHORT_IMAGE_HEADERS: dict[tuple[str, str], str] = {
    ("短路过程", "短路电流Imax"): "短路电流Imax",
    ("短路过程", "应力Vpeak_本管"): "应力Vpeak_本管",
    ("短路过程", "短路能量Esc_对管"): "短路能量Esc_对管",
    ("短路过程", "应力Vpeak_对管"): "应力Vpeak_对管",
    ("短路过程", "Desat动作时间"): "Desat动作时间",
}

ImageMap = Mapping[tuple[str, str], str | Path]
ReportProgressCallback = Callable[[int, int, str], None]

REPORT_IMAGE_DISPLAY_SIZE_PX = (320, 240)
DPT_WAVEFORM_BLOCK_ROWS = 51
DPT_WAVEFORM_BLOCK_STRIDE = 53
_IMAGE_SLOT_PADDING_PX = 4
_WAVEFORM_HEADER_ROW_HEIGHT_PX = 26
_WAVEFORM_HEADER_FONT_SIZE = 10
_WAVEFORM_LEFT_LABEL_FONT_SIZE = 20
_WAVEFORM_STATE_FONT_SIZE = 11
_DPT_REPORT_DATA_FONT_SIZE = 12
_REPORT_VIEW_DEFAULT_SCREEN_WIDTH_PX = 1920
_REPORT_VIEW_HORIZONTAL_MARGIN_PX = 160
_REPORT_VIEW_MIN_ZOOM = 55
_REPORT_VIEW_MAX_ZOOM = 100
_FILENAME_SETPOINT_TOLERANCE = 0.05
_FALLBACK_SETPOINT_TOLERANCE = 10.0


@dataclass(frozen=True)
class ReportWriteSummary:
    report_path: Path
    data_sheet: str
    data_row: int
    data_row_end: int | None = None
    data_rows_written: int = 1
    waveform_sheet: str | None = None
    waveform_anchor_row: int | None = None
    images_written: int = 0


@dataclass(frozen=True)
class _DptDataGroup:
    start_row: int
    end_row: int
    phase: str
    temp: str
    condition: str


@dataclass(frozen=True)
class _ShortTemperatureGroup:
    start_row: int
    end_row: int
    phase_family: str
    temp: object


@dataclass(frozen=True)
class _WorksheetRowBlock:
    start_row: int
    end_row: int
    sort_key: tuple[object, ...]


@dataclass(frozen=True)
class _CellSnapshot:
    block_index: int
    row_offset: int
    column: int
    value: object
    data_type: str
    style: object
    hyperlink: object
    comment: object


@dataclass(frozen=True)
class _DptDataTarget:
    row: int
    group_index: int
    group_start_row: int
    row_offset: int
    inserted_row: bool = False


@dataclass(frozen=True)
class _DptSetpointMatch:
    vdc: float | None
    idc: float | None
    vdc_from_filename: bool = False
    idc_from_filename: bool = False


def _format_temperature_number(value: float) -> str:
    fv = float(value)
    if abs(fv - round(fv)) < 0.05:
        return str(int(round(fv)))
    return f"{fv:.1f}".rstrip("0").rstrip(".")


def _format_temperature_label(value: float) -> str:
    return f"{_format_temperature_number(value)}℃"


def _parse_temperature_number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if match is None:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _canonical_temperature_text(value: object) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    text = text.replace("°C", "℃").replace("DEG", "℃")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?C", text):
        text = f"{text[:-1]}℃"
    return text


def _normalized_temperature_labels(
    temperature_labels: TemperatureLabels | None = None,
) -> dict[str, tuple[str, float]]:
    labels = {
        code: (display, float(value))
        for code, (display, value) in TEMP_LABELS.items()
    }
    if not temperature_labels:
        return labels
    for code in TEMP_LABELS:
        if code not in temperature_labels:
            continue
        raw = temperature_labels[code]
        value = _parse_temperature_number(raw)
        if value is None:
            continue
        raw_text = str(raw).strip() if isinstance(raw, str) else ""
        display = (
            raw_text
            if raw_text
            and (
                "℃" in raw_text
                or "°" in raw_text
                or raw_text.upper().endswith("C")
            )
            else _format_temperature_label(value)
        )
        labels[code] = (display, value)
    return labels


def _read_report_temperature_identities(wb) -> dict[str, object]:
    """Read stable RT/HT/LT identities saved by an earlier report write."""

    identities: dict[str, object] = {}
    for code in TEMP_LABELS:
        name = f"{_TEMPERATURE_IDENTITY_NAME_PREFIX}{code}"
        defined_name = wb.defined_names.get(name)
        if (
            defined_name is None
            or defined_name.description != _TEMPERATURE_IDENTITY_MAGIC
        ):
            continue
        raw = str(defined_name.attr_text or "")
        if len(raw) >= 2 and raw.startswith('"') and raw.endswith('"'):
            identities[code] = raw[1:-1].replace('""', '"')
    return identities


def _temperature_identity_labels_for_workbook(
    wb,
    temperature_labels: TemperatureLabels | None,
) -> _TemperatureIdentityLabels:
    return _TemperatureIdentityLabels(
        temperature_labels,
        previous=_read_report_temperature_identities(wb),
    )


def _write_report_temperature_identities(
    wb,
    temperature_labels: TemperatureLabels | None,
) -> None:
    """Persist stable codes without adding or changing report worksheets."""

    labels = _normalized_temperature_labels(temperature_labels)
    for code in TEMP_LABELS:
        name = f"{_TEMPERATURE_IDENTITY_NAME_PREFIX}{code}"
        existing = wb.defined_names.get(name)
        if (
            existing is not None
            and existing.description != _TEMPERATURE_IDENTITY_MAGIC
        ):
            raise ValueError(
                f"报告已存在保留名称“{name}”，"
                "无法安全保存温度工况身份"
            )
        display = labels[code][0].replace('"', '""')
        wb.defined_names.add(
            DefinedName(
                name=name,
                description=_TEMPERATURE_IDENTITY_MAGIC,
                hidden=True,
                attr_text=f'"{display}"',
            )
        )


def _temperature_code_from_token(
    token: object,
    temperature_labels: TemperatureLabels | None = None,
) -> str | None:
    labels = _normalized_temperature_labels(temperature_labels)
    text = str(token or "").strip()
    code_text = text.upper()
    if code_text in TEMP_LABELS:
        return code_text
    canonical = _canonical_temperature_text(text)
    candidates: set[str] = set()
    for code, (display, _value) in labels.items():
        if canonical == _canonical_temperature_text(display):
            candidates.add(code)
    previous = getattr(temperature_labels, "previous", {})
    for code, display in previous.items():
        if canonical == _canonical_temperature_text(display):
            candidates.add(code)
    value = _parse_temperature_number(text)
    if value is not None:
        for code, (_display, expected) in labels.items():
            if abs(value - expected) < 0.05:
                candidates.add(code)
        for code, display in previous.items():
            previous_value = _parse_temperature_number(display)
            if previous_value is not None and abs(value - previous_value) < 0.05:
                candidates.add(code)
        for code, (_display, expected) in TEMP_LABELS.items():
            if abs(value - float(expected)) < 0.05:
                candidates.add(code)
    if len(candidates) == 1:
        return next(iter(candidates))
    # Duplicate/custom labels are ambiguous by definition.  Failing the
    # lookup is safer than silently writing another RT/HT/LT condition.
    return None


def _temperature_cell_matches(
    value: object,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    return _temperature_code_from_token(value, temperature_labels) == temp_code


def _temperature_values_match(left: object, right: object) -> bool:
    """Compare actual temperatures without collapsing them to RT/HT/LT."""

    left_number = _parse_temperature_number(left)
    right_number = _parse_temperature_number(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) < 0.05
    left_text = _canonical_temperature_text(left)
    right_text = _canonical_temperature_text(right)
    return bool(left_text) and left_text == right_text


def _temperature_cell_matches_display(
    value: object,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    display = _temperature_display(temp_code, temperature_labels)
    if _temperature_values_match(value, display):
        return True
    # A literal RT/HT/LT cell is an uninitialized template placeholder. It may
    # be claimed once and immediately rewritten to the actual display value.
    return str(value or "").strip().upper() == temp_code.upper()


def _temperature_tokens_match(
    left: object,
    right: object,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    left_code = str(left or "").strip().upper()
    right_code = str(right or "").strip().upper()
    if left_code in TEMP_LABELS:
        return _temperature_cell_matches_display(
            right,
            left_code,
            temperature_labels,
        )
    if right_code in TEMP_LABELS:
        return _temperature_cell_matches_display(
            left,
            right_code,
            temperature_labels,
        )
    return _temperature_values_match(left, right)


def _temperature_display(
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> str:
    return _normalized_temperature_labels(temperature_labels)[temp_code][0]


def _temperature_number_for_report_order(
    value: object,
    temperature_labels: TemperatureLabels | None = None,
) -> float | None:
    number = _parse_temperature_number(value)
    if number is not None:
        return number
    code = str(value or "").strip().upper()
    if code not in TEMP_LABELS:
        code = _temperature_code_from_token(value, temperature_labels) or ""
    if code not in TEMP_LABELS:
        return None
    return _normalized_temperature_labels(temperature_labels)[code][1]


def _temperature_report_order_key(value: float) -> tuple[int, float]:
    """Place non-negative temperatures first, then colder negatives lower."""

    number = float(value)
    return (0, number) if number >= 0 else (1, abs(number))


def _phase_report_order(phase_code: str) -> int:
    code = str(phase_code or "").strip().upper()
    try:
        return PHASE_CODES.index(code)
    except ValueError:
        return len(PHASE_CODES)


def _phase_family_report_order(phase_family: str) -> int:
    family = str(phase_family or "").strip().upper()[:1]
    try:
        return "UVW".index(family)
    except ValueError:
        return 3


def _phase_temp_label(
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> str:
    return f"{phase_code}_{_temperature_display(temp_code, temperature_labels)}"


def _merged_range_containing(
    ws: Worksheet,
    row: int,
    col: int,
) -> CellRange | None:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def _set_merged_cell_value(
    ws: Worksheet,
    row: int,
    col: int,
    value: object,
) -> None:
    rng = _merged_range_containing(ws, row, col)
    if rng is not None:
        ws.cell(rng.min_row, rng.min_col).value = value
        return
    ws.cell(row, col).value = value


def dpt_report_image_params_for_result(
    result: ExtractResult,
) -> tuple[tuple[str, str], ...]:
    return tuple(
        param
        for param in DPT_REPORT_IMAGE_PARAMS
        if (
            param == DPT_OVERVIEW_IMAGE_PARAM
            or (
                not result.is_metric_unavailable(*param)
                and (
                    not result.single_pulse_mode
                    or param[0] not in _SINGLE_PULSE_SKIPPED_IMAGE_SECTIONS
                )
            )
        )
    )


def _path_parts(path: str) -> list[str]:
    return [p for p in re.split(r"[\\/]+", str(path)) if p]


def _infer_phase_code(path: str, result: ExtractResult) -> str:
    for part in reversed(_path_parts(path)):
        stem = Path(part).stem.upper()
        for code in PHASE_CODES:
            if re.search(rf"(?<![A-Z0-9]){code}(?![A-Z0-9])", stem):
                return code
    for value in (result.profile_code, result.phase):
        text = str(value or "").upper()
        for code in PHASE_CODES:
            if text == code:
                return code
    return ""


def _resolve_phase_code(
    path: str,
    result: ExtractResult,
    phase_code: str | None = None,
) -> str:
    """Use the frozen report-page phase/bridge before path inference."""

    if phase_code is None:
        return _infer_phase_code(path, result)
    code = str(phase_code).strip().upper()
    if code not in PHASE_CODES:
        raise ValueError(f"不支持的相位/桥臂代码：{phase_code}")
    return code


def _infer_temp_code(path: str) -> str:
    for part in reversed(_path_parts(path)):
        stem = Path(part).stem.upper()
        for code in TEMP_LABELS:
            if re.search(rf"(?<![A-Z0-9]){code}(?![A-Z0-9])", stem):
                return code
        if re.search(r"(?<!\d)25(?:℃|C|DEG)?(?!\d)", stem):
            return "RT"
        if re.search(r"(?<!\d)150(?:℃|C|DEG)?(?!\d)", stem):
            return "HT"
        if re.search(r"(?<!\d)-?40(?:℃|C|DEG)?(?!\d)", stem):
            return "LT"
    return "RT"


def _resolve_temp_code(path: str, temperature_code: str | None = None) -> str:
    """Use an explicit UI/report condition before falling back to path inference."""

    if temperature_code is None:
        return _infer_temp_code(path)
    code = str(temperature_code).strip().upper()
    if code not in TEMP_LABELS:
        raise ValueError(f"不支持的温度工况代码：{temperature_code}")
    return code


def _phase_sheet_prefix(phase_code: str) -> str:
    if not phase_code:
        return "U"
    return phase_code[0]


_CONDITION_NUMBER = r"(-?\d+(?:\.\d+)?)"
_RG_ON_TERMS = r"RG[\s_\-]*ON|RGON|R[\s_\-]*ON|RON|开通(?:栅极)?(?:电阻)?"
_RG_OFF_TERMS = r"RG[\s_\-]*OFF|RGOFF|R[\s_\-]*OFF|ROFF|关断(?:栅极)?(?:电阻)?"
_CG_TERMS = r"CG|栅极电容|驱动电容"
_RG_TERMS = r"RG|栅极电阻|驱动电阻"
_INLINE_SEP = r"[ _\-\t]*"
_CONDITION_TOKEN_PATTERNS: tuple[tuple[str, tuple[re.Pattern[str], ...]], ...] = (
    (
        "rg_on",
        (
            re.compile(
                rf"(?:{_RG_ON_TERMS})\s*(?:=|:|：)?\s*{_CONDITION_NUMBER}",
                re.IGNORECASE,
            ),
            re.compile(
                rf"{_CONDITION_NUMBER}{_INLINE_SEP}(?:OHM|R|Ω)?{_INLINE_SEP}(?:{_RG_ON_TERMS})",
                re.IGNORECASE,
            ),
        ),
    ),
    (
        "rg_off",
        (
            re.compile(
                rf"(?:{_RG_OFF_TERMS})\s*(?:=|:|：)?\s*{_CONDITION_NUMBER}",
                re.IGNORECASE,
            ),
            re.compile(
                rf"{_CONDITION_NUMBER}{_INLINE_SEP}(?:OHM|R|Ω)?{_INLINE_SEP}(?:{_RG_OFF_TERMS})",
                re.IGNORECASE,
            ),
        ),
    ),
    (
        "cg",
        (
            re.compile(
                rf"(?:{_CG_TERMS})\s*(?:=|:|：)?\s*{_CONDITION_NUMBER}",
                re.IGNORECASE,
            ),
            re.compile(
                rf"{_CONDITION_NUMBER}{_INLINE_SEP}(?:NF|PF|UF)?{_INLINE_SEP}(?:{_CG_TERMS})",
                re.IGNORECASE,
            ),
        ),
    ),
    (
        "rg",
        (
            re.compile(
                rf"(?:^|[^A-Z0-9])(?:{_RG_TERMS})(?![\s_\-]*(?:ON|OFF))\s*(?:=|:|：)?\s*{_CONDITION_NUMBER}",
                re.IGNORECASE,
            ),
            re.compile(
                rf"{_CONDITION_NUMBER}{_INLINE_SEP}(?:OHM|R|Ω)?{_INLINE_SEP}(?:{_RG_TERMS})(?![\s_\-]*(?:ON|OFF))",
                re.IGNORECASE,
            ),
        ),
    ),
)
_CONDITION_TOKEN_ORDER = ("rg_on", "rg_off", "rg", "cg")
_CONDITION_LINE_KEY_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("rg_on", re.compile(rf"(?:{_RG_ON_TERMS})", re.IGNORECASE)),
    ("rg_off", re.compile(rf"(?:{_RG_OFF_TERMS})", re.IGNORECASE)),
    (
        "rg",
        re.compile(
            rf"(?:^|[^A-Z0-9])(?:{_RG_TERMS})(?![\s_\-]*(?:ON|OFF))",
            re.IGNORECASE,
        ),
    ),
    ("cg", re.compile(rf"(?:{_CG_TERMS})", re.IGNORECASE)),
)


def _condition_signature(text: object) -> dict[str, float]:
    if text is None:
        return {}
    normalized = str(text).upper().replace("Ω", "OHM").replace("欧姆", "OHM")
    tokens: dict[str, float] = {}
    for key, patterns in _CONDITION_TOKEN_PATTERNS:
        for pattern in patterns:
            match = pattern.search(normalized)
            if match:
                value = next(group for group in match.groups() if group is not None)
                tokens[key] = float(value)
                break
    return tokens


def _condition_value_matches(expected: float, actual: float) -> bool:
    return abs(float(expected) - float(actual)) <= max(0.05, abs(float(expected)) * 0.02)


def _condition_signature_matches(
    source: Mapping[str, float],
    candidate: Mapping[str, float],
) -> bool:
    if not source:
        return True
    if not candidate:
        return False
    for key, expected in source.items():
        if key in candidate:
            if not _condition_value_matches(expected, candidate[key]):
                return False
            continue
        if key == "rg":
            if any(
                _condition_value_matches(expected, candidate[alt])
                for alt in ("rg_on", "rg_off")
                if alt in candidate
            ):
                continue
        return False
    return True


def _format_condition_signature(signature: Mapping[str, float]) -> str:
    parts: list[str] = []
    labels = {
        "rg_on": "Rg_on",
        "rg_off": "Rg_off",
        "rg": "Rg",
        "cg": "Cg",
    }
    for key in _CONDITION_TOKEN_ORDER:
        if key not in signature:
            continue
        parts.append(f"{labels[key]}{_format_condition_number(signature[key])}")
    return "_".join(parts)


def _format_condition_number(value: float) -> str:
    fv = float(value)
    if abs(fv - round(fv)) < 0.05:
        return str(int(round(fv)))
    return f"{fv:g}"


def _format_condition_line(key: str, value: float) -> str:
    labels = {
        "rg_on": ("Rg_on", "ohm"),
        "rg_off": ("Rg_off", "ohm"),
        "rg": ("Rg", "ohm"),
        "cg": ("Cg", "nf"),
    }
    label, unit = labels[key]
    return f"{label} = {_format_condition_number(value)} {unit}"


def _condition_line_key(line: str) -> str | None:
    normalized = str(line or "").upper().replace("Ω", "OHM").replace("欧姆", "OHM")
    for key, pattern in _CONDITION_LINE_KEY_PATTERNS:
        if pattern.search(normalized):
            return key
    return None


def _merged_condition_text(
    existing: object,
    signature: Mapping[str, float],
) -> str:
    lines = [line for line in str(existing or "").splitlines() if line.strip()]
    used: set[str] = set()
    for idx, line in enumerate(lines):
        key = _condition_line_key(line)
        if key is None or key not in signature:
            continue
        lines[idx] = _format_condition_line(key, signature[key])
        used.add(key)
    for key in _CONDITION_TOKEN_ORDER:
        if key in signature and key not in used:
            lines.append(_format_condition_line(key, signature[key]))
    return "\n".join(lines)


def _format_condition_key(
    vdc: float | None,
    idc: float | None,
    condition_signature: Mapping[str, float] | None = None,
) -> str:
    def fmt(v: float | None) -> str:
        if v is None:
            return ""
        fv = float(v)
        if abs(fv - round(fv)) < 0.05:
            return str(int(round(fv)))
        return f"{fv:g}"

    suffix = _format_condition_signature(condition_signature or {})
    if vdc is None and idc is None:
        return suffix
    base = f"{fmt(vdc)}V_{fmt(idc)}A"
    return f"{base}_{suffix}" if suffix else base


def _condition_values(text: str | None) -> tuple[float | None, float | None]:
    if text is None:
        return None, None
    return parse_setpoints_from_filename(str(text))


def _dpt_setpoint_match(result: ExtractResult) -> _DptSetpointMatch:
    fn_v, fn_i = parse_setpoints_from_filename(result.source_path)
    vdc_from_filename = fn_v is not None
    idc_from_filename = fn_i is not None and fn_i > 0
    vdc, idc = _match_setpoints(result)
    return _DptSetpointMatch(
        vdc=vdc,
        idc=idc,
        vdc_from_filename=vdc_from_filename,
        idc_from_filename=idc_from_filename,
    )


def _condition_setpoint_match(text: str | None) -> _DptSetpointMatch:
    vdc, idc = _condition_values(text)
    return _DptSetpointMatch(
        vdc=vdc,
        idc=idc,
        vdc_from_filename=vdc is not None,
        idc_from_filename=idc is not None and idc > 0,
    )


def _setpoint_value_matches(
    expected: float | None,
    actual: float | None,
    *,
    expected_from_filename: bool,
) -> bool:
    if expected is None or actual is None:
        return False
    tolerance = (
        _FILENAME_SETPOINT_TOLERANCE
        if expected_from_filename
        else _FALLBACK_SETPOINT_TOLERANCE
    )
    return abs(float(actual) - float(expected)) <= tolerance


def _merged_value(ws: Worksheet, row: int, col: int):
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return None


def _merged_ranges_with_value(
    ws: Worksheet,
    *,
    min_col: int,
    max_col: int,
) -> list[tuple[CellRange, str]]:
    items: list[tuple[CellRange, str]] = []
    for rng in ws.merged_cells.ranges:
        if rng.min_col == min_col and rng.max_col == max_col:
            value = ws.cell(rng.min_row, rng.min_col).value
            if value is not None:
                items.append((rng, str(value)))
    return sorted(items, key=lambda item: item[0].min_row)


def _left_merged_ranges_with_value(ws: Worksheet) -> list[tuple[CellRange, str]]:
    items: list[tuple[CellRange, str]] = []
    for rng in ws.merged_cells.ranges:
        if rng.min_col != 1 or rng.max_col > 8:
            continue
        value = ws.cell(rng.min_row, rng.min_col).value
        if value is not None:
            items.append((rng, str(value)))
    return sorted(items, key=lambda item: item[0].min_row)


def _left_merged_range_at(
    ws: Worksheet,
    row: int,
    *,
    fallback_cols: int = 8,
    fallback_rows: int = 17,
) -> CellRange:
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col <= 8 and rng.min_row == row:
            return rng
    return CellRange(
        min_col=1,
        max_col=fallback_cols,
        min_row=row,
        max_row=row + fallback_rows - 1,
    )


def _numeric_cell_value(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        value = text
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _data_row_setpoints(ws: Worksheet, row: int) -> tuple[float | None, float | None]:
    return (
        _numeric_cell_value(_merged_value(ws, row, COL_VOLTAGE)),
        _numeric_cell_value(_merged_value(ws, row, COL_CURRENT)),
    )


def _data_row_matches_setpoints(
    ws: Worksheet,
    row: int,
    setpoints: _DptSetpointMatch,
) -> bool:
    row_vdc, row_idc = _data_row_setpoints(ws, row)
    return _setpoint_value_matches(
        setpoints.vdc,
        row_vdc,
        expected_from_filename=setpoints.vdc_from_filename,
    ) and _setpoint_value_matches(
        setpoints.idc,
        row_idc,
        expected_from_filename=setpoints.idc_from_filename,
    )


def _data_row_has_payload(ws: Worksheet, row: int) -> bool:
    for col in range(COL_VOLTAGE, LAST_COL + 1):
        value = _merged_value(ws, row, col)
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max(ws.max_column, LAST_COL) + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def _shift_row_dimensions_for_insert(
    ws: Worksheet,
    insert_row: int,
    amount: int,
) -> None:
    """Mirror inserted rows in explicit row dimensions; openpyxl does not."""

    for row in sorted(
        (row for row in ws.row_dimensions if row >= insert_row),
        reverse=True,
    ):
        dimension = ws.row_dimensions[row]
        del ws.row_dimensions[row]
        dimension.index = row + amount
        ws.row_dimensions[row + amount] = dimension


def _shift_merged_range_for_insert(
    rng: CellRange,
    *,
    group_start_row: int,
    insert_row: int,
) -> CellRange:
    min_row = rng.min_row
    max_row = rng.max_row
    if rng.min_row == group_start_row and rng.max_row == insert_row - 1 and rng.min_col <= 3:
        max_row += 1
    elif rng.min_row >= insert_row:
        min_row += 1
        max_row += 1
    elif rng.min_row < insert_row <= rng.max_row:
        max_row += 1
    return CellRange(
        min_col=rng.min_col,
        max_col=rng.max_col,
        min_row=min_row,
        max_row=max_row,
    )


def _insert_dpt_data_row(data_ws: Worksheet, group_start_row: int, insert_row: int) -> int:
    source_row = max(group_start_row, insert_row - 1)
    merged_ranges = [CellRange(str(rng)) for rng in data_ws.merged_cells.ranges]
    for rng in list(data_ws.merged_cells.ranges):
        data_ws.unmerge_cells(str(rng))

    if insert_row <= data_ws.max_row:
        data_ws.insert_rows(insert_row)
        _shift_row_dimensions_for_insert(data_ws, insert_row, 1)
    _copy_row_style(data_ws, source_row, insert_row)

    for rng in merged_ranges:
        shifted = _shift_merged_range_for_insert(
            rng,
            group_start_row=group_start_row,
            insert_row=insert_row,
        )
        data_ws.merge_cells(
            start_row=shifted.min_row,
            end_row=shifted.max_row,
            start_column=shifted.min_col,
            end_column=shifted.max_col,
        )
    return insert_row


def _dpt_data_groups(data_ws: Worksheet) -> list[_DptDataGroup]:
    phase_ranges = [
        item
        for item in _merged_ranges_with_value(data_ws, min_col=1, max_col=1)
        if item[0].min_row >= DATA_ROW
    ]
    groups: list[_DptDataGroup] = []
    for idx, (phase_rng, phase_value) in enumerate(phase_ranges):
        next_row = (
            phase_ranges[idx + 1][0].min_row
            if idx + 1 < len(phase_ranges)
            else data_ws.max_row + 1
        )
        groups.append(
            _DptDataGroup(
                start_row=phase_rng.min_row,
                end_row=next_row,
                phase=str(phase_value),
                temp=str(_merged_value(data_ws, phase_rng.min_row, 2) or ""),
                condition=str(_merged_value(data_ws, phase_rng.min_row, COL_CONDITION) or ""),
            )
        )
    return groups


def _row_block_index(
    blocks: Sequence[_WorksheetRowBlock],
    row: int,
) -> int | None:
    for index, block in enumerate(blocks):
        if block.start_row <= row <= block.end_row:
            return index
    return None


def _shift_drawing_anchor_rows(drawing, row_delta: int) -> None:
    if row_delta == 0:
        return
    anchor = drawing.anchor
    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        drawing.anchor = f"{get_column_letter(col)}{row + row_delta}"
        return
    start = getattr(anchor, "_from", None)
    if start is None:
        return
    start.row = int(start.row) + row_delta
    end = getattr(anchor, "to", None)
    if end is not None:
        end.row = int(end.row) + row_delta


def _reorder_worksheet_row_blocks(
    ws: Worksheet,
    blocks: Sequence[_WorksheetRowBlock],
) -> bool:
    """Stable-sort contiguous row blocks while keeping drawings attached."""

    source_blocks = sorted(blocks, key=lambda block: block.start_row)
    if len(source_blocks) < 2:
        return False
    for previous, current in zip(source_blocks, source_blocks[1:]):
        if previous.end_row + 1 != current.start_row:
            raise ValueError(
                f"报告工作表“{ws.title}”的温度块不连续："
                f"{previous.start_row}:{previous.end_row} 与 "
                f"{current.start_row}:{current.end_row}"
            )

    ordered_indices = sorted(
        range(len(source_blocks)),
        key=lambda index: (source_blocks[index].sort_key, index),
    )
    if ordered_indices == list(range(len(source_blocks))):
        return False

    target_starts: dict[int, int] = {}
    target_row = source_blocks[0].start_row
    for source_index in ordered_indices:
        block = source_blocks[source_index]
        target_starts[source_index] = target_row
        target_row += block.end_row - block.start_row + 1

    region_start = source_blocks[0].start_row
    region_end = source_blocks[-1].end_row
    merged_ranges = [CellRange(str(rng)) for rng in ws.merged_cells.ranges]
    moved_merged_ranges: list[CellRange] = []
    for rng in merged_ranges:
        overlaps = rng.max_row >= region_start and rng.min_row <= region_end
        if not overlaps:
            continue
        if rng.min_row < region_start or rng.max_row > region_end:
            raise ValueError(
                f"报告工作表“{ws.title}”存在跨越温度排序区域的合并单元格：{rng}"
            )
        block_index = _row_block_index(source_blocks, rng.min_row)
        if (
            block_index is None
            or _row_block_index(source_blocks, rng.max_row) != block_index
        ):
            raise ValueError(
                f"报告工作表“{ws.title}”存在跨温度块合并单元格：{rng}"
            )
        block = source_blocks[block_index]
        row_delta = target_starts[block_index] - block.start_row
        moved_merged_ranges.append(_shift_range_rows(rng, row_delta))

    cell_snapshots: list[_CellSnapshot] = []
    for (row, column), cell in list(ws._cells.items()):
        if not region_start <= row <= region_end:
            continue
        block_index = _row_block_index(source_blocks, row)
        if block_index is None:
            continue
        block = source_blocks[block_index]
        cell_snapshots.append(
            _CellSnapshot(
                block_index=block_index,
                row_offset=row - block.start_row,
                column=column,
                value=copy(getattr(cell, "_value", None)),
                data_type=str(getattr(cell, "data_type", "n")),
                style=copy(getattr(cell, "_style", None)),
                hyperlink=copy(getattr(cell, "_hyperlink", None)),
                comment=copy(getattr(cell, "comment", None)),
            )
        )

    row_dimension_snapshots: list[tuple[int, int, object]] = []
    for row, dimension in list(ws.row_dimensions.items()):
        if not region_start <= row <= region_end:
            continue
        block_index = _row_block_index(source_blocks, row)
        if block_index is None:
            continue
        block = source_blocks[block_index]
        row_dimension_snapshots.append(
            (block_index, row - block.start_row, copy(dimension))
        )

    drawing_moves: list[tuple[object, int]] = []
    for drawing in [*getattr(ws, "_images", []), *getattr(ws, "_charts", [])]:
        anchor = drawing.anchor
        if isinstance(anchor, str):
            row, _col = coordinate_to_tuple(anchor)
        else:
            start = getattr(anchor, "_from", None)
            if start is None:
                continue
            row = int(start.row) + 1
        block_index = _row_block_index(source_blocks, row)
        if block_index is None:
            continue
        block = source_blocks[block_index]
        drawing_moves.append(
            (drawing, target_starts[block_index] - block.start_row)
        )

    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= region_start and rng.min_row <= region_end:
            ws.unmerge_cells(str(rng))
    for key in [
        key for key in ws._cells if region_start <= key[0] <= region_end
    ]:
        del ws._cells[key]
    for row in [
        row for row in ws.row_dimensions if region_start <= row <= region_end
    ]:
        del ws.row_dimensions[row]

    for snapshot in cell_snapshots:
        row = target_starts[snapshot.block_index] + snapshot.row_offset
        cell = ws.cell(row, snapshot.column)
        cell._value = copy(snapshot.value)
        cell.data_type = snapshot.data_type
        cell._style = copy(snapshot.style)
        cell._hyperlink = copy(snapshot.hyperlink)
        if cell._hyperlink is not None:
            cell._hyperlink.ref = cell.coordinate
        cell.comment = copy(snapshot.comment)

    for block_index, row_offset, dimension in row_dimension_snapshots:
        row = target_starts[block_index] + row_offset
        dimension.index = row
        ws.row_dimensions[row] = dimension

    for rng in moved_merged_ranges:
        ws.merge_cells(str(rng))
    for drawing, row_delta in drawing_moves:
        _shift_drawing_anchor_rows(drawing, row_delta)
    return True


def _require_temperature_order_number(
    ws: Worksheet,
    value: object,
    temperature_labels: TemperatureLabels | None = None,
) -> float:
    number = _temperature_number_for_report_order(value, temperature_labels)
    if number is None:
        raise ValueError(
            f"报告工作表“{ws.title}”存在无法识别的温度：{value!r}；"
            "为避免数据与图片错位，已停止写入"
        )
    return number


def _normalize_dpt_data_temperature_order(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    groups = _dpt_data_groups(ws)
    blocks: list[_WorksheetRowBlock] = []
    for index, group in enumerate(groups):
        if index + 1 < len(groups):
            end_row = groups[index + 1].start_row - 1
        else:
            phase_range = _merged_range_containing(ws, group.start_row, 1)
            end_row = (
                phase_range.max_row
                if phase_range is not None
                else group.end_row - 1
            )
        number = _require_temperature_order_number(
            ws,
            group.temp,
            temperature_labels,
        )
        blocks.append(
            _WorksheetRowBlock(
                start_row=group.start_row,
                end_row=end_row,
                sort_key=(
                    _temperature_report_order_key(number),
                    _phase_report_order(group.phase),
                ),
            )
        )
    return _reorder_worksheet_row_blocks(ws, blocks)


def _normalize_dpt_waveform_temperature_order(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    labels = [
        (rng, parts)
        for rng, value in _left_merged_ranges_with_value(ws)
        if (parts := _phase_temp_token_parts(value)) is not None
        and parts[0] in PHASE_CODES
    ]
    blocks: list[_WorksheetRowBlock] = []
    for index, (rng, parts) in enumerate(labels):
        number = _require_temperature_order_number(
            ws,
            parts[1],
            temperature_labels,
        )
        end_row = (
            labels[index + 1][0].min_row - 1
            if index + 1 < len(labels)
            else rng.min_row + DPT_WAVEFORM_BLOCK_STRIDE - 1
        )
        blocks.append(
            _WorksheetRowBlock(
                start_row=rng.min_row,
                end_row=end_row,
                sort_key=(
                    _temperature_report_order_key(number),
                    _phase_report_order(parts[0]),
                ),
            )
        )
    return _reorder_worksheet_row_blocks(ws, blocks)


def _sync_dpt_condition_cell(
    data_ws: Worksheet,
    group_start_row: int,
    signature: Mapping[str, float],
) -> None:
    if not signature:
        return
    current = _merged_value(data_ws, group_start_row, COL_CONDITION)
    merged = _merged_condition_text(current, signature)
    if merged:
        _set_merged_cell_value(data_ws, group_start_row, COL_CONDITION, merged)


def _next_dpt_data_row(
    data_ws: Worksheet,
    start_row: int,
    end_row_exclusive: int,
    *,
    setpoints: _DptSetpointMatch,
) -> tuple[int, bool]:
    for row in range(start_row, end_row_exclusive):
        if _data_row_matches_setpoints(data_ws, row, setpoints):
            return row, False
    for row in range(start_row, end_row_exclusive):
        if not _data_row_has_payload(data_ws, row):
            return row, False
    return _insert_dpt_data_row(data_ws, start_row, end_row_exclusive), True


def _dpt_group_containing(data_ws: Worksheet, group_start_row: int) -> _DptDataGroup | None:
    for group in _dpt_data_groups(data_ws):
        if group.start_row == group_start_row:
            return group
        if group.start_row < group_start_row < group.end_row:
            return group
    return None


def _prepare_dpt_data_rows(
    data_ws: Worksheet,
    target: _DptDataTarget,
    results: Sequence[ExtractResult],
) -> list[_DptDataTarget]:
    """Reserve consecutive report rows without overwriting unrelated conditions."""
    if not results:
        return []
    targets = [target]
    for offset in range(1, len(results)):
        setpoints = _dpt_setpoint_match(results[offset])
        row = target.row + offset
        group = _dpt_group_containing(data_ws, target.group_start_row)
        needs_insert = group is None or row >= group.end_row
        if not needs_insert and _data_row_has_payload(data_ws, row):
            needs_insert = not _data_row_matches_setpoints(data_ws, row, setpoints)
        if needs_insert:
            row = _insert_dpt_data_row(data_ws, target.group_start_row, row)
        targets.append(
            _DptDataTarget(
                row=row,
                group_index=target.group_index,
                group_start_row=target.group_start_row,
                row_offset=row - target.group_start_row,
                inserted_row=needs_insert,
            )
        )
    return targets


def _insert_dpt_temperature_group(
    data_ws: Worksheet,
    source_group: _DptDataGroup,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
    *,
    insert_row: int | None = None,
) -> None:
    """Clone an empty four-row template group for another actual temperature."""

    source_start = source_group.start_row
    if insert_row is None:
        insert_row = source_group.end_row
    row_count = DPT_DATA_GROUP_TEMPLATE_ROWS
    merged_ranges = [CellRange(str(rng)) for rng in data_ws.merged_cells.ranges]
    cloned_ranges = [
        CellRange(
            min_col=rng.min_col,
            max_col=rng.max_col,
            min_row=insert_row + (rng.min_row - source_start),
            max_row=insert_row + (rng.max_row - source_start),
        )
        for rng in merged_ranges
        if source_start <= rng.min_row
        and rng.max_row < source_start + row_count
    ]
    for rng in list(data_ws.merged_cells.ranges):
        data_ws.unmerge_cells(str(rng))

    data_ws.insert_rows(insert_row, row_count)
    _shift_row_dimensions_for_insert(data_ws, insert_row, row_count)
    copied_source_start = (
        source_start + row_count if insert_row <= source_start else source_start
    )
    max_column = max(data_ws.max_column, LAST_COL)
    for offset in range(row_count):
        source_row = copied_source_start + offset
        target_row = insert_row + offset
        _copy_row_style(data_ws, source_row, target_row)
        for col in range(1, max_column + 1):
            source = data_ws.cell(source_row, col)
            target = data_ws.cell(target_row, col)
            target.value = copy(source.value) if col <= COL_CONDITION else None

    for rng in merged_ranges:
        shifted = (
            CellRange(
                min_col=rng.min_col,
                max_col=rng.max_col,
                min_row=rng.min_row + row_count,
                max_row=rng.max_row + row_count,
            )
            if rng.min_row >= insert_row
            else rng
        )
        data_ws.merge_cells(str(shifted))
    for rng in cloned_ranges:
        data_ws.merge_cells(str(rng))
    cloned_range_texts = {str(rng) for rng in cloned_ranges}
    for col in (1, 2, COL_CONDITION):
        source_header = next(
            (
                rng
                for rng in merged_ranges
                if rng.min_row == source_start
                and rng.min_col == col
                and rng.max_col == col
                and rng.max_row >= source_start + row_count - 1
            ),
            None,
        )
        if source_header is None:
            continue
        target_header = CellRange(
            min_col=col,
            max_col=col,
            min_row=insert_row,
            max_row=insert_row + row_count - 1,
        )
        if str(target_header) not in cloned_range_texts:
            data_ws.merge_cells(str(target_header))

    _set_merged_cell_value(data_ws, insert_row, 1, phase_code.upper())
    _set_merged_cell_value(
        data_ws,
        insert_row,
        2,
        _temperature_display(temp_code, temperature_labels),
    )


def _dpt_temperature_group_insert_row(
    groups: Sequence[_DptDataGroup],
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
    *,
    fallback_row: int,
) -> int:
    temperature = _temperature_number_for_report_order(
        _temperature_display(temp_code, temperature_labels),
        temperature_labels,
    )
    if temperature is None:
        return fallback_row
    target_key = (
        _temperature_report_order_key(temperature),
        _phase_report_order(phase_code),
    )
    last_known_end: int | None = None
    for group in groups:
        group_temperature = _temperature_number_for_report_order(
            group.temp,
            temperature_labels,
        )
        if group_temperature is None:
            continue
        group_key = (
            _temperature_report_order_key(group_temperature),
            _phase_report_order(group.phase),
        )
        if group_key > target_key:
            return group.start_row
        last_known_end = group.end_row
    return last_known_end if last_known_end is not None else fallback_row


def _ensure_dpt_temperature_group(
    data_ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    groups = _dpt_data_groups(data_ws)
    exact_groups = [
        group
        for group in groups
        if group.phase.upper() == phase_code.upper()
        and _temperature_cell_matches_display(
            group.temp,
            temp_code,
            temperature_labels,
        )
    ]
    if exact_groups:
        return

    seed_groups = [
        (index, group)
        for index, group in enumerate(groups)
        if group.phase.upper() == phase_code.upper()
        and _temperature_cell_matches(group.temp, temp_code, temperature_labels)
    ]
    if not seed_groups:
        display_temp = _temperature_display(temp_code, temperature_labels)
        raise ValueError(
            f"报告模板工作表“{data_ws.title}”缺少工况组："
            f"{phase_code.upper()} / {display_temp}；为避免写入错误行，已停止写入"
        )
    _source_index, source_group = seed_groups[-1]
    insert_row = _dpt_temperature_group_insert_row(
        groups,
        phase_code,
        temp_code,
        temperature_labels,
        fallback_row=source_group.end_row,
    )
    _insert_dpt_temperature_group(
        data_ws,
        source_group,
        phase_code,
        temp_code,
        temperature_labels,
        insert_row=insert_row,
    )


def _find_dpt_data_target(
    result: ExtractResult,
    data_ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> _DptDataTarget:
    setpoints = _dpt_setpoint_match(result)
    source_condition = _condition_signature(result.source_path)
    matching_groups: list[tuple[int, _DptDataGroup]] = []
    for group_index, group in enumerate(_dpt_data_groups(data_ws)):
        if group.phase.upper() != phase_code.upper():
            continue
        if not _temperature_cell_matches_display(
            group.temp,
            temp_code,
            temperature_labels,
        ):
            continue
        matching_groups.append((group_index, group))

    if source_condition:
        condition_groups = [
            item
            for item in matching_groups
            if _condition_signature_matches(
                source_condition,
                _condition_signature(item[1].condition),
            )
        ]
        if condition_groups:
            matching_groups = condition_groups

    for group_index, group in matching_groups:
        row, inserted = _next_dpt_data_row(
            data_ws,
            group.start_row,
            group.end_row,
            setpoints=setpoints,
        )
        target = _DptDataTarget(
            row=row,
            group_index=group_index,
            group_start_row=group.start_row,
            row_offset=row - group.start_row,
            inserted_row=inserted,
        )
        _set_merged_cell_value(
            data_ws,
            group.start_row,
            2,
            _temperature_display(temp_code, temperature_labels),
        )
        _sync_dpt_condition_cell(data_ws, group.start_row, source_condition)
        return target

    display_temp = _temperature_display(temp_code, temperature_labels)
    raise ValueError(
        f"报告模板工作表“{data_ws.title}”缺少工况组："
        f"{phase_code.upper()} / {display_temp}；为避免写入错误行，已停止写入"
    )


def _normalize_dpt_data_temperature_cells(
    data_ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    for group in _dpt_data_groups(data_ws):
        raw = str(group.temp or "").strip().upper()
        if raw in TEMP_LABELS:
            _set_merged_cell_value(
                data_ws,
                group.start_row,
                2,
                _temperature_display(raw, temperature_labels),
            )


def _dpt_waveform_group_base(data_ws: Worksheet, group_index: int) -> int:
    base = 1
    for group in _dpt_data_groups(data_ws)[:group_index]:
        base += _dpt_group_payload_row_count(data_ws, group) * DPT_WAVEFORM_BLOCK_STRIDE
    return base


def _dpt_group_payload_row_count(data_ws: Worksheet, group: _DptDataGroup) -> int:
    return sum(
        1
        for row in range(group.start_row, group.end_row)
        if _data_row_has_payload(data_ws, row)
    )


def _dpt_waveform_row_index(data_ws: Worksheet, group_start_row: int, row: int) -> int:
    return sum(
        1
        for candidate in range(group_start_row, row)
        if _data_row_has_payload(data_ws, candidate)
    )


def _copy_waveform_row(ws: Worksheet, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target.value = source.value
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def _shift_range_rows(rng: CellRange, row_delta: int) -> CellRange:
    return CellRange(
        min_col=rng.min_col,
        max_col=rng.max_col,
        min_row=rng.min_row + row_delta,
        max_row=rng.max_row + row_delta,
    )


def _shift_image_anchors_for_insert(
    ws: Worksheet,
    insert_row: int,
    amount: int,
) -> None:
    """Keep existing drawing anchors aligned when waveform rows are inserted."""
    for image in getattr(ws, "_images", []):
        anchor = image.anchor
        if isinstance(anchor, str):
            row, col = coordinate_to_tuple(anchor)
            if row >= insert_row:
                image.anchor = f"{get_column_letter(col)}{row + amount}"
            continue

        start = getattr(anchor, "_from", None)
        if start is None or int(start.row) + 1 < insert_row:
            continue
        start.row = int(start.row) + amount
        end = getattr(anchor, "to", None)
        if end is not None:
            end.row = int(end.row) + amount


def _insert_dpt_waveform_block(ws: Worksheet, insert_row: int) -> None:
    amount = DPT_WAVEFORM_BLOCK_STRIDE
    source_row = (
        insert_row - amount
        if insert_row > amount
        else max(1, insert_row)
    )
    source_end = source_row + amount - 1
    max_row_before = ws.max_row
    merged_ranges = [CellRange(str(rng)) for rng in ws.merged_cells.ranges]
    copied_ranges = [
        _shift_range_rows(rng, insert_row - source_row)
        for rng in merged_ranges
        if source_row <= rng.min_row and rng.max_row <= source_end
    ]

    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    if insert_row <= ws.max_row:
        ws.insert_rows(insert_row, amount)
        _shift_row_dimensions_for_insert(ws, insert_row, amount)
        _shift_image_anchors_for_insert(ws, insert_row, amount)
    copied_source_row = (
        source_row + amount if insert_row <= source_row else source_row
    )
    for offset in range(amount):
        row_to_copy = copied_source_row + offset
        if row_to_copy > max_row_before and offset >= DPT_WAVEFORM_BLOCK_ROWS:
            row_to_copy = max(1, copied_source_row - (amount - offset))
        _copy_waveform_row(ws, row_to_copy, insert_row + offset)

    restored: list[CellRange] = []
    for rng in merged_ranges:
        if rng.min_row >= insert_row:
            restored.append(_shift_range_rows(rng, amount))
        elif rng.min_row < insert_row <= rng.max_row:
            restored.append(
                CellRange(
                    min_col=rng.min_col,
                    max_col=rng.max_col,
                    min_row=rng.min_row,
                    max_row=rng.max_row + amount,
                )
            )
        else:
            restored.append(rng)

    seen: set[tuple[int, int, int, int]] = set()
    for rng in [*restored, *copied_ranges]:
        key = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        if key in seen:
            continue
        seen.add(key)
        ws.merge_cells(
            start_row=rng.min_row,
            end_row=rng.max_row,
            start_column=rng.min_col,
            end_column=rng.max_col,
        )


def _set_range_value(ws: Worksheet, rng: CellRange, value: object) -> None:
    ws.cell(rng.min_row, rng.min_col).value = value


def _dpt_waveform_label(ws: Worksheet, row: int) -> str:
    return str(_merged_value(ws, row, 1) or "").strip()


def _dpt_waveform_block_has_images(ws: Worksheet, anchor_row: int) -> bool:
    block = CellRange(
        min_col=1,
        max_col=max(1, ws.max_column),
        min_row=anchor_row,
        max_row=anchor_row + DPT_WAVEFORM_BLOCK_ROWS - 1,
    )
    for image in getattr(ws, "_images", []):
        pos = _image_anchor_position(image)
        if pos is None:
            continue
        row, col = pos
        if block.min_row <= row <= block.max_row and block.min_col <= col <= block.max_col:
            return True
    return False


def _phase_temp_parts(
    label: str,
    temperature_labels: TemperatureLabels | None = None,
) -> tuple[str, str] | None:
    match = re.fullmatch(r"\s*([UVW][HL])\s*_?\s*(.+?)\s*", str(label or ""), re.I)
    if match is None:
        return None
    temp_code = _temperature_code_from_token(match.group(2), temperature_labels)
    if temp_code is None:
        return None
    return match.group(1).upper(), temp_code


def _phase_temp_token_parts(label: str) -> tuple[str, str] | None:
    match = re.fullmatch(r"\s*([UVW][HL])\s*_?\s*(.+?)\s*", str(label or ""), re.I)
    if match is None:
        return None
    return match.group(1).upper(), match.group(2).strip()


def _phase_temperature_matches_display(
    label: str,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    parts = _phase_temp_token_parts(label)
    if parts is None or parts[0] != phase_code.upper():
        return False
    return _temperature_cell_matches_display(
        parts[1],
        temp_code,
        temperature_labels,
    )


def _dpt_waveform_block_has_data_match(
    data_ws: Worksheet,
    phase_temp_label: str,
    condition_label: str | None,
    temperature_labels: TemperatureLabels | None = None,
    *,
    require_condition_signature: bool = True,
) -> bool:
    phase_temp = _phase_temp_token_parts(phase_temp_label)
    if phase_temp is None:
        return False
    phase_code, temperature_token = phase_temp
    setpoints = _condition_setpoint_match(condition_label)
    if setpoints.vdc is None or setpoints.idc is None:
        return False
    condition_signature = _condition_signature(condition_label)
    for group in _dpt_data_groups(data_ws):
        if group.phase.upper() != phase_code:
            continue
        if not _temperature_tokens_match(
            group.temp,
            temperature_token,
            temperature_labels,
        ):
            continue
        if (
            require_condition_signature
            and condition_signature
            and not _condition_signature_matches(
                condition_signature,
                _condition_signature(group.condition),
            )
        ):
            continue
        for row in range(group.start_row, group.end_row):
            if _data_row_matches_setpoints(data_ws, row, setpoints):
                return True
    return False


def _dpt_waveform_block_is_reusable(
    ws: Worksheet,
    data_ws: Worksheet,
    anchor_row: int,
    existing_label: str,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    if _dpt_waveform_block_has_images(ws, anchor_row):
        return False
    condition_label = _dpt_waveform_label(ws, anchor_row + 17)
    return not _dpt_waveform_block_has_data_match(
        data_ws,
        existing_label,
        condition_label,
        temperature_labels,
    )


def _ensure_dpt_waveform_anchor(
    ws: Worksheet,
    data_ws: Worksheet,
    target: _DptDataTarget,
    phase_code: str,
    temp_code: str,
    vdc: float | None,
    idc: float | None,
    condition_signature: Mapping[str, float] | None = None,
    temperature_labels: TemperatureLabels | None = None,
) -> int:
    group_base = _dpt_waveform_group_base(data_ws, target.group_index)
    row_index = _dpt_waveform_row_index(data_ws, target.group_start_row, target.row)
    anchor_row = group_base + row_index * DPT_WAVEFORM_BLOCK_STRIDE
    phase_temp = _phase_temp_label(phase_code, temp_code, temperature_labels)
    existing_label = _dpt_waveform_label(ws, anchor_row)
    existing_matches_target = _phase_temperature_matches_display(
        existing_label,
        phase_code,
        temp_code,
        temperature_labels,
    )
    inserted_row_occupied = target.inserted_row and bool(existing_label)
    mismatched_live_block = (
        existing_label
        and not existing_matches_target
        and (
            target.inserted_row
            or not _dpt_waveform_block_is_reusable(
                ws,
                data_ws,
                anchor_row,
                existing_label,
                temperature_labels,
            )
        )
    )
    if (
        anchor_row + DPT_WAVEFORM_BLOCK_ROWS - 1 > ws.max_row
        or inserted_row_occupied
        or mismatched_live_block
    ):
        _insert_dpt_waveform_block(ws, anchor_row)

    _set_range_value(ws, _left_merged_range_at(ws, anchor_row), phase_temp)
    _set_range_value(
        ws,
        _left_merged_range_at(ws, anchor_row + 17),
        _format_condition_key(vdc, idc, condition_signature),
    )
    _set_range_value(ws, _left_merged_range_at(ws, anchor_row + 34), "总概览图")
    return anchor_row


def _sync_dpt_waveform_insertions(
    ws: Worksheet,
    data_ws: Worksheet,
    targets: Sequence[_DptDataTarget],
) -> None:
    """Mirror every inserted data row with one waveform block insertion."""

    for target in sorted(
        (item for item in targets if item.inserted_row),
        key=lambda item: (item.group_index, item.row),
    ):
        group_base = _dpt_waveform_group_base(data_ws, target.group_index)
        row_index = _dpt_waveform_row_index(
            data_ws,
            target.group_start_row,
            target.row,
        )
        _insert_dpt_waveform_block(
            ws,
            group_base + row_index * DPT_WAVEFORM_BLOCK_STRIDE,
        )


def _set_value(ws: Worksheet, row: int, col: int, value) -> None:
    if value is None:
        return
    if isinstance(value, str) and not value.strip():
        return
    ws.cell(row, col, value)


def _normalize_dpt_written_data_row_style(ws: Worksheet, row: int) -> None:
    """Keep DPT report value cells visually consistent even on older templates."""
    max_col = max(ws.max_column, LAST_COL)
    for col in range(COL_VOLTAGE, min(max_col, LAST_COL) + 1):
        cell = ws.cell(row, col)
        font = copy(cell.font)
        font.sz = _DPT_REPORT_DATA_FONT_SIZE
        cell.font = font
        alignment = copy(cell.alignment)
        alignment.horizontal = "center"
        alignment.vertical = "center"
        cell.alignment = alignment


def _set_metric_value(
    ws: Worksheet,
    row: int,
    col: int,
    result: ExtractResult,
    section: str,
    name: str,
    value,
) -> None:
    if result.is_metric_unavailable(section, name):
        ws.cell(row, col).value = None
        return
    _set_value(ws, row, col, value)


_DPT_COLUMN_BOUNDS = {
    "关断过程": (6, 17),
    "开通": (18, 30),
    "反向恢复": (31, 37),
    "汇总": (38, 40),
}


def _dpt_header_row_has_values(ws: Worksheet) -> bool:
    for col in range(1, min(ws.max_column, 40) + 1):
        if str(ws.cell(HEADER_NAME_ROW, col).value or "").strip():
            return True
    return False


def _dpt_header_col(
    ws: Worksheet,
    section: str,
    headers: tuple[str, ...],
) -> int | None:
    lo, hi = _DPT_COLUMN_BOUNDS.get(section, (1, min(ws.max_column, 40)))
    hi = min(hi, max(ws.max_column, hi))
    normalized_headers = tuple(_normalized(header) for header in headers)
    for col in range(lo, hi + 1):
        text = _normalized(ws.cell(HEADER_NAME_ROW, col).value)
        if not text:
            continue
        if any(text == header or text.startswith(header) for header in normalized_headers if header):
            return col
    return None


def _dpt_metric_col(
    ws: Worksheet,
    section: str,
    fallback_col: int,
    *headers: str,
    require_header: bool = False,
) -> int | None:
    found = _dpt_header_col(ws, section, tuple(headers))
    if found is not None:
        return found
    if require_header and _dpt_header_row_has_values(ws):
        return None
    return fallback_col


def _set_dpt_metric_value(
    ws: Worksheet,
    row: int,
    col: int | None,
    result: ExtractResult,
    section: str,
    name: str,
    value,
) -> None:
    if col is None:
        return
    _set_metric_value(ws, row, col, result, section, name, value)


def _copy_column_style(ws: Worksheet, source_col: int, target_col: int) -> None:
    max_row = max(ws.max_row, HEADER_UNIT_ROW)
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row in range(1, max_row + 1):
        src = ws.cell(row, source_col)
        dst = ws.cell(row, target_col)
        dst._style = copy(src._style)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def _unmerge_dpt_header_rows(ws: Worksheet) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= 2 and rng.max_row <= 2 and rng.min_col >= 1:
            ws.unmerge_cells(str(rng))


def _merge_dpt_data_headers_like_template(ws: Worksheet) -> None:
    for spec in ("A1:E2", "F1:Q2", "R1:AK1", "R2:AD2", "AE2:AK2"):
        try:
            ws.merge_cells(spec)
        except ValueError:
            pass
    ws.cell(1, 1).value = ws.cell(1, 1).value or "信息"
    ws.cell(1, 6).value = "关断过程"
    ws.cell(1, 18).value = "开通过程"
    ws.cell(2, 18).value = "开通"
    ws.cell(2, 31).value = "反向恢复"


def _rewrite_dpt_header_row(ws: Worksheet) -> None:
    for col, name, unit, _fill in COLUMNS:
        ws.cell(HEADER_NAME_ROW, col).value = name
        unit_cell = ws.cell(HEADER_UNIT_ROW, col)
        if not isinstance(unit_cell, MergedCell):
            unit_cell.value = unit


def _ensure_dpt_data_power_columns(ws: Worksheet) -> bool:
    """Upgrade old A:AK DPT data sheets to the Pmax/Pdmax A:AN layout in-place."""
    present = {
        section: _dpt_header_col(
            ws,
            section,
            tuple(dict.fromkeys((power_metric_name(section), "Pdmax"))),
        )
        is not None
        for section in ("关断过程", "开通", "反向恢复")
    }
    if all(present.values()):
        needs_header_rewrite = any(
            _dpt_header_col(ws, section, (power_metric_name(section),)) is None
            for section in ("关断过程", "开通", "反向恢复")
        )
        if needs_header_rewrite:
            _rewrite_dpt_header_row(ws)
        return needs_header_rewrite

    _unmerge_dpt_header_rows(ws)
    insertion_plan = [
        ("反向恢复", "Err"),
        ("开通", "Eon"),
        ("关断过程", "Eoff"),
    ]
    for section, loss_header in insertion_plan:
        if present.get(section):
            continue
        loss_col = _dpt_header_col(ws, section, (loss_header,))
        if loss_col is None:
            continue
        ws.insert_cols(loss_col)
        _copy_column_style(ws, loss_col + 1, loss_col)
        ws.cell(HEADER_NAME_ROW, loss_col).value = power_metric_name(section)
        ws.cell(HEADER_UNIT_ROW, loss_col).value = "KW"
        for row in range(DATA_ROW, ws.max_row + 1):
            ws.cell(row, loss_col).value = None

    _merge_dpt_data_headers_like_template(ws)
    _rewrite_dpt_header_row(ws)
    return True


def _write_dpt_data(ws: Worksheet, row: int, result: ExtractResult) -> None:
    off = result.turn_off
    on = result.turn_on
    rr = result.reverse_recovery
    vdc, idc = _match_setpoints(result)

    def col(section: str, fallback_col: int, *headers: str, require_header: bool = False) -> int | None:
        return _dpt_metric_col(
            ws,
            section,
            fallback_col,
            *headers,
            require_header=require_header,
        )

    _set_value(ws, row, 4, _num(vdc, 1))
    _set_value(ws, row, 5, _num(idc, 1))
    _set_dpt_metric_value(
        ws,
        row,
        col("关断过程", COL_OFF["delta_vce"], "ΔVce"),
        result,
        "关断过程",
        "ΔVce",
        _num(off.delta_vce),
    )
    _set_dpt_metric_value(
        ws,
        row,
        col("关断过程", COL_OFF["ic_off_max"], "Ic_off_max"),
        result,
        "关断过程",
        "Ic_off_max",
        _num(off.ic_off_max),
    )
    _set_dpt_metric_value(
        ws,
        row,
        col("关断过程", COL_OFF["vce_off_max"], "Vce_off_max"),
        result,
        "关断过程",
        "Vce_off_max",
        _num(off.vce_off_max),
    )
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["dvdt"], "dv/dt"), result, "关断过程", "dv/dt", _num(off.dvdt))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["didt"], "di/dt"), result, "关断过程", "di/dt", _num(off.didt))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["ls_off"], "Ls_off"), result, "关断过程", "Ls_off", _num(off.ls_off))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["toff"], "Toff"), result, "关断过程", "Toff", _num(off.toff))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["td_off"], "Td_off"), result, "关断过程", "Td_off", _num(off.td_off))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["tf"], "Tf"), result, "关断过程", "Tf", _num(off.tf))
    _set_dpt_metric_value(
        ws,
        row,
        col("关断过程", COL_OFF["crosstalk"], "串扰电压"),
        result,
        "关断过程",
        "串扰电压",
        _crosstalk_str(off.crosstalk_vmax, off.crosstalk_vmin),
    )
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["pmax"], "Pmax", "Pdmax", require_header=True), result, "关断过程", "Pmax", _num(off.pmax, 3))
    _set_dpt_metric_value(ws, row, col("关断过程", COL_OFF["eoff"], "Eoff"), result, "关断过程", "Eoff", _num(off.eoff, 3))

    _set_dpt_metric_value(ws, row, col("开通", COL_ON["delta_vce"], "ΔVce"), result, "开通", "ΔVce", _num(on.delta_vce))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["ic_on_max"], "Ic_on_max"), result, "开通", "Ic_on_max", _num(on.ic_on_max))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["vce_on_max"], "Vce_on_max"), result, "开通", "Vce_on_max", _num(on.vce_on_max))
    _set_dpt_metric_value(
        ws,
        row,
        col("开通", COL_ON["turn_on_current"], "开通电流"),
        result,
        "开通",
        "开通电流",
        _num(on.turn_on_current),
    )
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["dvdt"], "dv/dt"), result, "开通", "dv/dt", _num(on.dvdt))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["didt"], "di/dt"), result, "开通", "di/dt", _num(on.didt))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["ls_on"], "Ls_on"), result, "开通", "Ls_on", _num(on.ls_on))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["ton"], "Ton"), result, "开通", "Ton", _num(on.ton))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["td_on"], "Td_on"), result, "开通", "Td_on", _num(on.td_on))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["tr"], "Tr"), result, "开通", "Tr", _num(on.tr))
    _set_dpt_metric_value(
        ws,
        row,
        col("开通", COL_ON["crosstalk"], "串扰电压"),
        result,
        "开通",
        "串扰电压",
        _crosstalk_str(on.crosstalk_vmax, on.crosstalk_vmin),
    )
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["pmax"], "Pmax", "Pdmax", require_header=True), result, "开通", "Pmax", _num(on.pmax, 3))
    _set_dpt_metric_value(ws, row, col("开通", COL_ON["eon"], "Eon"), result, "开通", "Eon", _num(on.eon, 3))

    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["irr"], "Irr"), result, "反向恢复", "Irr", _num(rr.irr))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["trr"], "Trr"), result, "反向恢复", "Trr", _num(rr.trr))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["vrr"], "Vrr"), result, "反向恢复", "Vrr", _num(rr.vrr))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["dvdt"], "Dvdt_max", "dv/dt"), result, "反向恢复", "dv/dt", _num(rr.dvdt_max))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["didt"], "Didt_Irr", "di/dt"), result, "反向恢复", "di/dt", _num(rr.didt_irr))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["pdmax"], "Pdmax", require_header=True), result, "反向恢复", "Pdmax", _num(rr.pdmax, 3))
    _set_dpt_metric_value(ws, row, col("反向恢复", COL_RR["err"], "Err"), result, "反向恢复", "Err", _num(rr.err, 3))

    etotal = off.eoff + on.eon + rr.err
    if any(
        result.is_metric_unavailable(*key)
        for key in (("关断过程", "Eoff"), ("开通", "Eon"), ("反向恢复", "Err"))
    ):
        tail_col = col("汇总", COL_TAIL["etotal"], "Etotal（all）", "Etotal")
        if tail_col is not None:
            ws.cell(row, tail_col).value = None
    else:
        tail_col = col("汇总", COL_TAIL["etotal"], "Etotal（all）", "Etotal")
        if tail_col is not None:
            _set_value(ws, row, tail_col, _num(etotal, 3) if etotal > 0 else None)
    _normalize_dpt_written_data_row_style(ws, row)


def _normalized(text: object) -> str:
    return (
        str(text or "")
        .upper()
        .replace("\n", "")
        .replace(" ", "")
        .replace("（", "(")
        .replace("）", ")")
        .replace("△", "Δ")
    )


def _header_matches(value: object, header_text: str) -> bool:
    needle = _normalized(header_text)
    haystack = _normalized(value)
    return bool(needle) and haystack.startswith(needle)


def _find_header_range(ws: Worksheet, header_row: int, header_text: str) -> CellRange | None:
    for rng in ws.merged_cells.ranges:
        if rng.min_row == header_row and rng.max_row == header_row:
            if _header_matches(ws.cell(rng.min_row, rng.min_col).value, header_text):
                return rng
    for col in range(1, ws.max_column + 1):
        if _header_matches(ws.cell(header_row, col).value, header_text):
            return CellRange(
                min_col=col,
                max_col=col,
                min_row=header_row,
                max_row=header_row,
            )
    return None


def _picture_range_below(
    ws: Worksheet,
    header_range: CellRange,
    *,
    rows: int,
) -> CellRange:
    min_row = header_range.min_row + 1
    max_row = header_range.min_row + rows
    for rng in ws.merged_cells.ranges:
        if (
            rng.min_col == header_range.min_col
            and rng.max_col == header_range.max_col
            and rng.min_row == min_row
            and rng.max_row == max_row
        ):
            return rng
    return CellRange(
        min_col=header_range.min_col,
        max_col=header_range.max_col,
        min_row=min_row,
        max_row=max_row,
    )


def _column_width_px(ws: Worksheet, col: int) -> int:
    letter = get_column_letter(col)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43
    return max(8, int(float(width) * 7 + 5))


def _row_height_px(ws: Worksheet, row: int) -> int:
    height = ws.row_dimensions[row].height
    if height is None:
        height = 15.0
    return max(8, int(float(height) * 96 / 72))


def _range_size_px(ws: Worksheet, rng: CellRange) -> tuple[int, int]:
    width = sum(_column_width_px(ws, col) for col in range(rng.min_col, rng.max_col + 1))
    height = sum(_row_height_px(ws, row) for row in range(rng.min_row, rng.max_row + 1))
    return max(1, width - 4), max(1, height - 4)


def _set_column_width_px(ws: Worksheet, col: int, px: int) -> None:
    letter = get_column_letter(col)
    width = (max(8, px) - 5) / 7
    ws.column_dimensions[letter].width = max(2.5, width)


def _set_row_height_px(ws: Worksheet, row: int, px: int) -> None:
    ws.row_dimensions[row].height = max(1.0, max(8, px) * 72 / 96)


def _style_text_cell(
    ws: Worksheet,
    rng: CellRange,
    *,
    font_size: float,
    shrink_to_fit: bool = True,
    bold: bool | None = None,
) -> None:
    cell = ws.cell(rng.min_row, rng.min_col)
    font = copy(cell.font)
    font.sz = font_size
    if bold is not None:
        font.bold = bold
    cell.font = font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
        shrink_to_fit=shrink_to_fit,
    )


def _merged_range_containing(ws: Worksheet, row: int, col: int) -> CellRange | None:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def _style_cell_at(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    font_size: float,
    shrink_to_fit: bool = True,
) -> None:
    rng = _merged_range_containing(ws, row, col)
    if rng is None:
        rng = CellRange(min_col=col, max_col=col, min_row=row, max_row=row)
    if ws.cell(rng.min_row, rng.min_col).value is None:
        return
    _style_text_cell(
        ws,
        rng,
        font_size=font_size,
        shrink_to_fit=shrink_to_fit,
    )


def _all_header_ranges_by_texts(ws: Worksheet, header_texts) -> list[CellRange]:
    needles = [_normalized(text) for text in header_texts if _normalized(text)]
    ranges: list[CellRange] = []
    seen: set[tuple[int, int, int, int]] = set()

    def matches(value: object) -> bool:
        return any(_header_matches(value, needle) for needle in needles)

    for header_rng in ws.merged_cells.ranges:
        if header_rng.min_row != header_rng.max_row:
            continue
        if not matches(ws.cell(header_rng.min_row, header_rng.min_col).value):
            continue
        _add_unique_range(ranges, seen, header_rng)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if _cell_in_merged_range(ws, row, col):
                continue
            if not matches(ws.cell(row, col).value):
                continue
            _add_unique_range(
                ranges,
                seen,
                CellRange(min_col=col, max_col=col, min_row=row, max_row=row),
            )
    return ranges


def _normalize_dpt_waveform_text(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    header_texts = {header_text for _block, header_text in _DPT_IMAGE_HEADERS.values()}
    for rng in _all_header_ranges_by_texts(ws, header_texts):
        _set_row_height_px(ws, rng.min_row, _WAVEFORM_HEADER_ROW_HEIGHT_PX)
        _style_text_cell(
            ws,
            rng,
            font_size=_WAVEFORM_HEADER_FONT_SIZE,
            shrink_to_fit=True,
            bold=True,
        )

    labels = _left_merged_ranges_with_value(ws)
    for phase_rng, phase_label in labels:
        phase_temp = _phase_temp_token_parts(phase_label)
        if phase_temp is None:
            continue
        raw_code = phase_temp[1].strip().upper()
        if raw_code in TEMP_LABELS:
            _set_range_value(
                ws,
                phase_rng,
                _phase_temp_label(phase_temp[0], raw_code, temperature_labels),
            )
        state_col = phase_rng.max_col + 1
        for offset in (0, 17, 34):
            row = phase_rng.min_row + offset
            _set_row_height_px(ws, row, _WAVEFORM_HEADER_ROW_HEIGHT_PX)
            label_rng = _left_merged_range_at(
                ws,
                row,
                fallback_cols=phase_rng.max_col,
                fallback_rows=17,
            )
            _style_text_cell(
                ws,
                label_rng,
                font_size=_WAVEFORM_LEFT_LABEL_FONT_SIZE,
                shrink_to_fit=True,
                bold=True,
            )
            _style_cell_at(
                ws,
                row,
                state_col,
                font_size=_WAVEFORM_STATE_FONT_SIZE,
                shrink_to_fit=True,
            )


def _normalize_short_picture_text(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    for rng in _all_header_ranges_by_texts(ws, _SHORT_IMAGE_HEADERS.values()):
        _set_row_height_px(ws, rng.min_row, _WAVEFORM_HEADER_ROW_HEIGHT_PX)
        _style_text_cell(
            ws,
            rng,
            font_size=_WAVEFORM_HEADER_FONT_SIZE,
            shrink_to_fit=True,
            bold=True,
        )
    for rng, value in _merged_ranges_with_value(ws, min_col=1, max_col=5):
        phase_temp = _phase_temp_token_parts(value)
        if phase_temp is not None:
            raw_code = phase_temp[1].strip().upper()
            if raw_code in TEMP_LABELS:
                _set_range_value(
                    ws,
                    rng,
                    _phase_temp_label(phase_temp[0], raw_code, temperature_labels),
                )
            _style_text_cell(
                ws,
                rng,
                font_size=_WAVEFORM_LEFT_LABEL_FONT_SIZE,
                shrink_to_fit=True,
                bold=True,
            )


def _used_columns_width_px(ws: Worksheet) -> int:
    max_col = ws.max_column
    for rng in ws.merged_cells.ranges:
        max_col = max(max_col, rng.max_col)
    return sum(_column_width_px(ws, col) for col in range(1, max_col + 1))


def _set_open_zoom_for_screen(
    ws: Worksheet,
    target_screen_width_px: int | None,
) -> None:
    screen_width = int(target_screen_width_px or _REPORT_VIEW_DEFAULT_SCREEN_WIDTH_PX)
    if screen_width <= 0:
        screen_width = _REPORT_VIEW_DEFAULT_SCREEN_WIDTH_PX
    usable_width = max(960, screen_width - _REPORT_VIEW_HORIZONTAL_MARGIN_PX)
    used_width = max(1, _used_columns_width_px(ws))
    zoom = int(usable_width * 100 / used_width)
    zoom = max(_REPORT_VIEW_MIN_ZOOM, min(_REPORT_VIEW_MAX_ZOOM, zoom))
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom


def _set_report_open_zoom(wb, target_screen_width_px: int | None) -> None:
    for ws in wb.worksheets:
        _set_open_zoom_for_screen(ws, target_screen_width_px)


def _fit_columns_to_width(ws: Worksheet, rng: CellRange, content_width_px: int) -> None:
    cols = list(range(rng.min_col, rng.max_col + 1))
    if not cols:
        return
    target_total = max(
        8 * len(cols),
        content_width_px + _IMAGE_SLOT_PADDING_PX,
    )
    base = target_total // len(cols)
    remainder = target_total - base * len(cols)
    for idx, col in enumerate(cols):
        _set_column_width_px(ws, col, int(base + (1 if idx < remainder else 0)))


def _fit_rows_to_height(ws: Worksheet, rng: CellRange, content_height_px: int) -> None:
    rows = list(range(rng.min_row, rng.max_row + 1))
    if not rows:
        return
    target_total = max(8 * len(rows), content_height_px + _IMAGE_SLOT_PADDING_PX)
    base = target_total // len(rows)
    remainder = target_total - base * len(rows)
    for idx, row in enumerate(rows):
        _set_row_height_px(ws, row, int(base + (1 if idx < remainder else 0)))


def _fit_range_to_image_size(
    ws: Worksheet,
    rng: CellRange,
    display_size: tuple[int, int],
) -> None:
    _fit_columns_to_width(ws, rng, display_size[0])
    _fit_rows_to_height(ws, rng, display_size[1])


def _fit_columns_to_image_aspect(ws: Worksheet, rng: CellRange, image: XLImage) -> None:
    if not image.width or not image.height:
        return
    _current_w, current_h = _range_size_px(ws, rng)
    _fit_columns_to_width(ws, rng, int(current_h * image.width / image.height))


def _image_aspect(image_path: str | Path) -> float | None:
    path = Path(image_path)
    if not path.exists():
        return None
    image = XLImage(str(path))
    if not image.width or not image.height:
        return None
    return image.width / image.height


def _prepare_uniform_image_slots(
    ws: Worksheet,
    items: list[tuple[str | Path, CellRange]],
    *,
    template_ranges: list[CellRange] | None = None,
) -> tuple[int, int] | None:
    ranges = template_ranges if template_ranges is not None else [rng for _image_path, rng in items]
    if not ranges:
        return None
    for rng in ranges:
        _fit_range_to_image_size(ws, rng, REPORT_IMAGE_DISPLAY_SIZE_PX)
    return REPORT_IMAGE_DISPLAY_SIZE_PX


def _cell_in_merged_range(ws: Worksheet, row: int, col: int) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return True
    return False


def _add_unique_range(
    ranges: list[CellRange],
    seen: set[tuple[int, int, int, int]],
    rng: CellRange,
) -> None:
    key = (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
    if key in seen:
        return
    seen.add(key)
    ranges.append(rng)


def _all_picture_ranges_by_headers(
    ws: Worksheet,
    header_texts,
    *,
    rows: int,
) -> list[CellRange]:
    needles = [_normalized(text) for text in header_texts if _normalized(text)]
    ranges: list[CellRange] = []
    seen: set[tuple[int, int, int, int]] = set()

    def matches(value: object) -> bool:
        return any(_header_matches(value, needle) for needle in needles)

    for header_rng in ws.merged_cells.ranges:
        if header_rng.min_row != header_rng.max_row:
            continue
        if not matches(ws.cell(header_rng.min_row, header_rng.min_col).value):
            continue
        _add_unique_range(
            ranges,
            seen,
            _picture_range_below(ws, header_rng, rows=rows),
        )

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if _cell_in_merged_range(ws, row, col):
                continue
            if not matches(ws.cell(row, col).value):
                continue
            _add_unique_range(
                ranges,
                seen,
                _picture_range_below(
                    ws,
                    CellRange(min_col=col, max_col=col, min_row=row, max_row=row),
                    rows=rows,
                ),
            )
    return ranges


def _all_dpt_overview_ranges(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> list[CellRange]:
    ranges: list[CellRange] = []
    seen: set[tuple[int, int, int, int]] = set()
    labels = _left_merged_ranges_with_value(ws)
    by_row = {rng.min_row: value for rng, value in labels}
    for phase_rng, phase_label in labels:
        if _phase_temp_parts(phase_label, temperature_labels) is None:
            continue
        condition_label = by_row.get(phase_rng.min_row + 17)
        if condition_label is None:
            continue
        cv, ci = _condition_values(condition_label)
        if cv is None or ci is None:
            continue
        _add_unique_range(
            ranges,
            seen,
            _left_merged_range_at(ws, phase_rng.min_row + 34),
        )
    return ranges


def _image_anchor_position(image) -> tuple[int, int] | None:
    anchor = image.anchor
    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        return row, col
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return int(marker.row) + 1, int(marker.col) + 1


def _remove_images_in_range(ws: Worksheet, rng: CellRange) -> None:
    kept = []
    for image in getattr(ws, "_images", []):
        pos = _image_anchor_position(image)
        if pos is None:
            kept.append(image)
            continue
        row, col = pos
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            continue
        kept.append(image)
    ws._images = kept


def _condition_matches_setpoints(
    condition_label: str | None,
    setpoints: _DptSetpointMatch,
    condition_signature: Mapping[str, float] | None = None,
) -> bool:
    cv, ci = _condition_values(condition_label)
    if not _setpoint_value_matches(
        setpoints.vdc,
        cv,
        expected_from_filename=setpoints.vdc_from_filename,
    ) or not _setpoint_value_matches(
        setpoints.idc,
        ci,
        expected_from_filename=setpoints.idc_from_filename,
    ):
        return False
    signature = condition_signature or {}
    if signature and not _condition_signature_matches(
        signature,
        _condition_signature(condition_label),
    ):
        return False
    return True


def _clear_dpt_waveform_block_payload(ws: Worksheet, anchor_row: int) -> None:
    block_range = CellRange(
        min_col=1,
        max_col=max(1, ws.max_column),
        min_row=anchor_row,
        max_row=anchor_row + DPT_WAVEFORM_BLOCK_ROWS - 1,
    )
    _remove_images_in_range(ws, block_range)
    for offset in (0, 17, 34):
        _set_range_value(ws, _left_merged_range_at(ws, anchor_row + offset), None)


def _clear_duplicate_dpt_waveform_blocks(
    ws: Worksheet,
    data_ws: Worksheet,
    target_anchor_row: int | None,
    phase_code: str,
    temp_code: str,
    setpoints: _DptSetpointMatch,
    condition_signature: Mapping[str, float] | None = None,
    temperature_labels: TemperatureLabels | None = None,
    protected_anchor_rows: set[int] | None = None,
) -> None:
    if target_anchor_row is None:
        return
    labels = _left_merged_ranges_with_value(ws)
    label_by_row = {rng.min_row: value for rng, value in labels}
    for phase_rng, phase_label in labels:
        anchor_row = phase_rng.min_row
        if anchor_row == target_anchor_row or anchor_row in (protected_anchor_rows or set()):
            continue
        if not _phase_temperature_matches_display(
            phase_label,
            phase_code,
            temp_code,
            temperature_labels,
        ):
            continue
        condition_label = label_by_row.get(anchor_row + 17)
        existing_setpoints = _condition_setpoint_match(condition_label)
        if existing_setpoints.vdc is None or existing_setpoints.idc is None:
            # Cleanup is destructive. If a legacy/custom label cannot prove
            # which V/I condition this block belongs to, preserve it instead
            # of treating a parse failure as evidence that the image is stale.
            continue
        if not _condition_matches_setpoints(
            condition_label,
            setpoints,
            condition_signature,
        ):
            if not _condition_matches_setpoints(
                condition_label,
                setpoints,
            ) and _dpt_waveform_block_has_data_match(
                data_ws,
                str(phase_label),
                str(condition_label or ""),
                temperature_labels,
                # A merged group condition is updated before duplicate cleanup.
                # Preserve another V/I block whose numeric data row still exists,
                # even when that earlier block carries a different Rg/Cg signature.
                # The fallback must not preserve the same V/I with a stale signature.
                require_condition_signature=False,
            ):
                continue
        _clear_dpt_waveform_block_payload(ws, anchor_row)


def _marker_after_pixels(
    ws: Worksheet,
    *,
    start_row: int,
    start_col: int,
    width_px: int,
    height_px: int,
) -> AnchorMarker:
    col = start_col
    remaining_w = max(0, int(width_px))
    while remaining_w > 0:
        col_width = _column_width_px(ws, col)
        if remaining_w < col_width:
            break
        remaining_w -= col_width
        col += 1

    row = start_row
    remaining_h = max(0, int(height_px))
    while remaining_h > 0:
        row_height = _row_height_px(ws, row)
        if remaining_h < row_height:
            break
        remaining_h -= row_height
        row += 1

    return AnchorMarker(
        col=max(0, col - 1),
        colOff=pixels_to_EMU(remaining_w),
        row=max(0, row - 1),
        rowOff=pixels_to_EMU(remaining_h),
    )


def _image_two_cell_anchor(
    ws: Worksheet,
    rng: CellRange,
    width_px: int,
    height_px: int,
) -> TwoCellAnchor:
    start = AnchorMarker(
        col=max(0, rng.min_col - 1),
        row=max(0, rng.min_row - 1),
    )
    end = _marker_after_pixels(
        ws,
        start_row=rng.min_row,
        start_col=rng.min_col,
        width_px=width_px,
        height_px=height_px,
    )
    return TwoCellAnchor(editAs="oneCell", _from=start, to=end)


def _insert_image(
    ws: Worksheet,
    image_path: str | Path,
    rng: CellRange,
    *,
    shrink_columns: bool = True,
    display_size: tuple[int, int] | None = None,
) -> None:
    path = Path(image_path)
    if not path.is_file():
        raise FileNotFoundError(f"报告图片不存在或不是文件：{path}")
    _remove_images_in_range(ws, rng)
    ws.cell(rng.min_row, rng.min_col).value = None
    image = XLImage(str(path))
    if shrink_columns:
        _fit_columns_to_image_aspect(ws, rng, image)
    max_w, max_h = _range_size_px(ws, rng)
    if image.width and image.height:
        # Keep the captured waveform's original aspect ratio. Only the sheet
        # columns are adapted; the image is never stretched independently.
        if display_size is not None:
            target_w, target_h = display_size
            target_aspect = target_w / max(1, target_h)
            image_aspect = image.width / image.height
            if abs(target_aspect - image_aspect) < 0.02:
                scale = min(max_w / target_w, max_h / target_h, 1.0)
                image.width = max(1, int(target_w * scale))
                image.height = max(1, int(target_h * scale))
            else:
                scale = min(target_w / image.width, target_h / image.height)
                image.width = max(1, int(image.width * scale))
                image.height = max(1, int(image.height * scale))
            image.anchor = _image_two_cell_anchor(ws, rng, image.width, image.height)
            ws.add_image(image)
            return
        scale = min(max_w / image.width, max_h / image.height)
        image.width = max(1, int(image.width * scale))
        image.height = max(1, int(image.height * scale))
    image.anchor = _image_two_cell_anchor(ws, rng, image.width, image.height)
    ws.add_image(image)


def _write_dpt_images(
    ws: Worksheet,
    anchor_row: int | None,
    images: ImageMap,
    result: ExtractResult,
    progress_callback: Callable[[int, int, str], None] | None = None,
    temperature_labels: TemperatureLabels | None = None,
) -> int:
    if anchor_row is None:
        _normalize_dpt_waveform_text(ws, temperature_labels)
        return 0
    allowed_params = set(dpt_report_image_params_for_result(result))
    block_rows = {
        "off": anchor_row,
        "on": anchor_row + 17,
        "rr": anchor_row + 34,
    }
    # A report write replaces the complete image state of the active
    # condition.  Clear every slot in that condition first so a missing or
    # unavailable metric cannot leave a picture from an earlier write.
    overview_target = _left_merged_range_at(ws, anchor_row + 34)
    _remove_images_in_range(ws, overview_target)
    parameter_targets: dict[tuple[str, str], CellRange] = {}
    for key, (block, header_text) in _DPT_IMAGE_HEADERS.items():
        header = _find_header_range(ws, block_rows[block], header_text)
        if header is None:
            continue
        target = _picture_range_below(ws, header, rows=16)
        parameter_targets[key] = target
        _remove_images_in_range(ws, target)
    written = 0
    overview_items: list[tuple[str | Path, CellRange]] = []
    overview_path = (
        images.get(DPT_OVERVIEW_IMAGE_PARAM)
        if DPT_OVERVIEW_IMAGE_PARAM in allowed_params
        else None
    )
    if overview_path is not None:
        overview_items.append((overview_path, overview_target))
    regular_items: list[tuple[str | Path, CellRange]] = []
    for key, image_path in images.items():
        if key == DPT_OVERVIEW_IMAGE_PARAM:
            continue
        if key not in allowed_params:
            continue
        target = parameter_targets.get(key)
        if target is None:
            continue
        regular_items.append((image_path, target))
    overview_display_size = _prepare_uniform_image_slots(
        ws,
        overview_items,
        template_ranges=_all_dpt_overview_ranges(ws, temperature_labels),
    )
    all_parameter_slots = _all_picture_ranges_by_headers(
        ws,
        {header_text for _block, header_text in _DPT_IMAGE_HEADERS.values()},
        rows=16,
    )
    display_size = _prepare_uniform_image_slots(
        ws,
        regular_items,
        template_ranges=all_parameter_slots,
    )
    _normalize_dpt_waveform_text(ws, temperature_labels)
    image_total = len(overview_items) + len(regular_items)
    for image_path, target in overview_items:
        _insert_image(
            ws,
            image_path,
            target,
            shrink_columns=False,
            display_size=overview_display_size,
        )
        written += 1
        if progress_callback is not None:
            progress_callback(written, image_total, "插入报告图片")
    for image_path, target in regular_items:
        _insert_image(
            ws,
            image_path,
            target,
            shrink_columns=False,
            display_size=display_size,
        )
        written += 1
        if progress_callback is not None:
            progress_callback(written, image_total, "插入报告图片")
    return written


def _short_target_row(
    ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> int:
    for row in range(5, ws.max_row + 1):
        phase = str(ws.cell(row, 2).value or "").upper()
        if phase != phase_code.upper():
            continue
        row_temp = _merged_value(ws, row, 1)
        if _temperature_cell_matches_display(
            row_temp,
            temp_code,
            temperature_labels,
        ):
            _set_merged_cell_value(
                ws,
                row,
                1,
                _temperature_display(temp_code, temperature_labels),
            )
            return row
    display_temp = _temperature_display(temp_code, temperature_labels)
    raise ValueError(
        f"报告模板工作表“{ws.title}”缺少短路工况行："
        f"{phase_code.upper()} / {display_temp}；为避免写入错误行，已停止写入"
    )


def _short_temperature_groups(ws: Worksheet) -> list[_ShortTemperatureGroup]:
    groups: list[_ShortTemperatureGroup] = []
    seen: set[tuple[int, int]] = set()
    for row in range(5, ws.max_row + 1):
        phase = str(ws.cell(row, 2).value or "").strip().upper()
        if phase not in PHASE_CODES:
            continue
        temperature_range = _merged_range_containing(ws, row, 1)
        start_row = temperature_range.min_row if temperature_range is not None else row
        end_row = temperature_range.max_row if temperature_range is not None else row
        key = (start_row, end_row)
        if key in seen:
            continue
        seen.add(key)
        groups.append(
            _ShortTemperatureGroup(
                start_row=start_row,
                end_row=end_row,
                phase_family=phase[0],
                temp=_merged_value(ws, row, 1),
            )
        )
    return sorted(groups, key=lambda group: group.start_row)


def _normalize_short_data_temperature_order(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    groups = _short_temperature_groups(ws)
    blocks: list[_WorksheetRowBlock] = []
    for index, group in enumerate(groups):
        end_row = (
            groups[index + 1].start_row - 1
            if index + 1 < len(groups)
            else group.end_row
        )
        number = _require_temperature_order_number(
            ws,
            group.temp,
            temperature_labels,
        )
        blocks.append(
            _WorksheetRowBlock(
                start_row=group.start_row,
                end_row=end_row,
                sort_key=(
                    _phase_family_report_order(group.phase_family),
                    _temperature_report_order_key(number),
                ),
            )
        )
    return _reorder_worksheet_row_blocks(ws, blocks)


def _normalize_short_picture_temperature_order(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> bool:
    labels = [
        (rng, parts)
        for rng, value in _merged_ranges_with_value(ws, min_col=1, max_col=5)
        if (parts := _phase_temp_token_parts(value)) is not None
        and parts[0] in PHASE_CODES
    ]
    blocks: list[_WorksheetRowBlock] = []
    for index, (rng, parts) in enumerate(labels):
        number = _require_temperature_order_number(
            ws,
            parts[1],
            temperature_labels,
        )
        end_row = (
            labels[index + 1][0].min_row - 1
            if index + 1 < len(labels)
            else ws.max_row
        )
        blocks.append(
            _WorksheetRowBlock(
                start_row=rng.min_row,
                end_row=end_row,
                sort_key=(
                    _phase_family_report_order(parts[0][0]),
                    _temperature_report_order_key(number),
                    _phase_report_order(parts[0]),
                ),
            )
        )
    return _reorder_worksheet_row_blocks(ws, blocks)


def _short_temperature_group_insert_row(
    ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
    *,
    fallback_row: int,
) -> int:
    temperature = _temperature_number_for_report_order(
        _temperature_display(temp_code, temperature_labels),
        temperature_labels,
    )
    if temperature is None:
        return fallback_row
    target_key = _temperature_report_order_key(temperature)
    phase_family = phase_code.upper()[0]
    last_family_end: int | None = None
    for group in _short_temperature_groups(ws):
        if group.phase_family != phase_family:
            continue
        group_temperature = _temperature_number_for_report_order(
            group.temp,
            temperature_labels,
        )
        if group_temperature is None:
            continue
        if _temperature_report_order_key(group_temperature) > target_key:
            return group.start_row
        last_family_end = group.end_row
    return last_family_end + 1 if last_family_end is not None else fallback_row


def _ensure_short_temperature_group(
    ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    for row in range(5, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").upper() != phase_code.upper():
            continue
        if _temperature_cell_matches_display(
            _merged_value(ws, row, 1),
            temp_code,
            temperature_labels,
        ):
            return

    seed_row = None
    for row in range(5, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").upper() != phase_code.upper():
            continue
        if _temperature_cell_matches(
            _merged_value(ws, row, 1),
            temp_code,
            temperature_labels,
        ):
            seed_row = row
    if seed_row is None:
        return

    temperature_range = _merged_range_containing(ws, seed_row, 1)
    source_start = temperature_range.min_row if temperature_range is not None else seed_row
    source_end = temperature_range.max_row if temperature_range is not None else seed_row
    row_count = source_end - source_start + 1
    insert_row = _short_temperature_group_insert_row(
        ws,
        phase_code,
        temp_code,
        temperature_labels,
        fallback_row=source_end + 1,
    )
    merged_ranges = [CellRange(str(rng)) for rng in ws.merged_cells.ranges]
    cloned_ranges = [
        CellRange(
            min_col=rng.min_col,
            max_col=rng.max_col,
            min_row=insert_row + (rng.min_row - source_start),
            max_row=insert_row + (rng.max_row - source_start),
        )
        for rng in merged_ranges
        if source_start <= rng.min_row and rng.max_row <= source_end
    ]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.insert_rows(insert_row, row_count)
    _shift_row_dimensions_for_insert(ws, insert_row, row_count)
    copied_source_start = (
        source_start + row_count if insert_row <= source_start else source_start
    )
    for offset in range(row_count):
        source_row = copied_source_start + offset
        target_row = insert_row + offset
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            target.value = copy(source.value) if col <= 3 else None
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
                target.number_format = source.number_format
    for rng in merged_ranges:
        shifted = (
            CellRange(
                min_col=rng.min_col,
                max_col=rng.max_col,
                min_row=rng.min_row + row_count,
                max_row=rng.max_row + row_count,
            )
            if rng.min_row >= insert_row
            else rng
        )
        ws.merge_cells(str(shifted))
    for rng in cloned_ranges:
        ws.merge_cells(str(rng))
    _set_merged_cell_value(
        ws,
        insert_row,
        1,
        _temperature_display(temp_code, temperature_labels),
    )


def _normalize_short_temperature_cells(
    ws: Worksheet,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    for row in range(5, ws.max_row + 1):
        raw = str(_merged_value(ws, row, 1) or "").strip().upper()
        if raw in TEMP_LABELS:
            _set_merged_cell_value(
                ws,
                row,
                1,
                _temperature_display(raw, temperature_labels),
            )


def _normalize_report_temperature_labels(
    wb,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    for ws in wb.worksheets:
        title = str(ws.title)
        if title.endswith("_双脉冲数据") or title.endswith("相_双脉冲数据"):
            _normalize_dpt_data_temperature_cells(ws, temperature_labels)
        elif title.endswith("_双脉冲波形") or title.endswith("相_双脉冲波形"):
            _normalize_dpt_waveform_text(ws, temperature_labels)
        elif title == "短路测试":
            _normalize_short_temperature_cells(ws, temperature_labels)
        elif title == "短路测试图片":
            _normalize_short_picture_text(ws, temperature_labels)


def _write_short_data(ws: Worksheet, row: int, result: ExtractResult) -> None:
    sc = result.short_circuit
    vdc = _short_voltage_from_filename(result.source_path)
    if vdc is None:
        vdc, _idc = _match_setpoints(result)
    _set_value(ws, row, 4, _num(vdc, 1))
    _set_metric_value(ws, row, 5, result, "短路过程", "短路电流Imax", _num(sc.ic_max, 3))
    _set_metric_value(ws, row, 6, result, "短路过程", "短路时间Tsc", _num(sc.tsc, 4))
    _set_metric_value(ws, row, 7, result, "短路过程", "短路能量Esc_本管", _num(sc.esc_dut, 4))
    _set_metric_value(ws, row, 8, result, "短路过程", "应力Vpeak_本管", _num(sc.vpeak_dut, 3))
    _set_metric_value(ws, row, 9, result, "短路过程", "短路能量Esc_对管", _num(sc.esc_other, 4))
    _set_metric_value(ws, row, 10, result, "短路过程", "应力Vpeak_对管", _num(sc.vpeak_other, 3))
    _set_metric_value(ws, row, 11, result, "短路过程", "Desat动作时间", _num(sc.desat_time, 4))


def _short_voltage_from_filename(path: str) -> float | None:
    vdc, _idc = parse_setpoints_from_filename(path)
    if vdc is not None:
        return vdc
    match = re.search(
        r"(?<![A-Z0-9])(\d+(?:\.\d+)?)V(?![A-Z])",
        Path(path).stem.upper(),
    )
    if match is None:
        return None
    return float(match.group(1))


def _format_measurement_with_unit(
    value: float | None,
    unit: str,
    *,
    digits: int,
) -> str | None:
    number = _num(value, digits)
    if number is None:
        return None
    text = f"{number:.{digits}f}".rstrip("0").rstrip(".")
    return f"{text}{unit}"


def _write_short_picture_conditions(
    ws: Worksheet,
    anchor_row: int,
    result: ExtractResult,
) -> None:
    values = {
        "VCE": _format_measurement_with_unit(
            _short_voltage_from_filename(result.source_path),
            "V",
            digits=1,
        ),
        "IMAX": _format_measurement_with_unit(
            result.short_circuit.ic_max,
            "A",
            digits=3,
        ),
    }
    if all(value is None for value in values.values()):
        return

    max_row = min(ws.max_row, anchor_row + 10)
    for row in range(anchor_row, max_row + 1):
        for col in range(1, 6):
            label = _normalized(ws.cell(row, col).value)
            key: str | None = None
            if label.startswith("VCE"):
                key = "VCE"
            elif label.startswith("IMAX"):
                key = "IMAX"
            if key is None:
                continue
            _set_value(ws, row + 1, col, values[key])


def _short_image_anchor_row(
    ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> int | None:
    for rng, value in _merged_ranges_with_value(ws, min_col=1, max_col=5):
        if _phase_temperature_matches_display(
            value,
            phase_code,
            temp_code,
            temperature_labels,
        ):
            _set_range_value(
                ws,
                rng,
                _phase_temp_label(phase_code, temp_code, temperature_labels),
            )
            return rng.min_row
    return None


def _ensure_short_picture_temperature_block(
    ws: Worksheet,
    phase_code: str,
    temp_code: str,
    temperature_labels: TemperatureLabels | None = None,
) -> None:
    labels = _merged_ranges_with_value(ws, min_col=1, max_col=5)
    if any(
        _phase_temperature_matches_display(
            value,
            phase_code,
            temp_code,
            temperature_labels,
        )
        for _rng, value in labels
    ):
        return
    phase_labels = [
        (rng, value, parts)
        for rng, value in labels
        if (parts := _phase_temp_token_parts(value)) is not None
        and parts[0] in PHASE_CODES
    ]
    seeds = [
        (index, rng, value)
        for index, (rng, value, _parts) in enumerate(phase_labels)
        if _phase_temp_parts(value, temperature_labels)
        == (phase_code.upper(), temp_code)
    ]
    if not seeds:
        return
    seed_index, seed_range, _seed_value = seeds[-1]
    source_start = seed_range.min_row
    source_next_row = (
        phase_labels[seed_index + 1][0].min_row
        if seed_index + 1 < len(phase_labels)
        else ws.max_row + 1
    )
    row_count = max(1, source_next_row - source_start)
    temperature = _temperature_number_for_report_order(
        _temperature_display(temp_code, temperature_labels),
        temperature_labels,
    )
    insert_row = source_next_row
    if temperature is not None:
        target_key = (
            _temperature_report_order_key(temperature),
            _phase_report_order(phase_code),
        )
        phase_family = phase_code.upper()[0]
        last_family_index: int | None = None
        for index, (rng, value, parts) in enumerate(phase_labels):
            if parts[0][0] != phase_family:
                continue
            block_temperature = _temperature_number_for_report_order(
                parts[1],
                temperature_labels,
            )
            if block_temperature is None:
                continue
            block_key = (
                _temperature_report_order_key(block_temperature),
                _phase_report_order(parts[0]),
            )
            if block_key > target_key:
                insert_row = rng.min_row
                break
            last_family_index = index
        else:
            if last_family_index is not None:
                insert_row = (
                    phase_labels[last_family_index + 1][0].min_row
                    if last_family_index + 1 < len(phase_labels)
                    else ws.max_row + 1
                )
    merged_ranges = [CellRange(str(rng)) for rng in ws.merged_cells.ranges]
    cloned_ranges = [
        CellRange(
            min_col=rng.min_col,
            max_col=rng.max_col,
            min_row=insert_row + (rng.min_row - source_start),
            max_row=insert_row + (rng.max_row - source_start),
        )
        for rng in merged_ranges
        if source_start <= rng.min_row and rng.max_row < source_start + row_count
    ]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.insert_rows(insert_row, row_count)
    _shift_row_dimensions_for_insert(ws, insert_row, row_count)
    _shift_image_anchors_for_insert(ws, insert_row, row_count)
    copied_source_start = (
        source_start + row_count if insert_row <= source_start else source_start
    )
    for offset in range(row_count):
        source_row = copied_source_start + offset
        target_row = insert_row + offset
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            target.value = copy(source.value)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
                target.number_format = source.number_format
    for rng in merged_ranges:
        shifted = (
            CellRange(
                min_col=rng.min_col,
                max_col=rng.max_col,
                min_row=rng.min_row + row_count,
                max_row=rng.max_row + row_count,
            )
            if rng.min_row >= insert_row
            else rng
        )
        ws.merge_cells(str(shifted))
    for rng in cloned_ranges:
        ws.merge_cells(str(rng))

    cloned_label = CellRange(
        min_col=seed_range.min_col,
        max_col=seed_range.max_col,
        min_row=insert_row,
        max_row=insert_row + (seed_range.max_row - seed_range.min_row),
    )
    _set_range_value(
        ws,
        cloned_label,
        _phase_temp_label(phase_code, temp_code, temperature_labels),
    )
    for row in range(insert_row, min(ws.max_row, insert_row + row_count - 1) + 1):
        for col in range(1, 6):
            label = _normalized(ws.cell(row, col).value)
            if (label.startswith("VCE") or label.startswith("IMAX")) and row < ws.max_row:
                ws.cell(row + 1, col).value = None


def _write_short_images(
    ws: Worksheet,
    anchor_row: int | None,
    images: ImageMap,
    result: ExtractResult,
    progress_callback: Callable[[int, int, str], None] | None = None,
    temperature_labels: TemperatureLabels | None = None,
) -> int:
    if anchor_row is None:
        _normalize_short_picture_text(ws, temperature_labels)
        return 0
    _write_short_picture_conditions(ws, anchor_row, result)
    # As with DPT reports, the active short-circuit condition owns a complete
    # set of slots.  Remove its prior pictures before filtering this run's
    # unavailable or missing images; other condition blocks remain untouched.
    parameter_targets: dict[tuple[str, str], CellRange] = {}
    for key, header_text in _SHORT_IMAGE_HEADERS.items():
        header = _find_header_range(ws, anchor_row, header_text)
        if header is None:
            continue
        target = _picture_range_below(ws, header, rows=9)
        parameter_targets[key] = target
        _remove_images_in_range(ws, target)
    written = 0
    items: list[tuple[str | Path, CellRange]] = []
    for key, image_path in images.items():
        if result.is_metric_unavailable(*key):
            continue
        target = parameter_targets.get(key)
        if target is None:
            continue
        items.append((image_path, target))
    all_parameter_slots = _all_picture_ranges_by_headers(
        ws,
        _SHORT_IMAGE_HEADERS.values(),
        rows=9,
    )
    display_size = _prepare_uniform_image_slots(
        ws,
        items,
        template_ranges=all_parameter_slots,
    )
    _normalize_short_picture_text(ws, temperature_labels)
    image_total = len(items)
    for image_path, target in items:
        _insert_image(
            ws,
            image_path,
            target,
            shrink_columns=False,
            display_size=display_size,
        )
        written += 1
        if progress_callback is not None:
            progress_callback(written, image_total, "插入报告图片")
    return written


def _as_report_results(result: ExtractResult | Sequence[ExtractResult]) -> list[ExtractResult]:
    if isinstance(result, ExtractResult):
        return [result]
    rows = list(result)
    if not rows:
        raise ValueError("没有可写入报告的结果")
    first_mode = rows[0].short_circuit_mode
    if any(row.short_circuit_mode != first_mode for row in rows):
        raise ValueError("不能混合写入双脉冲和短路测试结果")
    return rows


def _expected_dpt_image_keys(
    images: ImageMap,
    result: ExtractResult,
) -> tuple[tuple[str, str], ...]:
    allowed = set(dpt_report_image_params_for_result(result))
    return tuple(key for key in images if key in allowed)


def _expected_short_image_keys(
    images: ImageMap,
    result: ExtractResult,
) -> tuple[tuple[str, str], ...]:
    return tuple(
        key
        for key in images
        if key in _SHORT_IMAGE_HEADERS and not result.is_metric_unavailable(*key)
    )


def _require_complete_report_images(
    report_kind: str,
    expected_keys: Sequence[tuple[str, str]],
    images_written: int,
) -> None:
    expected_count = len(expected_keys)
    if images_written == expected_count:
        return
    expected_text = "、".join(f"{section}/{name}" for section, name in expected_keys)
    raise ValueError(
        f"{report_kind}报告图片写入不完整：应写入 {expected_count} 张，"
        f"实际写入 {images_written} 张。请检查模板中的工况图片块和参数标题。"
        f"待写图片：{expected_text}"
    )


def write_report_template(
    result: ExtractResult | Sequence[ExtractResult],
    report_path: str | Path,
    *,
    images: ImageMap | None = None,
    target_screen_width_px: int | None = None,
    progress_callback: ReportProgressCallback | None = None,
    temperature_labels: TemperatureLabels | None = None,
    temperature_code: str | None = None,
    phase_code: str | None = None,
    image_result_index: int | None = None,
) -> ReportWriteSummary:
    path = Path(report_path)
    results = _as_report_results(result)
    result0 = results[0]
    active_index = 0 if image_result_index is None else int(image_result_index)
    if active_index < 0 or active_index >= len(results):
        raise ValueError(f"报告截图结果索引越界：{active_index}")
    image_result = results[active_index]
    progress_total = max(6, 6 + len(images or {}))

    def emit(step: int, label: str) -> None:
        if progress_callback is not None:
            progress_callback(
                max(0, min(int(step), progress_total)),
                progress_total,
                label,
            )

    def emit_image(done: int, total: int, label: str) -> None:
        if progress_callback is not None:
            progress_callback(
                max(0, min(int(done), max(0, int(total)))),
                max(0, int(total)),
                label,
            )

    emit(0, "打开报告文件")
    wb = load_workbook(path)
    emit(1, "读取报告模板")
    temperature_labels = _temperature_identity_labels_for_workbook(
        wb,
        temperature_labels,
    )
    image_map: ImageMap = images or {}
    phase_code = _resolve_phase_code(result0.source_path, result0, phase_code)
    temp_code = _resolve_temp_code(result0.source_path, temperature_code)

    if result0.short_circuit_mode:
        if len(results) != 1:
            raise ValueError("短路测试报告不支持多行结果")
        result = result0
        data_sheet = "短路测试"
        picture_sheet = "短路测试图片"
        if data_sheet not in wb.sheetnames:
            raise ValueError("报告模板缺少工作表：短路测试")
        data_ws = wb[data_sheet]
        _normalize_short_data_temperature_order(data_ws, temperature_labels)
        if picture_sheet in wb.sheetnames:
            _normalize_short_picture_temperature_order(
                wb[picture_sheet],
                temperature_labels,
            )
        _ensure_short_temperature_group(
            data_ws,
            phase_code,
            temp_code,
            temperature_labels,
        )
        data_row = _short_target_row(
            data_ws,
            phase_code,
            temp_code,
            temperature_labels,
        )
        _write_short_data(data_ws, data_row, result)
        _normalize_short_temperature_cells(data_ws, temperature_labels)
        emit(2, "写入报告数据")
        expected_image_keys = _expected_short_image_keys(image_map, result)
        if expected_image_keys and picture_sheet not in wb.sheetnames:
            raise ValueError(f"报告模板缺少工作表：{picture_sheet}，无法写入报告图片")
        images_written = 0
        anchor_row = None
        if picture_sheet in wb.sheetnames:
            image_ws = wb[picture_sheet]
            _ensure_short_picture_temperature_block(
                image_ws,
                phase_code,
                temp_code,
                temperature_labels,
            )
            anchor_row = _short_image_anchor_row(
                image_ws,
                phase_code,
                temp_code,
                temperature_labels,
            )
            if expected_image_keys and anchor_row is None:
                raise ValueError(
                    f"报告模板工作表“{picture_sheet}”缺少图片工况块："
                    f"{_phase_temp_label(phase_code, temp_code, temperature_labels)}"
                )
            images_written = _write_short_images(
                image_ws,
                anchor_row,
                image_map,
                result,
                progress_callback=emit_image,
                temperature_labels=temperature_labels,
            )
        _require_complete_report_images(
            "短路",
            expected_image_keys,
            images_written,
        )
        emit(progress_total - 2, "整理报告版式")
        _normalize_report_temperature_labels(wb, temperature_labels)
        _set_report_open_zoom(wb, target_screen_width_px)
        _write_report_temperature_identities(wb, temperature_labels)
        emit(progress_total - 1, "保存报告文件")
        wb.save(path)
        return ReportWriteSummary(
            report_path=path,
            data_sheet=data_sheet,
            data_row=data_row,
            waveform_sheet=picture_sheet if picture_sheet in wb.sheetnames else None,
            waveform_anchor_row=anchor_row,
            images_written=images_written,
        )

    sheet_prefix = _phase_sheet_prefix(phase_code)
    data_sheet = f"{sheet_prefix}相_双脉冲数据"
    waveform_sheet = f"{sheet_prefix}相_双脉冲波形"
    if data_sheet not in wb.sheetnames:
        raise ValueError(f"报告模板缺少工作表：{data_sheet}")
    data_ws = wb[data_sheet]
    _ensure_dpt_data_power_columns(data_ws)
    expected_image_keys = _expected_dpt_image_keys(image_map, image_result)
    if expected_image_keys and waveform_sheet not in wb.sheetnames:
        raise ValueError(f"报告模板缺少工作表：{waveform_sheet}，无法写入报告图片")
    waveform_ws = wb[waveform_sheet] if waveform_sheet in wb.sheetnames else None
    _normalize_dpt_data_temperature_order(data_ws, temperature_labels)
    if waveform_ws is not None:
        _normalize_dpt_waveform_temperature_order(
            waveform_ws,
            temperature_labels,
        )
    _ensure_dpt_temperature_group(
        data_ws,
        phase_code,
        temp_code,
        temperature_labels,
    )
    target = _find_dpt_data_target(
        result0,
        data_ws,
        phase_code,
        temp_code,
        temperature_labels,
    )
    condition_signature = _condition_signature(image_result.source_path)
    _sync_dpt_condition_cell(data_ws, target.group_start_row, condition_signature)
    data_targets = _prepare_dpt_data_rows(data_ws, target, results)
    data_rows = [item.row for item in data_targets]
    data_row = data_rows[0]
    for row, row_result in zip(data_rows, results):
        _write_dpt_data(data_ws, row, row_result)
    _normalize_dpt_data_temperature_cells(data_ws, temperature_labels)
    emit(2, "写入报告数据")
    if waveform_ws is not None:
        _sync_dpt_waveform_insertions(waveform_ws, data_ws, data_targets)
    setpoints = _dpt_setpoint_match(image_result)
    waveform_anchor_rows: list[int] = []
    if waveform_ws is not None:
        for row_target, row_result in zip(data_targets, results):
            stable_target = _DptDataTarget(
                row=row_target.row,
                group_index=row_target.group_index,
                group_start_row=row_target.group_start_row,
                row_offset=row_target.row_offset,
                inserted_row=False,
            )
            row_vdc, row_idc = _match_setpoints(row_result)
            waveform_anchor_rows.append(
                _ensure_dpt_waveform_anchor(
                    waveform_ws,
                    data_ws,
                    stable_target,
                    phase_code,
                    temp_code,
                    row_vdc,
                    row_idc,
                    _condition_signature(row_result.source_path),
                    temperature_labels,
                )
            )
    waveform_anchor_row = (
        waveform_anchor_rows[active_index]
        if waveform_anchor_rows
        else None
    )
    if expected_image_keys and waveform_anchor_row is None:
        raise ValueError(
            f"报告模板工作表“{waveform_sheet}”缺少图片工况块："
            f"{_phase_temp_label(phase_code, temp_code, temperature_labels)}"
        )
    if waveform_ws is not None:
        _clear_duplicate_dpt_waveform_blocks(
            waveform_ws,
            data_ws,
            waveform_anchor_row,
            phase_code,
            temp_code,
            setpoints,
            condition_signature,
            temperature_labels,
            protected_anchor_rows=set(waveform_anchor_rows),
        )
    images_written = (
        _write_dpt_images(
            waveform_ws,
            waveform_anchor_row,
            image_map,
            image_result,
            progress_callback=emit_image,
            temperature_labels=temperature_labels,
        )
        if waveform_ws is not None
        else 0
    )
    _require_complete_report_images(
        "双脉冲",
        expected_image_keys,
        images_written,
    )
    emit(progress_total - 2, "整理报告版式")
    _normalize_report_temperature_labels(wb, temperature_labels)
    _set_report_open_zoom(wb, target_screen_width_px)
    _write_report_temperature_identities(wb, temperature_labels)
    emit(progress_total - 1, "保存报告文件")
    wb.save(path)
    return ReportWriteSummary(
        report_path=path,
        data_sheet=data_sheet,
        data_row=data_row,
        data_row_end=data_rows[-1],
        data_rows_written=len(data_rows),
        waveform_sheet=waveform_sheet if waveform_ws is not None else None,
        waveform_anchor_row=waveform_anchor_row,
        images_written=images_written,
    )
